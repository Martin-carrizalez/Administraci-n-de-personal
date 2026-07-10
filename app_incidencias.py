import streamlit as st
import os
# Import protegido: si nomina_module falta o truena, la app sigue funcionando
# (solo se deshabilita esa pestaña). Un módulo secundario no debe tumbar el login.
try:
    from nomina_module import render_pendientes_nomina
except Exception as _e_nomina:
    render_pendientes_nomina = None
    _ERROR_NOMINA = str(_e_nomina)
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from datetime import datetime, date, timedelta
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import qrcode
from datetime import timezone
import pytz
from reportlab.platypus import Image as RLImage

# ─────────────────────────────────────────────
# FLAGS DE FUNCIONALIDAD
# ─────────────────────────────────────────────
HABILITAR_CUMPLEANOS = False  # Cambiar a True cuando Ángel lo autorice

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

TIPO_LABELS = {
    "ECO": "Día económico",
    "PSE": "Pase de salida sin retorno",
    "PSR": "Pase de salida con retorno",
    "PEN": "Pase de entrada",
    "COM": "Comisión",
    "CHO": "Cambio de horario",
    "RGU": "Reposición de guardias",
}
if HABILITAR_CUMPLEANOS:
    TIPO_LABELS["CUM"] = "Día de cumpleaños"

DIAS_SEMANA = ["LUN", "MAR", "MIE", "JUE", "VIE"]

def formato_jefe_pdf(jefe_raw: str) -> str:
    """Convierte el valor de la columna JEFE_INMEDIATO del Sheet Usuarios
    (ej. 'Maricela Esquivel Domínguez — Dir. Desarrollo Académico')
    al formato del PDF con salto de línea entre nombre y cargo.
    El catálogo de jefes vive en el Sheet, no en el código."""
    jefe_raw = str(jefe_raw or "").strip()
    if not jefe_raw:
        return ""
    # Separa nombre y cargo por el guion largo (—) o el guion normal (-) con espacios
    for sep in (" — ", " – ", " - "):
        if sep in jefe_raw:
            nombre, cargo = jefe_raw.split(sep, 1)
            return f"{nombre.strip()}<br/>{cargo.strip()}"
    return jefe_raw

DIRECTORA_NOMBRE = "Claudia Gisela Ramírez Monroy"
DIRECTORA_CARGO  = "Encargada del Despacho de la Dirección de Formación Continua"
NOMBRE_DIRECTORA = f"{DIRECTORA_NOMBRE}<br/>{DIRECTORA_CARGO}"
DRIVE_ANEXOS_FOLDER = "1LnQjrhjEKgKxFTJD8USLoiCpCKQHfOQC"

COLUMNAS_HORARIO = {
    "LUN": ("ENTRADA_LUN", "SALIDA_LUN"),
    "MAR": ("ENTRADA_MAR", "SALIDA_MAR"),
    "MIE": ("ENTRADA_MIE", "SALIDA_MIE"),
    "JUE": ("ENTRADA_JUE", "SALIDA_JUE"),
    "VIE": ("ENTRADA_VIE", "SALIDA_VIE"),
    "SAB": ("ENTRADA_SAB", "SALIDA_SAB"),
    "DOM": ("ENTRADA_DOM", "SALIDA_DOM"),
}

COLS_INCIDENCIAS = [
    "ID", "FOLIO", "RFC", "NOMBRE", "TIPO", "FECHA_SOLICITUD",
    "FECHA_INICIO", "FECHA_FIN", "DIAS", "HORAS_PASE", "HORA_RETORNO",
    "MOTIVO", "TIENE_ANEXO", "LINK_ANEXO", "ESTADO", "AUTORIZADO_POR",
    "FECHA_AUTORIZACION", "OBSERVACIONES"
]

# ─────────────────────────────────────────────
# CONEXIÓN GOOGLE SHEETS
# ─────────────────────────────────────────────
@st.cache_resource
def get_client():
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=SCOPES
    )
    return gspread.authorize(creds)

@st.cache_data(ttl=300)
def cargar_empleados():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_economicos_id"])
    ws = sh.worksheet("Empleados")
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data) if data else pd.DataFrame(columns=[
        "ID", "RFC", "CURP", "PATERNO", "MATERNO", "NOMBRE",
        "PLAZA", "PUESTO", "BASE/INTERINO", "QNA FIN", "C.C.T.",
        "CENTRO DE TRABAJO", "DIAS DISPONIBLES", "DIAS TOTALES"
    ])

@st.cache_data(ttl=300)
def cargar_usuarios():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws = sh.worksheet("Usuarios")
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data) if data else pd.DataFrame(columns=["RFC", "Correo electrónico institucional"])

@st.cache_data(ttl=300)
def cargar_solicitudes_eco():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_economicos_id"])
    ws = sh.worksheet("Solicitudes")
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data) if data else pd.DataFrame()

@st.cache_data(ttl=300)
def cargar_horarios():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws = sh.worksheet("empleados")
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data) if data else pd.DataFrame()

@st.cache_data(ttl=300)
def cargar_directorio_nomina():
    """Lee la tab Directorio_Nomina del Sheet del checador.
    Columnas esperadas: ID, NOMBRE_COMPLETO, CORREO, JEFE_INMEDIATO, CORREO_JEFE, CC_FIJO
    El CC_FIJO solo se llena en las primeras filas (correos de asistentes)."""
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    try:
        ws = sh.worksheet("Directorio_Nomina")
    except gspread.WorksheetNotFound:
        return pd.DataFrame(columns=["ID","NOMBRE_COMPLETO","CORREO","JEFE_INMEDIATO","CORREO_JEFE","CC_FIJO"])
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data) if data else pd.DataFrame(columns=["ID","NOMBRE_COMPLETO","CORREO","JEFE_INMEDIATO","CORREO_JEFE","CC_FIJO"])

@st.cache_data(ttl=300)
def cargar_incidencias():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws = sh.worksheet("Incidencias")
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data) if data else pd.DataFrame(columns=COLS_INCIDENCIAS)

@st.cache_data(ttl=300)
def cargar_historial_horarios():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws = sh.worksheet("HISTORIAL_HORARIOS")
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data) if data else pd.DataFrame(columns=[
        "RFC","NOMBRE","FECHA_INICIO","FECHA_FIN",
        "ENTRADA_LUN","SALIDA_LUN","ENTRADA_MAR","SALIDA_MAR",
        "ENTRADA_MIE","SALIDA_MIE","ENTRADA_JUE","SALIDA_JUE",
        "ENTRADA_VIE","SALIDA_VIE"
    ])

def _leer_hoja_con_reintento(ws, intentos=3, espera=5):
    """Lee get_all_records reintentando si Google devuelve 429 (cuota excedida)."""
    import time as _time
    for i in range(intentos):
        try:
            return ws.get_all_records()
        except gspread.exceptions.APIError as e:
            if "429" in str(e) and i < intentos - 1:
                _time.sleep(espera)
                continue
            raise
    return []

def cargar_festivos():
    """Lee la tab festivos con el MISMO método que el conteo de faltas
    (load_fest_ch), que sí funciona. Sin cache para evitar servir vacíos."""
    import time as _time
    for intento in range(3):
        try:
            client = get_client()
            ws = client.open_by_key(st.secrets["sheet_checador_id"]).worksheet("festivos")
            data = ws.get_all_records()
            return pd.DataFrame(data).fillna("") if data else pd.DataFrame()
        except gspread.exceptions.APIError as e:
            if "429" in str(e) and intento < 2:
                _time.sleep(5)
                continue
            raise
    return pd.DataFrame()

ASISTENCIA_TAB = "Asistencia_Mes"
ASISTENCIA_HEADERS = ["RFC", "NOMBRE", "PERIODO", "FALTAS", "JUSTIFICADAS",
                      "NO_JUSTIFICADAS", "RETARDOS", "DIAS_FALTA", "AVISAR", "FECHA_PROCESO"]

def _ws_asistencia():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    try:
        return sh.worksheet(ASISTENCIA_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(ASISTENCIA_TAB, rows=2, cols=len(ASISTENCIA_HEADERS))
        ws.append_row(ASISTENCIA_HEADERS)
        return ws

def guardar_asistencia_mes(filas: list[dict], periodo: str):
    """Reemplaza el contenido de Asistencia_Mes con el último reporte procesado.
    SIN clear() previo: se escribe encima y el sobrante se limpia solo si la
    escritura tuvo éxito. Si Google falla (429), los datos anteriores quedan
    intactos en vez de perder la tab completa."""
    ws = _ws_asistencia()
    rows = [ASISTENCIA_HEADERS]
    ahora = datetime.now(pytz.timezone("America/Mexico_City")).strftime("%Y-%m-%d %H:%M")
    for f in filas:
        rows.append([
            f.get("RFC",""), f.get("NOMBRE",""), periodo,
            f.get("FALTAS",0), f.get("JUSTIFICADAS",0), f.get("NO_JUSTIFICADAS",0),
            f.get("RETARDOS",0), f.get("DIAS_FALTA",""), f.get("AVISAR",""), ahora
        ])
    ws.update(rows, value_input_option="USER_ENTERED")
    # Éxito: limpiar filas viejas que hayan quedado debajo de la tabla nueva
    try:
        if ws.row_count > len(rows):
            ws.batch_clear([f"A{len(rows) + 1}:J{ws.row_count}"])
    except Exception:
        pass
    cargar_asistencia_mes.clear()

@st.cache_data(ttl=300)
def cargar_asistencia_mes():
    try:
        ws = _ws_asistencia()
        data = ws.get_all_records(numericise_ignore=["all"])
        return pd.DataFrame(data) if data else pd.DataFrame(columns=ASISTENCIA_HEADERS)
    except Exception:
        return pd.DataFrame(columns=ASISTENCIA_HEADERS)

# ── Observaciones de faltas: histórico permanente por RFC + fecha ──
OBS_TAB = "Observaciones_Faltas"
OBS_HEADERS = ["RFC", "NOMBRE", "FECHA_FALTA", "OBSERVACION", "FECHA_REGISTRO"]

def _ws_observaciones():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    try:
        return sh.worksheet(OBS_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(OBS_TAB, rows=2, cols=len(OBS_HEADERS))
        ws.append_row(OBS_HEADERS)
        return ws

@st.cache_data(ttl=300)
def cargar_observaciones():
    try:
        ws = _ws_observaciones()
        data = ws.get_all_records(numericise_ignore=["all"])
        return pd.DataFrame(data) if data else pd.DataFrame(columns=OBS_HEADERS)
    except Exception:
        return pd.DataFrame(columns=OBS_HEADERS)

def guardar_observaciones(nuevas: list[dict]):
    """Agrega/actualiza observaciones por RFC+fecha sin borrar el histórico.
    nuevas: [{RFC, NOMBRE, FECHA_FALTA, OBSERVACION}, ...]"""
    if not nuevas:
        return
    ws = _ws_observaciones()
    existentes = ws.get_all_records(numericise_ignore=["all"])
    # índice por (RFC, FECHA_FALTA) → número de fila (2-based)
    idx = {}
    for i, row in enumerate(existentes, start=2):
        idx[(str(row.get("RFC","")).upper(), str(row.get("FECHA_FALTA","")))] = i
    ahora = datetime.now(pytz.timezone("America/Mexico_City")).strftime("%Y-%m-%d %H:%M")
    from gspread.cell import Cell
    actualizaciones, nuevas_filas = [], []
    for n in nuevas:
        rfc_n = str(n.get("RFC","")).upper()
        fecha_n = str(n.get("FECHA_FALTA",""))
        obs = str(n.get("OBSERVACION","")).strip()
        if not obs:
            continue
        clave = (rfc_n, fecha_n)
        if clave in idx:
            fila = idx[clave]
            actualizaciones.append(Cell(fila, 4, obs))           # col OBSERVACION
            actualizaciones.append(Cell(fila, 5, ahora))         # col FECHA_REGISTRO
        else:
            nuevas_filas.append([rfc_n, n.get("NOMBRE",""), fecha_n, obs, ahora])
    if actualizaciones:
        ws.update_cells(actualizaciones, value_input_option="USER_ENTERED")
    if nuevas_filas:
        ws.append_rows(nuevas_filas, value_input_option="USER_ENTERED")
    cargar_observaciones.clear()


# ── Justificaciones de Dirección: faltas justificadas con Vo.Bo. de la titular ──
# Se guardan en el Sheet y se aplican ANTES de generar el reporte, para que
# Excel y PDF salgan ya corregidos (adiós edición manual del Excel).
JUSTIF_DIR_TAB = "Justificaciones_Direccion"
JUSTIF_DIR_HEADERS = ["RFC", "NOMBRE", "FECHA", "MOTIVO", "REGISTRADO"]

def _ws_justif_dir():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    try:
        return sh.worksheet(JUSTIF_DIR_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(JUSTIF_DIR_TAB, rows=2, cols=len(JUSTIF_DIR_HEADERS))
        ws.append_row(JUSTIF_DIR_HEADERS)
        return ws

@st.cache_data(ttl=60)
def cargar_justif_direccion():
    try:
        ws = _ws_justif_dir()
        data = ws.get_all_records(numericise_ignore=["all"])
        return pd.DataFrame(data) if data else pd.DataFrame(columns=JUSTIF_DIR_HEADERS)
    except Exception:
        return pd.DataFrame(columns=JUSTIF_DIR_HEADERS)

def guardar_justif_direccion(nuevas: list[dict]):
    """nuevas: [{RFC, NOMBRE, FECHA (YYYY-MM-DD), MOTIVO}, ...].
    No duplica: si ya existe RFC+FECHA, actualiza el motivo."""
    if not nuevas:
        return
    ws = _ws_justif_dir()
    existentes = ws.get_all_records(numericise_ignore=["all"])
    idx = {}
    for i, row in enumerate(existentes, start=2):
        idx[(str(row.get("RFC","")).upper().strip(), str(row.get("FECHA","")).strip())] = i
    ahora = datetime.now(pytz.utc).astimezone(pytz.timezone("America/Mexico_City")).strftime("%Y-%m-%d %H:%M")
    from gspread.cell import Cell
    actualizaciones, filas_nuevas = [], []
    for n in nuevas:
        rfc_n = str(n.get("RFC","")).upper().strip()
        fecha_n = str(n.get("FECHA","")).strip()
        if not rfc_n or not fecha_n:
            continue
        clave = (rfc_n, fecha_n)
        if clave in idx:
            fila = idx[clave]
            actualizaciones.append(Cell(fila, 4, str(n.get("MOTIVO",""))))
            actualizaciones.append(Cell(fila, 5, ahora))
        else:
            filas_nuevas.append([rfc_n, n.get("NOMBRE",""), fecha_n, str(n.get("MOTIVO","")), ahora])
    if actualizaciones:
        ws.update_cells(actualizaciones, value_input_option="USER_ENTERED")
    if filas_nuevas:
        ws.append_rows(filas_nuevas, value_input_option="USER_ENTERED")
    cargar_justif_direccion.clear()

def eliminar_justif_direccion(rfc: str, fecha: str):
    """Quita una justificación (para deshacer errores). Localiza la fila EN VIVO."""
    try:
        ws = _ws_justif_dir()
        data = ws.get_all_records(numericise_ignore=["all"])
        for i, row in enumerate(data, start=2):
            if (str(row.get("RFC","")).upper().strip() == str(rfc).upper().strip()
                    and str(row.get("FECHA","")).strip() == str(fecha).strip()):
                ws.delete_rows(i)
                break
        cargar_justif_direccion.clear()
    except Exception as e:
        st.error(f"No se pudo eliminar la justificación: {e}")



# ─────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────
def generar_folio(tipo: str, incidencias_df: pd.DataFrame) -> str:
    anio = datetime.now().year
    prefijo = f"{tipo}-{anio}-"
    # Para ECO buscar en Solicitudes
    if tipo == "ECO":
        try:
            sol = cargar_solicitudes_eco()
            col_folio = next((c for c in sol.columns if "FOLIO" in c.upper()), None)
            if col_folio:
                existentes = sol[sol[col_folio].astype(str).str.startswith(prefijo)][col_folio].tolist()
                nums = []
                for f in existentes:
                    partes = f.split("-")
                    if len(partes) == 3 and partes[2].isdigit():
                        nums.append(int(partes[2]))
                return f"{prefijo}{str(max(nums) + 1 if nums else 1).zfill(4)}"
        except Exception:
            pass
        return f"{prefijo}0001"
    if incidencias_df.empty:
        return f"{prefijo}0001"
    existentes = incidencias_df[
        incidencias_df["FOLIO"].astype(str).str.startswith(prefijo)
    ]["FOLIO"].tolist()
    if not existentes:
        return f"{prefijo}0001"
    nums = []
    for f in existentes:
        partes = f.split("-")
        if len(partes) == 3 and partes[2].isdigit():
            nums.append(int(partes[2]))
    return f"{prefijo}{str(max(nums) + 1 if nums else 1).zfill(4)}"

def festivos_en_periodo(festivos_df: pd.DataFrame, ini: date, fin: date) -> list:
    """Devuelve [(fecha, descripcion), ...] de los festivos que caen dentro
    del período [ini, fin], ordenados por fecha. Expande rangos.
    Usa parseo flexible (pd.to_datetime) porque el Sheet puede devolver
    las fechas en distintos formatos (YYYY-MM-DD, DD/MM/YYYY, serial, etc.)."""
    out = []
    if festivos_df is None or not hasattr(festivos_df, "empty") or festivos_df.empty:
        return out
    # Normalizar ini/fin a date puro (la app puede pasarlos como datetime/Timestamp
    # con hora, lo que rompe la comparación date <= datetime).
    if hasattr(ini, "date") and not isinstance(ini, date):
        ini = ini.date()
    elif isinstance(ini, datetime):
        ini = ini.date()
    if hasattr(fin, "date") and not isinstance(fin, date):
        fin = fin.date()
    elif isinstance(fin, datetime):
        fin = fin.date()
    col_desc = next((c for c in festivos_df.columns if "DESCRIP" in str(c).upper()), None)
    col_fi   = next((c for c in festivos_df.columns if "INICIO" in str(c).upper()), None)
    col_ff   = next((c for c in festivos_df.columns if "FIN" in str(c).upper()), None)
    if not col_fi:
        return out
    def _parse_fecha(v):
        """Parsea fecha desde texto (ISO, DD/MM), datetime, o serial de Sheets."""
        if v is None or v == "":
            return None
        # Número serial de Google Sheets (días desde 1899-12-30)
        try:
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.replace(".","",1).isdigit()):
                serial = float(v)
                if serial > 30000:  # rango razonable de fechas modernas
                    return (datetime(1899,12,30) + timedelta(days=int(serial))).date()
        except Exception:
            pass
        ts = pd.to_datetime(v, errors="coerce")
        return ts.date() if not pd.isna(ts) else None

    for _, row in festivos_df.iterrows():
        try:
            fi = _parse_fecha(row.get(col_fi, ""))
            ff = _parse_fecha(row.get(col_ff, "") if col_ff else row.get(col_fi, ""))
            if fi is None:
                continue
            if ff is None:
                ff = fi
            desc = str(row[col_desc]).strip() if col_desc else ""
            d = fi
            while d <= ff:
                if ini <= d <= fin:
                    out.append((d, desc))
                d += timedelta(days=1)
        except Exception:
            pass
    return sorted(out, key=lambda x: x[0])

def festivos_a_set(festivos_df: pd.DataFrame) -> set:
    """Expande los rangos de la tab festivos a un set de fechas individuales."""
    festivos_set = set()
    for _, row in festivos_df.iterrows():
        try:
            fi = datetime.strptime(str(row["FECHA_INICIO"]), "%Y-%m-%d").date()
            ff = datetime.strptime(str(row["FECHA_FIN"]), "%Y-%m-%d").date()
            d = fi
            while d <= ff:
                festivos_set.add(d)
                d += timedelta(days=1)
        except Exception:
            pass
    return festivos_set

def dias_habiles_entre(fecha_inicio: date, fecha_fin: date, festivos_df: pd.DataFrame) -> int:
    festivos_set = festivos_a_set(festivos_df)
    total = 0
    d = fecha_inicio
    while d <= fecha_fin:
        if d.weekday() < 5 and d not in festivos_set:
            total += 1
        d += timedelta(days=1)
    return total

def dias_economicos_usados(rfc: str, solicitudes_df: pd.DataFrame) -> int:
    df = solicitudes_df[solicitudes_df["RFC"].astype(str).str.upper() == rfc.upper()]
    # Solo cuentan las APROBADAS: la columna tiene nombre de admin.
    # Las rechazadas escriben "RECHAZADO — ..." en la misma columna y NO deben descontar días.
    aprob = df["Aprobado Por"].astype(str).str.strip()
    df_aprobados = df[(aprob != "") & (~aprob.str.upper().str.startswith("RECHAZADO"))]
    total = 0
    for _, row in df_aprobados.iterrows():
        try:
            total += int(row["Dias Solicitados"])
        except Exception:
            pass
    return total

def actualizar_dias_disponibles_sheet(rfc_objetivo: str = None):
    """Recalcula DIAS DISPONIBLES (= DIAS TOTALES - usados aprobados) y lo escribe
    en la hoja Empleados del Sheet de económicos.
    Si rfc_objetivo se indica, solo actualiza esa fila; si no, actualiza todas.
    Devuelve (actualizados, errores)."""
    from gspread.cell import Cell
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_economicos_id"])
    ws = sh.worksheet("Empleados")
    registros = ws.get_all_records(numericise_ignore=["all"])
    if not registros:
        return 0, []
    headers = list(registros[0].keys())
    try:
        col_disp = headers.index("DIAS DISPONIBLES") + 1
    except ValueError:
        return 0, ["No existe la columna 'DIAS DISPONIBLES' en la hoja Empleados"]
    col_rfc = next((h for h in headers if h.upper() == "RFC"), "RFC")
    col_tot = next((h for h in headers if "DIAS TOTALES" in h.upper()), "DIAS TOTALES")

    solicitudes = cargar_solicitudes_eco()
    celdas, errores = [], []
    for i, row in enumerate(registros, start=2):  # fila 1 = headers
        rfc = str(row.get(col_rfc, "")).upper().strip()
        if not rfc:
            continue
        if rfc_objetivo and rfc != rfc_objetivo.upper().strip():
            continue
        try:
            total = int(row.get(col_tot, 0) or 0)
        except Exception:
            total = 0
        usados = dias_economicos_usados(rfc, solicitudes)
        disponibles = total - usados
        celdas.append(Cell(i, col_disp, disponibles))
    if celdas:
        ws.update_cells(celdas, value_input_option="USER_ENTERED")
    cargar_empleados.clear()
    return len(celdas), errores

def horas_pases_mes(rfc: str, incidencias_df: pd.DataFrame) -> float:
    """Suma las horas de pases autorizados del mes en curso usando FECHA_INICIO."""
    import pytz
    tz_mx = pytz.timezone("America/Mexico_City")
    ahora = datetime.now(pytz.utc).astimezone(tz_mx)
    df = incidencias_df[
        (incidencias_df["RFC"].astype(str).str.upper() == rfc.upper()) &
        (incidencias_df["TIPO"].isin(["PSE","PSR","PEN"])) &
        (incidencias_df["ESTADO"].astype(str).str.contains("AUTORIZADO"))
    ].copy()
    if df.empty:
        return 0.0
    df["FECHA_DT"] = pd.to_datetime(df["FECHA_INICIO"], errors="coerce")
    df = df[(df["FECHA_DT"].dt.year == ahora.year) & (df["FECHA_DT"].dt.month == ahora.month)]
    total = 0.0
    for _, row in df.iterrows():
        try:
            val = str(row.get("HORAS_PASE", 0) or 0).replace(",", ".")
            total += float(val)
        except Exception:
            pass
    return round(total, 2)

def rfc_a_fecha_nacimiento(rfc: str):
    """Extrae fecha de nacimiento del RFC (posiciones 4-9: AAMMDD)."""
    try:
        anio = int(rfc[4:6])
        mes  = int(rfc[6:8])
        dia  = int(rfc[8:10])
        pivote = datetime.now().year % 100
        anio_completo = 2000 + anio if anio <= pivote else 1900 + anio
        return date(anio_completo, mes, dia)
    except Exception:
        return None

def cumpleanos_laboral(rfc: str) -> date:
    """Retorna el día hábil de cumpleaños (lunes si cae sábado, viernes si domingo)."""
    fn = rfc_a_fecha_nacimiento(rfc)
    if not fn:
        return None
    hoy = date.today()
    cumple = date(hoy.year, fn.month, fn.day)
    if cumple < hoy:
        cumple = date(hoy.year + 1, fn.month, fn.day)
    if cumple.weekday() == 5:  # sábado → lunes
        cumple += timedelta(days=2)
    elif cumple.weekday() == 6:  # domingo → viernes
        cumple -= timedelta(days=2)
    return cumple

def rfc_oculto(rfc: str) -> str:
    """Muestra solo los primeros 4 caracteres del RFC."""
    return rfc[:4] + "****" + rfc[-3:] if len(rfc) >= 7 else rfc

# ─────────────────────────────────────────────
# GENERACIÓN DE COMPROBANTE PDF
# ─────────────────────────────────────────────
def generar_comprobante_pdf(datos: dict) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.2*cm, bottomMargin=1.2*cm)
    styles = getSampleStyleSheet()

    estilo_titulo   = ParagraphStyle("t", parent=styles["Normal"], fontSize=13, fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)
    estilo_sub      = ParagraphStyle("s", parent=styles["Normal"], fontSize=10, fontName="Helvetica", alignment=TA_CENTER, spaceAfter=2)
    estilo_folio    = ParagraphStyle("f", parent=styles["Normal"], fontSize=16, fontName="Helvetica-Bold", alignment=TA_CENTER, textColor=colors.HexColor("#002F6C"), spaceAfter=6)
    estilo_aviso    = ParagraphStyle("a", parent=styles["Normal"], fontSize=8, fontName="Helvetica-Oblique", alignment=TA_CENTER, textColor=colors.HexColor("#888888"))

    elementos = []
    # Logo SEJ
    import os
    logo_path = "logos_gris.png"
    if os.path.exists(logo_path):
        logo_rl = RLImage(logo_path, width=6*cm, height=1.5*cm)
        tabla_header = Table([[logo_rl, Paragraph(
            "<b>Dirección de Formación Continua</b><br/>Área de Recursos Humanos · SEJ",
            ParagraphStyle("hdr", parent=styles["Normal"], fontSize=9, fontName="Helvetica", alignment=TA_CENTER)
        )]], colWidths=[7*cm, 10*cm])
        tabla_header.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(1,0),(1,0),"CENTER")]))
        elementos.append(tabla_header)
    else:
        elementos.append(Paragraph("Secretaría de Educación Jalisco", estilo_titulo))
        elementos.append(Paragraph("Dirección de Formación Continua · Área de Recursos Humanos", estilo_sub))
    elementos.append(Spacer(1, 0.1*cm))
    elementos.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#002F6C")))
    elementos.append(Spacer(1, 0.1*cm))
    elementos.append(Paragraph("COMPROBANTE DE CAPTURA DE INCIDENCIA", estilo_titulo))
    elementos.append(Spacer(1, 0.1*cm))
    elementos.append(Paragraph(f"FOLIO: {datos['folio']}", estilo_folio))
    elementos.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
    elementos.append(Spacer(1, 0.15*cm))

    tabla_datos = [
        ["Nombre completo:", datos["nombre"]],
        ["Filiación:", rfc_oculto(datos["rfc"])],
        ["Tipo de incidencia:", datos.get("subtipo_label", datos["tipo_label"])],
        ["Fecha de solicitud:", datos["fecha_solicitud"]],
    ]
    if datos["tipo"] == "CHO":
        tabla_datos.insert(4, ["Fecha de aplicación:", datos["fecha_inicio"]])
    else:
        tabla_datos.insert(4, ["Fecha inicio:", datos["fecha_inicio"]])
        tabla_datos.insert(5, ["Fecha fin:", datos["fecha_fin"]])
    horas_val = datos.get("horas_pase", 0)
    tabla_datos += [
        ["Días solicitados:", "N/A" if datos["tipo"] in ["PSE", "CHO"] else str(datos["dias"])],
        ["Horas de pase:", (f"{horas_val}h" if horas_val else "No registradas") if datos["tipo"] == "PSE" else "N/A"],
        ["Motivo / Descripción:", Paragraph(str(datos["motivo"]), ParagraphStyle("mot", parent=styles["Normal"], fontSize=9, fontName="Helvetica"))],
        ["Documento anexo:", "Sí, se presentará en RH" if datos["tiene_anexo"] else "No aplica"],
    ]

    t = Table(tabla_datos, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",    (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.HexColor("#F4F6F9"), colors.white]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",(0, 0), (-1, -1), 8),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]))
    elementos.append(t)
    elementos.append(Spacer(1, 0.1*cm))
    elementos.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
    elementos.append(Spacer(1, 0.1*cm))

    color_estado = colors.HexColor("#F59E0B")
    et = Table([[Paragraph("<b>ESTADO: PENDIENTE DE AUTORIZACIÓN</b>",
                           ParagraphStyle("e", parent=styles["Normal"], fontSize=10,
                                          fontName="Helvetica-Bold", textColor=color_estado,
                                          alignment=TA_CENTER))]], colWidths=[16*cm])
    et.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 1, color_estado),
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#FFFBEB")),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    elementos.append(et)
    elementos.append(Spacer(1, 0.1*cm))
    # ── Bloque de firmas ─────────────────────────
    elementos.append(Paragraph("<b>REQUISITO OBLIGATORIO: SECCIÓN DE FIRMAS</b>",
        ParagraphStyle("tf", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", spaceAfter=6)))

    estilo_firma = ParagraphStyle("firmas", parent=styles["Normal"], fontSize=8, fontName="Helvetica", alignment=TA_CENTER)
    nombre_interesado = datos.get("nombre", "Servidor(a) Público(a)")
    jefe_pdf_texto    = datos.get("jefe_inmediato", "Nombre y Firma")
    # Director(a) General: viene del Sheet (columna DIRECTOR_GENERAL); si no, usa la constante
    director_pdf      = datos.get("director_general") or NOMBRE_DIRECTORA

    # ¿El jefe inmediato es la misma persona que el/la director(a) general?
    def _solo_nombre(txt):
        return str(txt or "").split("<br/>")[0].split("—")[0].split(" - ")[0].strip().lower()
    jefe_es_director = _solo_nombre(jefe_pdf_texto) and _solo_nombre(jefe_pdf_texto) == _solo_nombre(director_pdf)

    firma_interesado  = Paragraph(f"<br/><br/>___________________________<br/><b>Firma del Interesado</b><br/>{nombre_interesado}", estilo_firma)
    firma_jefe        = Paragraph(f"<br/><br/>___________________________<br/><b>Autoriza Jefe(a) Inmediato</b><br/>{jefe_pdf_texto}", estilo_firma)
    firma_vob         = Paragraph(f"<br/><br/>___________________________<br/><b>Vo.Bo. Titular del Área</b><br/>{director_pdf}", estilo_firma)

    if datos["tipo"] in ["ECO", "CHO", "RGU"]:
        if jefe_es_director:
            # La titular es a su vez la jefa inmediata: solo se pone Vo.Bo. del titular.
            t_firmas = Table([[firma_interesado, firma_vob]], colWidths=[8*cm, 8*cm])
        else:
            t_firmas = Table([[firma_interesado, firma_jefe, firma_vob]], colWidths=[5.3*cm, 5.3*cm, 5.4*cm])
    else:
        t_firmas = Table([[firma_interesado, firma_jefe]], colWidths=[8*cm, 8*cm])

    t_firmas.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    elementos.append(t_firmas)
    elementos.append(Spacer(1, 0.1*cm))
    elementos.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
    elementos.append(Spacer(1, 0.1*cm))

    # ── Instrucciones finales ─────────────────────
    elementos.append(Paragraph("<b>INSTRUCCIONES DE ENTREGA:</b>",
        ParagraphStyle("tit_i", parent=styles["Normal"], fontSize=8, fontName="Helvetica-Bold", spaceAfter=3)))
    for inst in [
        "1. Recaba las firmas físicas obligatorias que se muestran arriba.",
        "2. Adjunta la documentación de soporte original si aplica (cita IMSS, oficio de comisión, etc.).",
        "3. Entrega el expediente completo en el Área de Recursos Humanos (Administración de Personal DFC).",
        "4. Resguarda tu copia digital. El folio es tu número de seguimiento oficial.",
    ]:
        elementos.append(Paragraph(inst, ParagraphStyle("i", parent=styles["Normal"],
                                                         fontSize=8, fontName="Helvetica",
                                                         leftIndent=10, spaceAfter=3)))

    # ── QR de seguridad ──────────────────────────
    url_validacion = f"https://gestion-personal-dfc.streamlit.app/?validar_folio={datos['folio']}"
    qr = qrcode.QRCode(version=1, box_size=10, border=1)
    qr.add_data(url_validacion)
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = BytesIO()
    img_qr.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)
    qr_flowable = RLImage(qr_buffer, width=2.5*cm, height=2.5*cm)
    tabla_qr = Table([
        [qr_flowable, Paragraph(
            "<b>ESCÁNER DE SEGURIDAD DIGITAL</b><br/>Escanea este QR para verificar la autenticidad de este documento en tiempo real. Cualquier alteración anula el comprobante.",
            ParagraphStyle("txt_qr", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#555555"))
        )]
    ], colWidths=[3*cm, 13*cm])
    tabla_qr.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "MIDDLE")]))
    elementos.append(tabla_qr)
    elementos.append(Spacer(1, 0.15*cm))
    elementos.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
    elementos.append(Spacer(1, 0.1*cm))
    elementos.append(Paragraph(f"Generado el {datos['fecha_solicitud']} · Sistema ARI — DFC · SEJ", estilo_aviso))
    doc.build(elementos)
    buffer.seek(0)
    return buffer.read()

# ─────────────────────────────────────────────
# ESCRITURA EN SHEETS
# ─────────────────────────────────────────────
def subir_anexo_drive(archivo, folio: str, rfc: str) -> str:
    """Sube un archivo a Google Drive y retorna el link de visualización."""
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        from google.oauth2.service_account import Credentials
        import io

        creds = Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=["https://www.googleapis.com/auth/drive"]
        )
        service = build("drive", "v3", credentials=creds)

        ext = archivo.name.split(".")[-1].lower()
        nombre_archivo = f"{folio}_{rfc}.{ext}"
        media = MediaIoBaseUpload(io.BytesIO(archivo.read()), mimetype=archivo.type)
        file_metadata = {
            "name": nombre_archivo,
            "parents": [DRIVE_ANEXOS_FOLDER],
        }
        archivo_drive = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id",
            supportsAllDrives=True
        ).execute()

        file_id = archivo_drive.get("id")
        # NOTA: sin permiso "anyone" — el acceso se hereda de la membresía
        # de la unidad compartida (solo RH y la service account).
        return f"https://drive.google.com/file/d/{file_id}/view"
    except Exception as e:
        return f"ERROR: {e}"

def _envio_duplicado(clave: str) -> bool:
    """Candado anti doble clic: True si esta MISMA solicitud ya se envió en los
    últimos 2 minutos de esta sesión (doble clic o rerun de Streamlit).
    Evita solicitudes y folios duplicados."""
    import hashlib
    import time as _time
    h = hashlib.md5(clave.encode()).hexdigest()
    ult = st.session_state.get("_ultimo_envio")
    if ult and ult[0] == h and (_time.time() - ult[1]) < 120:
        return True
    st.session_state["_ultimo_envio"] = (h, _time.time())
    return False

def _error_amable(e: Exception, contexto: str = ""):
    """Traduce errores de cuota de Google a un mensaje humano en vez de traceback."""
    if "429" in str(e) or "Quota" in str(e) or "quota" in str(e):
        st.error("⏳ El sistema está ocupado en este momento (muchas personas usándolo a la vez). "
                 "Espera 15 segundos y vuelve a intentar. Tu información no se perdió.")
    else:
        st.error(f"Error {contexto}: {e}")

def guardar_dia_economico(datos: dict):
    """Guarda día económico en tab Solicitudes del Sheet económicos — igual que app3."""
    # Candado: la misma persona + mismas fechas en <2 min = doble clic
    if _envio_duplicado(f"ECO|{datos.get('rfc','')}|{datos.get('fecha_inicio','')}|{datos.get('fecha_fin','')}"):
        st.warning("Esta solicitud ya se envió hace un momento. Si crees que es un error, espera 2 minutos.")
        return
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_economicos_id"])
    ws = sh.worksheet("Solicitudes")
    todas = ws.get_all_records(numericise_ignore=["all"])
    nuevo_id = len(todas) + 1
    # FOLIO recalculado EN VIVO sobre lo recién leído (no sobre cache):
    # cierra la ventana de folios duplicados por concurrencia.
    try:
        anio_f = datetime.now().year
        pref = f"ECO-{anio_f}-"
        nums = []
        for t in todas:
            fol = str(t.get("FOLIO", ""))
            if fol.startswith(pref):
                pp = fol.split("-")
                if len(pp) == 3 and pp[2].isdigit():
                    nums.append(int(pp[2]))
        datos["folio"] = f"{pref}{str(max(nums) + 1 if nums else 1).zfill(4)}"
    except Exception:
        pass  # conserva el folio previamente generado
    emp_id = ""
    try:
        empleados = cargar_empleados()
        match = empleados[empleados["RFC"].astype(str).str.upper() == datos["rfc"].upper()]
        if not match.empty:
            emp_id = match.iloc[0].get("ID", "")
    except Exception:
        pass
    link_anexo = ""
    if datos.get("tiene_anexo") and datos.get("archivo_anexo"):
        try:
            link_anexo = subir_anexo_drive(datos["archivo_anexo"], datos["folio"], datos["rfc"])
        except Exception as e:
            st.warning(f"No se pudo subir el anexo: {e}")
    fila = [
        nuevo_id,                    # ID
        emp_id,                      # EmpleadoID
        datos["rfc"],                 # RFC
        datos["nombre"],              # Nombre Completo
        "economico",                 # Tipo Permiso
        datos["fecha_inicio"],        # Fecha Inicio
        datos["fecha_fin"],           # Fecha Fin
        datos["dias"],                # Dias Solicitados
        datos["motivo"],              # Motivo
        datos["fecha_solicitud"],     # Fecha Registro
        "",                          # Aprobado Por
        datos["nombre"],              # Registrado Por
        datos["folio"],               # FOLIO
        link_anexo,                  # LINK_ANEXO
    ]
    ws.append_row(fila, value_input_option="USER_ENTERED")
    cargar_solicitudes_eco.clear()
    return True

def guardar_incidencia(datos: dict):
    # Candado: la misma persona + tipo + fechas en <2 min = doble clic
    if _envio_duplicado(f"{datos.get('tipo','')}|{datos.get('rfc','')}|{datos.get('fecha_inicio','')}|{datos.get('fecha_fin','')}"):
        st.warning("Esta solicitud ya se envió hace un momento. Si crees que es un error, espera 2 minutos.")
        return False
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws = sh.worksheet("Incidencias")
    todas = ws.get_all_records(numericise_ignore=["all"])
    # FOLIO recalculado EN VIVO sobre lo recién leído (cierra duplicados por concurrencia)
    try:
        anio_f = datetime.now().year
        pref = f"{datos.get('tipo','')}-{anio_f}-"
        nums = []
        for t in todas:
            fol = str(t.get("FOLIO", ""))
            if fol.startswith(pref):
                pp = fol.split("-")
                if len(pp) == 3 and pp[2].isdigit():
                    nums.append(int(pp[2]))
        datos["folio"] = f"{pref}{str(max(nums) + 1 if nums else 1).zfill(4)}"
    except Exception:
        pass
    fila = [
        len(todas) + 1,
        datos["folio"],
        datos["rfc"],
        datos["nombre"],
        datos["tipo"],
        datos["fecha_solicitud"],
        datos["fecha_inicio"],
        datos["fecha_fin"],
        datos["dias"],
        datos.get("horas_pase", ""),
        datos.get("hora_retorno", ""),
        datos["motivo"],
        "SÍ" if datos["tiene_anexo"] else "NO",
        datos.get("link_anexo", ""),
        "PENDIENTE",
        "", "", "",
    ]
    ws.append_row(fila, value_input_option="USER_ENTERED")
    cargar_incidencias.clear()
    return True

def autorizar_incidencia(folio: str, obs: str = ""):
    try:
        import pytz
        tz_mx = pytz.timezone("America/Mexico_City")
        ahora = datetime.now(pytz.utc).astimezone(tz_mx).strftime("%Y-%m-%d %H:%M")
        client = get_client()
        sh = client.open_by_key(st.secrets["sheet_checador_id"])
        ws = sh.worksheet("Incidencias")
        headers = [h.upper().strip() for h in ws.row_values(1)]
        data = ws.get_all_records(numericise_ignore=["all"])
        if "ESTADO" not in headers:
            st.error("No se encontró columna ESTADO en el Sheet.")
            return
        from gspread.cell import Cell
        for i, row in enumerate(data, start=2):
            if str(row.get("FOLIO", "")) == folio:
                celdas = [
                    Cell(i, headers.index("ESTADO") + 1,             "AUTORIZADO"),
                    Cell(i, headers.index("AUTORIZADO_POR") + 1,     st.session_state.get("nombre", "admin")),
                    Cell(i, headers.index("FECHA_AUTORIZACION") + 1, ahora),
                ]
                if obs and "OBSERVACIONES" in headers:
                    celdas.append(Cell(i, headers.index("OBSERVACIONES") + 1, obs))
                ws.update_cells(celdas, value_input_option="USER_ENTERED")
                break
        cargar_incidencias.clear()
    except gspread.exceptions.APIError:
        st.error("⏳ Google Sheets está saturado temporalmente. Espera 5 segundos e intenta de nuevo.")
    except Exception as e:
        _error_amable(e, "al autorizar")

def rechazar_incidencia(folio: str, obs: str):
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws = sh.worksheet("Incidencias")
    headers = ws.row_values(1)
    data = ws.get_all_records(numericise_ignore=["all"])
    from gspread.cell import Cell
    ahora = datetime.now(pytz.utc).astimezone(pytz.timezone("America/Mexico_City")).strftime("%Y-%m-%d %H:%M")
    for i, row in enumerate(data, start=2):
        if str(row.get("FOLIO", "")) == folio:
            ws.update_cells([
                Cell(i, headers.index("ESTADO") + 1,             "RECHAZADO"),
                Cell(i, headers.index("AUTORIZADO_POR") + 1,     st.session_state.get("nombre", "admin")),
                Cell(i, headers.index("FECHA_AUTORIZACION") + 1, ahora),
                Cell(i, headers.index("OBSERVACIONES") + 1,      obs),
            ], value_input_option="USER_ENTERED")
            break
    cargar_incidencias.clear()

def _localizar_fila_eco(ws, folio: str = "", rfc: str = "", fecha_inicio: str = "", fecha_registro: str = ""):
    """Relee el Sheet EN VIVO y localiza la fila de la solicitud.
    Prioridad: FOLIO exacto; si no hay folio, RFC + Fecha Inicio + Fecha Registro
    (para solicitudes viejas sin folio). Devuelve (num_fila, headers_norm) o (None, headers_norm).
    NUNCA usar índices de DataFrames cacheados: si alguien insertó/borró filas
    en el Sheet, se aprobaría la solicitud equivocada."""
    headers = [h.upper().strip() for h in ws.row_values(1)]
    data = ws.get_all_records(numericise_ignore=["all"])
    folio = str(folio or "").strip()
    for i, row in enumerate(data, start=2):
        row_norm = {str(k).upper().strip(): str(v).strip() for k, v in row.items()}
        if folio and row_norm.get("FOLIO", "") == folio:
            return i, headers
        if not folio and rfc:
            if (row_norm.get("RFC", "").upper() == rfc.upper().strip()
                    and row_norm.get("FECHA INICIO", "") == str(fecha_inicio).strip()
                    and row_norm.get("FECHA REGISTRO", "") == str(fecha_registro).strip()):
                return i, headers
    return None, headers

def rechazar_dia_economico(obs: str, folio: str = "", rfc: str = "", fecha_inicio: str = "", fecha_registro: str = ""):
    try:
        client = get_client()
        sh = client.open_by_key(st.secrets["sheet_economicos_id"])
        ws = sh.worksheet("Solicitudes")
        fila, headers = _localizar_fila_eco(ws, folio, rfc, fecha_inicio, fecha_registro)
        if fila is None:
            st.error(f"No encontré la solicitud {folio or rfc} en el Sheet. Recarga la página e intenta de nuevo.")
            return
        col_aprobado = headers.index("APROBADO POR") + 1 if "APROBADO POR" in headers else None
        if col_aprobado:
            ws.update_cell(fila, col_aprobado, f"RECHAZADO — {obs}")
        cargar_solicitudes_eco.clear()
        st.warning("Solicitud rechazada.")
    except Exception as e:
        _error_amable(e, "al rechazar")

def aprobar_dia_economico(nombre_admin: str, folio: str = "", rfc: str = "", fecha_inicio: str = "", fecha_registro: str = ""):
    """Escribe el nombre del admin en Aprobado Por, localizando la fila por FOLIO en vivo."""
    try:
        import pytz
        from gspread.cell import Cell
        tz_mx = pytz.timezone("America/Mexico_City")
        ahora = datetime.now(pytz.utc).astimezone(tz_mx).strftime("%Y-%m-%d %H:%M")
        client = get_client()
        sh = client.open_by_key(st.secrets["sheet_economicos_id"])
        ws = sh.worksheet("Solicitudes")
        fila, headers = _localizar_fila_eco(ws, folio, rfc, fecha_inicio, fecha_registro)
        if fila is None:
            st.error(f"No encontré la solicitud {folio or rfc} en el Sheet. Recarga la página e intenta de nuevo.")
            return
        celdas = []
        if "APROBADO POR" in headers:
            celdas.append(Cell(fila, headers.index("APROBADO POR") + 1, nombre_admin))
        col_fecha_aut = next((j + 1 for j, h in enumerate(headers) if "FECHA" in h and "AUTORIZACION" in h), None)
        if col_fecha_aut:
            celdas.append(Cell(fila, col_fecha_aut, ahora))
        if celdas:
            ws.update_cells(celdas, value_input_option="USER_ENTERED")
        cargar_solicitudes_eco.clear()
        # Actualizar el saldo de DIAS DISPONIBLES de ese empleado en el Sheet Empleados
        try:
            rfc_sol = rfc
            if not rfc_sol and "RFC" in headers:
                rfc_sol = ws.cell(fila, headers.index("RFC") + 1).value
            if rfc_sol:
                actualizar_dias_disponibles_sheet(rfc_sol)
        except Exception:
            pass
        st.success("Día económico autorizado correctamente.")
    except Exception as e:
        _error_amable(e, "al aprobar")

def autorizar_cambio_horario(emp_id: str, horario_nuevo: dict, folio: str, fecha_inicio_cho: str = None):
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws_emp = sh.worksheet("empleados")
    registros = ws_emp.get_all_records(numericise_ignore=["all"])
    headers   = ws_emp.row_values(1)
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    fecha_aplicacion = fecha_inicio_cho if fecha_inicio_cho else fecha_hoy

    for i, row in enumerate(registros, start=2):
        if str(row.get("RFC", "")).upper().strip() == str(emp_id).upper().strip() or str(row.get("NOMBRE", "")).upper().strip() == str(emp_id).upper().strip():
            # Guardar horario ANTERIOR en HISTORIAL_HORARIOS antes de sobrescribir
            try:
                ws_hist = sh.worksheet("HISTORIAL_HORARIOS")
                hist_data = ws_hist.get_all_records(numericise_ignore=["all"])
                fila_hist = [
                    str(row.get("RFC", emp_id)),
                    str(row.get("NOMBRE", "")),
                    "",  # FECHA_INICIO — se llenará con la del registro anterior
                    (datetime.strptime(fecha_aplicacion, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d"),  # FECHA_FIN — día anterior al cambio
                    str(row.get("ENTRADA_LUN", "")), str(row.get("SALIDA_LUN", "")),
                    str(row.get("ENTRADA_MAR", "")), str(row.get("SALIDA_MAR", "")),
                    str(row.get("ENTRADA_MIE", "")), str(row.get("SALIDA_MIE", "")),
                    str(row.get("ENTRADA_JUE", "")), str(row.get("SALIDA_JUE", "")),
                    str(row.get("ENTRADA_VIE", "")), str(row.get("SALIDA_VIE", "")),
                ]
                ws_hist.append_row(fila_hist, value_input_option="USER_ENTERED")
                # Agregar también el nuevo horario con FECHA_INICIO
                fila_nuevo = [
                    str(row.get("RFC", emp_id)),
                    str(row.get("NOMBRE", "")),
                    str(fecha_aplicacion),  # FECHA_INICIO
                    "",  # FECHA_FIN — vacío porque es el vigente
                    horario_nuevo.get("LUN",{}).get("entrada",""), horario_nuevo.get("LUN",{}).get("salida",""),
                    horario_nuevo.get("MAR",{}).get("entrada",""), horario_nuevo.get("MAR",{}).get("salida",""),
                    horario_nuevo.get("MIE",{}).get("entrada",""), horario_nuevo.get("MIE",{}).get("salida",""),
                    horario_nuevo.get("JUE",{}).get("entrada",""), horario_nuevo.get("JUE",{}).get("salida",""),
                    horario_nuevo.get("VIE",{}).get("entrada",""), horario_nuevo.get("VIE",{}).get("salida",""),
                ]
                ws_hist.append_row(fila_nuevo, value_input_option="USER_ENTERED")
            except Exception as e:
                st.warning(f"No se pudo guardar historial: {e}")

            # Sobrescribir horario actual en tab empleados
            from gspread.cell import Cell
            celdas_horario = []
            for dia, (col_e, col_s) in COLUMNAS_HORARIO.items():
                if dia in horario_nuevo:
                    if col_e in headers:
                        celdas_horario.append(Cell(i, headers.index(col_e) + 1, horario_nuevo[dia]["entrada"]))
                    if col_s in headers:
                        celdas_horario.append(Cell(i, headers.index(col_s) + 1, horario_nuevo[dia]["salida"]))
            if celdas_horario:
                ws_emp.update_cells(celdas_horario, value_input_option="USER_ENTERED")
            break
    ws_inc = sh.worksheet("Incidencias")
    inc_h  = ws_inc.row_values(1)
    inc_d  = ws_inc.get_all_records(numericise_ignore=["all"])
    from gspread.cell import Cell
    ahora_inc = datetime.now(pytz.utc).astimezone(pytz.timezone("America/Mexico_City")).strftime("%Y-%m-%d %H:%M")
    for i, row in enumerate(inc_d, start=2):
        if str(row.get("FOLIO", "")) == folio:
            ws_inc.update_cells([
                Cell(i, inc_h.index("ESTADO") + 1,             "AUTORIZADO"),
                Cell(i, inc_h.index("AUTORIZADO_POR") + 1,     st.session_state.get("nombre", "admin")),
                Cell(i, inc_h.index("FECHA_AUTORIZACION") + 1, ahora_inc),
            ], value_input_option="USER_ENTERED")
            break
    cargar_incidencias.clear()

# ─────────────────────────────────────────────
# AUTENTICACIÓN
# ─────────────────────────────────────────────
def login():
    col_izq, col_centro, col_der = st.columns([1.1, 1.8, 1.1])
    with col_centro:
        with st.container(border=True):
            col_l, col_img, col_r = st.columns([1, 2, 1])
            with col_img:
                st.image("dfc_logo.png", use_container_width=True)
            st.markdown("<h3 style='text-align:center;margin-bottom:0;font-size:22px'>🪪 Ingreso al Sistema</h3>", unsafe_allow_html=True)
            st.markdown("<p style='text-align:center;color:gray;font-size:13px'>Dirección de Formación Continua · Personal</p>", unsafe_allow_html=True)
            st.divider()
            correo    = st.text_input("Correo electrónico institucional", placeholder="nombre@jalisco.gob.mx")
            rfc_input = st.text_input("RFC (Contraseña)", type="password", placeholder="XXXX000000XXX")
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
            entrar = st.button("Iniciar Sesión", type="primary", use_container_width=True)

    if entrar:
        if not correo or not rfc_input:
            st.error("Ingresa tu correo y RFC.")
            return

        admin_correo = st.secrets.get("admin_correo", "")
        admin_rfc    = st.secrets.get("admin_rfc", "")
        if correo.lower() == admin_correo.lower() and rfc_input.upper() == admin_rfc.upper():
            st.session_state["rol"]    = "admin"
            st.session_state["correo"] = correo
            st.session_state["rfc"]    = rfc_input.upper()
            st.session_state["nombre"] = "Administrador RH"
            st.rerun()
            return

        usuarios  = cargar_usuarios()
        match_usr = usuarios[usuarios["RFC"].astype(str).str.upper() == rfc_input.upper()]
        if match_usr.empty:
            st.error("RFC no encontrado. Verifica tus datos.")
            return

        usr_row           = match_usr.iloc[0]
        correo_registrado = str(usr_row.get("Correo electrónico institucional", "")).strip().lower()
        if not correo_registrado:
            # SEGURIDAD: sin correo registrado NO se permite el acceso.
            # Antes, un correo vacío en el Sheet dejaba entrar con solo el RFC.
            st.error("Tu RFC no tiene correo registrado en el sistema. Contacta a RH para completar tu registro.")
            return
        if correo.strip().lower() != correo_registrado:
            st.error("El correo no coincide con el RFC registrado.")
            return

        empleados = cargar_empleados()
        match_emp = empleados[empleados["RFC"].astype(str).str.upper() == rfc_input.upper()]
        if not match_emp.empty:
            emp_row         = match_emp.iloc[0]
            nombre_completo = f"{emp_row.get('PATERNO','')} {emp_row.get('MATERNO','')} {emp_row.get('NOMBRE','')}".strip()
            emp_dict        = emp_row.to_dict()
        else:
            # Buscar nombre en tab empleados del reloj checador por RFC
            try:
                horarios_tmp = cargar_horarios()
                match_hor = horarios_tmp[horarios_tmp["RFC"].astype(str).str.upper().str.strip() == rfc_input.upper().strip()]
                if not match_hor.empty:
                    nombre_completo = match_hor.iloc[0].get("NOMBRE", correo.split("@")[0].replace(".", " ").title())
                else:
                    nombre_completo = correo.split("@")[0].replace(".", " ").title()
            except Exception:
                nombre_completo = correo.split("@")[0].replace(".", " ").title()
            emp_dict = {}

        jefe_auto = str(usr_row.get("JEFE_INMEDIATO", "")).strip()
        jefe_pdf_auto = formato_jefe_pdf(jefe_auto)
        director_auto = formato_jefe_pdf(str(usr_row.get("DIRECTOR_GENERAL", "")).strip())
        st.session_state["director_general"] = director_auto
        st.session_state["rol"]             = "empleado"
        st.session_state["correo"]          = correo
        st.session_state["rfc"]             = rfc_input.upper()
        st.session_state["nombre"]          = nombre_completo
        st.session_state["empleado_row"]    = emp_dict
        st.session_state["jefe_inmediato"]  = jefe_pdf_auto
        st.rerun()

# ─────────────────────────────────────────────
# VISTA EMPLEADO
# ─────────────────────────────────────────────
def mapear_emojis_estado(estado: str) -> str:
    est = str(estado).upper()
    if "AUTORIZADO" in est or "✅" in est: return "✅ AUTORIZADO"
    if "RECHAZADO"  in est or "🔴" in est: return "🔴 RECHAZADO"
    return "🟡 PENDIENTE"

# ── TAB RELOJ CHECADOR (solo admin) ──────────────
# ── Reloj checador: extraído a checador_module.py (ver motor v2.2) ──
try:
    import checador_module as _checador_mod
    _ERROR_CHECADOR = ""
except Exception as _e_chk:
    _checador_mod = None
    _ERROR_CHECADOR = str(_e_chk)

def render_checador():
    if _checador_mod is None:
        st.error(f"El módulo del checador no está disponible: {_ERROR_CHECADOR}")
        return
    _checador_mod.render_checador({
        "cargar_festivos": cargar_festivos,
        "festivos_en_periodo": festivos_en_periodo,
        "get_client": get_client,
        "cargar_justif_direccion": cargar_justif_direccion,
        "guardar_justif_direccion": guardar_justif_direccion,
        "eliminar_justif_direccion": eliminar_justif_direccion,
        "cargar_observaciones": cargar_observaciones,
        "guardar_observaciones": guardar_observaciones,
        "guardar_asistencia_mes": guardar_asistencia_mes,
        "DIRECTORA_NOMBRE": DIRECTORA_NOMBRE,
        "DIRECTORA_CARGO": DIRECTORA_CARGO,
    })




def vista_empleado():
    rfc      = st.session_state["rfc"]
    nombre   = st.session_state["nombre"]
    emp_data = st.session_state.get("empleado_row", {})

    st.markdown(f'### Sesión iniciada como: {nombre}')
    st.caption('Filiación: ' + rfc_oculto(rfc) + ' · ' + emp_data.get('PUESTO', ''))

    # Alerta de falta no justificada (solo si RH marcó avisar a este empleado)
    try:
        _asis = cargar_asistencia_mes()
        if not _asis.empty and "RFC" in _asis.columns:
            _mi = _asis[_asis["RFC"].astype(str).str.upper() == str(rfc).upper()]
            if not _mi.empty:
                _f = _mi.iloc[0]
                _nojust = int(_f.get("NO_JUSTIFICADAS", 0) or 0)
                if str(_f.get("AVISAR","")).upper() == "SI" and _nojust > 0:
                    _dias = _f.get("DIAS_FALTA","")
                    st.error(
                        f"⚠️ Tienes {_nojust} falta(s) sin justificar"
                        + (f" ({_dias})" if _dias else "")
                        + ". Acude a Recursos Humanos o presenta tu justificante."
                    )
    except Exception:
        pass

    # Alerta días económicos vencen 31 dic
    hoy      = date.today()
    fin_anio = date(hoy.year, 12, 31)
    dias_para_vencer = (fin_anio - hoy).days
    if dias_para_vencer <= 60:
        st.warning(f"⚠️ Tus días económicos disponibles vencen el 31 de diciembre ({dias_para_vencer} días restantes).")

    # ── Carga de datos ───────────────────────────
    solicitudes = cargar_solicitudes_eco()
    empleados   = cargar_empleados()
    incidencias = cargar_incidencias()
    horarios_df = cargar_horarios()

    emp_row          = empleados[empleados["RFC"].astype(str).str.upper() == rfc]
    dias_totales     = int(emp_row["DIAS TOTALES"].iloc[0])     if not emp_row.empty else 0
    dias_usados      = dias_economicos_usados(rfc, solicitudes)
    dias_disponibles = dias_totales - dias_usados
    horas_pases      = horas_pases_mes(rfc, incidencias)

    # ── Tablero de saldos ────────────────────────
    col1, _col2 = st.columns(2)
    col1.metric("📅 Días económicos disponibles", dias_disponibles, delta=f"-{dias_usados} ejercidos")

    st.divider()

    # ── Horario actual ───────────────────────────
    with st.expander("🕐 Mi horario registrado", expanded=False):
        emp_hor = horarios_df[horarios_df["RFC"].astype(str).str.upper().str.strip() == rfc.upper().strip()]
        if not emp_hor.empty:
            row_hor = emp_hor.iloc[0]
            resumen = []
            for dia, (col_e, col_s) in COLUMNAS_HORARIO.items():
                ent = row_hor.get(col_e, "")
                sal = row_hor.get(col_s, "")
                if ent:
                    resumen.append({"Día": dia, "Entrada": ent, "Salida": sal})
            if resumen:
                st.dataframe(pd.DataFrame(resumen), use_container_width=True, hide_index=True)
            else:
                st.caption("Sin horario registrado.")
        else:
            st.caption("No se encontró tu horario en el sistema.")

    # ── Mis solicitudes ──────────────────────────
    st.markdown("#### 📋 Mis solicitudes este mes")
    import pytz
    tz_mx  = pytz.timezone("America/Mexico_City")
    ahora  = datetime.now(pytz.utc).astimezone(tz_mx)
    # Solo incidencias del mes actual
    mis_inc_all = incidencias[incidencias["RFC"].astype(str).str.upper() == rfc].copy()
    if not mis_inc_all.empty:
        mis_inc_all["FECHA_DT"] = pd.to_datetime(mis_inc_all["FECHA_INICIO"], errors="coerce")
        mis_inc = mis_inc_all[(mis_inc_all["FECHA_DT"].dt.year == ahora.year) & (mis_inc_all["FECHA_DT"].dt.month == ahora.month)].copy()
    else:
        mis_inc = mis_inc_all
    # Solo días económicos del mes actual
    sol_hist_all = solicitudes[solicitudes["RFC"].astype(str).str.upper().str.strip() == rfc.upper().strip()].copy() if "RFC" in solicitudes.columns else pd.DataFrame()
    if not sol_hist_all.empty:
        col_fi_eco = next((c for c in sol_hist_all.columns if "Inicio" in c or "INICIO" in c), None)
        if col_fi_eco:
            sol_hist_all["FECHA_DT"] = pd.to_datetime(sol_hist_all[col_fi_eco], errors="coerce")
            sol_hist = sol_hist_all[
                (sol_hist_all["FECHA_DT"].dt.year == ahora.year) &
                (sol_hist_all["FECHA_DT"].dt.month == ahora.month)
            ].copy()
            # Mes sin solicitudes = tabla vacía. NO mostrar el histórico como
            # respaldo: hacía que en julio aparecieran solicitudes de junio.
        else:
            sol_hist = sol_hist_all
    else:
        sol_hist = sol_hist_all

    if not sol_hist.empty:
        sol_hist = sol_hist.rename(columns={
            "Tipo Permiso":     "TIPO",
            "Fecha Inicio":     "FECHA_INICIO",
            "Fecha Fin":        "FECHA_FIN",
            "Dias Solicitados": "DIAS",
            "Aprobado Por":     "AUTORIZADO_POR",
            "Fecha Registro":   "FECHA_SOLICITUD",
        })
        col_folio_sol = next((c for c in sol_hist.columns if "FOLIO" in c.upper()), None)
        if col_folio_sol and col_folio_sol != "FOLIO":
            sol_hist["FOLIO"] = sol_hist[col_folio_sol]
        elif "FOLIO" not in sol_hist.columns:
            sol_hist["FOLIO"] = "HISTÓRICO"
        sol_hist["HORAS_PASE"]         = ""
        sol_hist["ESTADO"]             = sol_hist["AUTORIZADO_POR"].apply(
            lambda x: "🔴 RECHAZADO" if str(x).strip().upper().startswith("RECHAZADO") else ("✅ AUTORIZADO" if str(x).strip() != "" else "🟡 PENDIENTE")
        )
        # Observaciones: texto después del — en Aprobado Por para rechazos
        sol_hist["OBSERVACIONES"] = sol_hist["AUTORIZADO_POR"].apply(
            lambda x: str(x).split("—",1)[1].strip() if str(x).upper().startswith("RECHAZADO") and "—" in str(x) else ""
        )
    if not mis_inc.empty:
        mis_inc["ESTADO"] = mis_inc["ESTADO"].apply(mapear_emojis_estado)

    cols_mostrar = ["FOLIO", "TIPO", "FECHA_INICIO", "FECHA_FIN", "DIAS", "HORAS_PASE", "ESTADO", "FECHA_AUTORIZACION", "OBSERVACIONES"]
    frames = []
    if not sol_hist.empty:
        frames.append(sol_hist[[c for c in cols_mostrar if c in sol_hist.columns]])
    if not mis_inc.empty:
        frames.append(mis_inc[[c for c in cols_mostrar if c in mis_inc.columns]])

    if not frames:
        st.info("No tienes solicitudes este mes.")
    else:
        df_consolidado = pd.concat(frames, ignore_index=True)
        st.dataframe(df_consolidado, use_container_width=True, hide_index=True)

        # ── Alertas de pendientes y rechazos ──────────
        # Pendientes
        col_estado = "ESTADO" if "ESTADO" in df_consolidado.columns else None
        if col_estado:
            pendientes_emp = df_consolidado[df_consolidado[col_estado].astype(str).str.contains("PENDIENTE", case=False)]
            if not pendientes_emp.empty:
                folios_pend = ", ".join(pendientes_emp["FOLIO"].astype(str).tolist())
                st.warning(f"⏳ Tienes {len(pendientes_emp)} solicitud(es) pendiente(s): **{folios_pend}**\n\n"
                           "Para que tu solicitud sea procesada debes **entregar el comprobante PDF firmado** "
                           "en el Área de Recursos Humanos (Administración de Personal DFC). "
                           "Si ya lo entregaste, espera a que el área lo valide.")
    st.divider()

    # ── Faltas informativas ──────────────────────
    asis = cargar_asistencia_mes()
    mi_asis = pd.DataFrame()
    if not asis.empty and "RFC" in asis.columns:
        mi_asis = asis[asis["RFC"].astype(str).str.upper() == str(rfc).upper()]

    # Título honesto: el contenido es del ÚLTIMO reporte procesado del checador,
    # que casi siempre es del mes anterior. Nunca decir "este mes" si no lo es.
    _titulo_faltas = "📊 Mis faltas del último período procesado"
    try:
        if not mi_asis.empty:
            _per = str(mi_asis.iloc[0].get("PERIODO", ""))
            _f_per = pd.to_datetime(_per.split("~")[0].strip(), errors="coerce")
            if not pd.isna(_f_per):
                _meses_es = {1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",
                             7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}
                _hoy_mx = datetime.now(pytz.timezone("America/Mexico_City"))
                if _f_per.year == _hoy_mx.year and _f_per.month == _hoy_mx.month:
                    _titulo_faltas = "📊 Mis faltas este mes"
                else:
                    _titulo_faltas = f"📊 Mis faltas — reporte de {_meses_es[_f_per.month]} {_f_per.year}"
    except Exception:
        pass
    with st.expander(_titulo_faltas, expanded=False):
        st.info("ℹ️ Este contador es informativo. Las faltas no justificadas las gestiona Recursos Humanos.")
        if mi_asis.empty:
            c1, c2, c3 = st.columns(3)
            c1.metric("Faltas totales",        0)
            c2.metric("Faltas justificadas",   0)
            c3.metric("Faltas no justificadas",0)
            st.caption("Aún no hay un reporte de asistencia procesado para ti.")
        else:
            fila = mi_asis.iloc[0]
            faltas_tot = int(fila.get("FALTAS", 0) or 0)
            justis     = int(fila.get("JUSTIFICADAS", 0) or 0)
            nojust     = int(fila.get("NO_JUSTIFICADAS", 0) or 0)
            c1, c2, c3 = st.columns(3)
            c1.metric("Faltas totales",        faltas_tot)
            c2.metric("Faltas justificadas",   justis)
            c3.metric("Faltas no justificadas",nojust)
            st.caption(f"Período: {fila.get('PERIODO','')}")
            if fila.get("DIAS_FALTA"):
                st.caption(f"Días que faltaste: {fila.get('DIAS_FALTA')}")


    st.divider()
    st.markdown("### Nueva solicitud")

    jefe_guardado = st.session_state.get("jefe_inmediato", "")
    if jefe_guardado:
        jefe_pdf = jefe_guardado
        st.caption(f"👤 Jefe inmediato: {jefe_guardado.split(chr(60))[0]}")
    else:
        try:
            usuarios_df = cargar_usuarios()
            jefes_unicos = sorted({
                str(j).strip() for j in usuarios_df.get("JEFE_INMEDIATO", [])
                if str(j).strip()
            })
        except Exception:
            jefes_unicos = []
        if jefes_unicos:
            jefe_sel = st.selectbox("👤 Jefe inmediato que autoriza", options=jefes_unicos)
            jefe_pdf = formato_jefe_pdf(jefe_sel)
        else:
            jefe_manual = st.text_input("👤 Jefe inmediato que autoriza (nombre — cargo)")
            jefe_pdf = formato_jefe_pdf(jefe_manual)

    tipo = st.selectbox(
        "Tipo de incidencia",
        options=list(TIPO_LABELS.keys()),
        format_func=lambda x: TIPO_LABELS[x],
    )

    # ── DÍA ECONÓMICO ──────────────────────────
    if tipo == "ECO":
        st.info(f"Tienes **{dias_disponibles}** día(s) económico(s) disponibles. Vencen el 31 de diciembre.")
        if dias_disponibles <= 0:
            st.warning("No tienes días económicos disponibles.")
            return
        # Modo de selección: rango o días sueltos
        modo = st.radio("Selección de fechas", ["Rango continuo", "Días sueltos"], horizontal=True)
        festivos = cargar_festivos()
        festivos_set = festivos_a_set(festivos)

        if modo == "Rango continuo":
            col1, col2 = st.columns(2)
            with col1:
                fi = st.date_input("Fecha inicio", value=date.today())
            with col2:
                ff = st.date_input("Fecha fin",    value=date.today())
            if ff < fi:
                st.error("La fecha fin no puede ser anterior a la fecha inicio.")
                return
            dias_hab = dias_habiles_entre(fi, ff, festivos)
            st.caption(f"Días hábiles solicitados: **{dias_hab}**")
            if dias_hab > dias_disponibles:
                st.error(f"Excedes tus días disponibles ({dias_disponibles}).")
                return
            fechas_seleccionadas = []
            d = fi
            while d <= ff:
                if d.weekday() < 5 and d not in festivos_set:
                    fechas_seleccionadas.append(d)
                d += timedelta(days=1)
            fi_final, ff_final = fi, ff
        else:
            st.caption("Selecciona cada día económico que necesitas tomar.")
            # Mostrar checkboxes para los próximos 60 días hábiles
            from datetime import timedelta as _td
            dias_opciones = []
            d = date.today()
            while len(dias_opciones) < 60:
                if d.weekday() < 5 and d not in festivos_set:
                    dias_opciones.append(d)
                d += _td(days=1)
            NOMBRES_DIA = ["Lun","Mar","Mié","Jue","Vie"]
            MESES = ["","Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
            fechas_seleccionadas = []
            cols_sel = st.columns(5)
            for idx, dia in enumerate(dias_opciones):
                label = f"{NOMBRES_DIA[dia.weekday()]} {dia.day}/{MESES[dia.month]}"
                if cols_sel[idx % 5].checkbox(label, key=f"eco_dia_{dia}"):
                    fechas_seleccionadas.append(dia)
            dias_hab = len(fechas_seleccionadas)
            if dias_hab > 0:
                st.caption(f"Días seleccionados: **{dias_hab}**")
            if dias_hab > dias_disponibles:
                st.error(f"Excedes tus días disponibles ({dias_disponibles}).")
                return
            fi_final = min(fechas_seleccionadas) if fechas_seleccionadas else date.today()
            ff_final = max(fechas_seleccionadas) if fechas_seleccionadas else date.today()

        motivo = st.text_area("Motivo (opcional)", max_chars=300)
        archivo_eco = st.file_uploader("Adjuntar documento de soporte (opcional)", type=["pdf","png","jpg","jpeg"], key="eco_anexo")
        tiene_anexo_eco = archivo_eco is not None

        # Construir motivo con fechas exactas
        if fechas_seleccionadas:
            fechas_str = ", ".join([f.strftime("%d/%m/%Y") for f in sorted(fechas_seleccionadas)])
            motivo_final = f"{motivo} | Fechas: {fechas_str}".strip(" |")
        else:
            motivo_final = motivo

        if dias_hab == 0:
            st.warning("Selecciona al menos un día.")
            return
        enviar_solicitud(rfc, nombre, tipo, fi_final, ff_final, dias_hab, 0.0, motivo_final, tiene_anexo_eco, incidencias, archivo_eco, jefe_inmediato=jefe_pdf)

    # ── PASES ────────────────────────────────────
    elif tipo in ["PSE", "PSR", "PEN"]:
        # El tipo ya define el subtipo
        subtipo = TIPO_LABELS[tipo]
        fecha   = st.date_input("Fecha del pase", value=date.today())

        col1, col2 = st.columns(2)
        hora_salida = hora_entrada = hora_retorno = None
        with col1:
            if tipo in ["PSE", "PSR"]:
                hora_salida = st.text_input("Hora de salida (HH:MM)", placeholder="08:37", max_chars=5)
                hora_salida = hora_salida.strip() if hora_salida else None
        with col2:
            if tipo == "PEN":
                hora_entrada = st.text_input("Hora de entrada (HH:MM)", placeholder="08:37", max_chars=5)
                hora_entrada = hora_entrada.strip() if hora_entrada else None
            elif tipo == "PSR":
                hora_retorno = st.text_input("Hora estimada de retorno (HH:MM)", placeholder="10:30", max_chars=5)
                hora_retorno = hora_retorno.strip() if hora_retorno else None

        # Calcular horas del pase
        horas_pase = 0.0
        if hora_salida and hora_retorno:
            try:
                sal  = datetime.combine(fecha, datetime.strptime(hora_salida.strip(),  "%H:%M").time())
                ret  = datetime.combine(fecha, datetime.strptime(hora_retorno.strip(), "%H:%M").time())
                diff = (ret - sal).total_seconds() / 3600
                horas_pase = round(max(diff, 0), 2)
                st.caption(f"Horas de ausencia: **{horas_pase}h**")
            except:
                st.warning("Formato inválido. Usa HH:MM (ej: 08:37)")
        elif hora_salida:
            try:
                horarios_df_pse = cargar_horarios()
                emp_hor_pse = horarios_df_pse[horarios_df_pse["RFC"].astype(str).str.upper().str.strip() == rfc.upper().strip()]
                nd_pse = ["LUN","MAR","MIE","JUE","VIE","SAB","DOM"][fecha.weekday()]
                salida_prog = str(emp_hor_pse.iloc[0].get(f"SALIDA_{nd_pse}","")).strip() if not emp_hor_pse.empty else ""
                if salida_prog:
                    sal_dt = datetime.combine(fecha, datetime.strptime(hora_salida.strip(), "%H:%M").time())
                    fin_dt = datetime.combine(fecha, datetime.strptime(salida_prog, "%H:%M").time())
                    horas_pase = round(max((fin_dt - sal_dt).total_seconds() / 3600, 0), 2)
                    st.caption(f"Horas de ausencia (hasta fin de jornada {salida_prog}): **{horas_pase}h**")
                else:
                    horas_pase = 0.0
            except:
                horas_pase = 0.0
                st.warning("No se pudo calcular automáticamente.")
        elif hora_entrada:
            # Pase de entrada: horas ausentes = desde hora programada hasta hora real de llegada
            try:
                horarios_df_pse = cargar_horarios()
                emp_hor_pse = horarios_df_pse[horarios_df_pse["RFC"].astype(str).str.upper().str.strip() == rfc.upper().strip()]
                nd_pse = ["LUN","MAR","MIE","JUE","VIE","SAB","DOM"][fecha.weekday()]
                entrada_prog = str(emp_hor_pse.iloc[0].get(f"ENTRADA_{nd_pse}","")).strip() if not emp_hor_pse.empty else ""
                if entrada_prog:
                    prog_dt  = datetime.combine(fecha, datetime.strptime(entrada_prog, "%H:%M").time())
                    real_dt  = datetime.combine(fecha, datetime.strptime(hora_entrada.strip(), "%H:%M").time())
                    horas_pase = round(max((real_dt - prog_dt).total_seconds() / 3600, 0), 2)
                    st.caption(f"Horas de ausencia (desde {entrada_prog} hasta {hora_entrada}): **{horas_pase}h**")
                else:
                    horas_pase = 0.0
            except:
                horas_pase = 0.0
                st.warning("No se pudo calcular automáticamente.")


        motivo = st.text_area("Motivo", max_chars=300)
        archivo_anexo = st.file_uploader("Adjuntar justificante (opcional)", type=["pdf","png","jpg","jpeg"])
        tiene_anexo   = archivo_anexo is not None
        detalle = subtipo
        if hora_salida:  detalle += f" | Salida: {hora_salida}"
        if hora_entrada: detalle += f" | Entrada: {hora_entrada}"
        if hora_retorno: detalle += f" | Retorno: {hora_retorno}"
        motivo_completo = (detalle + "\n" + motivo).strip()
        enviar_solicitud(rfc, nombre, tipo, fecha, fecha, 0, horas_pase, motivo_completo, tiene_anexo, incidencias, archivo_anexo, subtipo_label=subtipo, hora_retorno=hora_retorno or "", jefe_inmediato=jefe_pdf)

    # ── COMISIÓN ────────────────────────────────
    elif tipo == "COM":
        col1, col2 = st.columns(2)
        with col1:
            fi = st.date_input("Fecha inicio comisión", value=date.today())
        with col2:
            ff = st.date_input("Fecha fin comisión",    value=date.today())
        if ff < fi:
            st.error("La fecha fin no puede ser anterior a la fecha inicio.")
        else:
            festivos = cargar_festivos()
            dias_hab = dias_habiles_entre(fi, ff, festivos)
            st.caption(f"Días de comisión: **{dias_hab}**")
        motivo      = st.text_area("Motivo de la comisión", max_chars=300)
        archivo_anexo = st.file_uploader("Adjuntar constancia/oficio (opcional)", type=["pdf","png","jpg","jpeg"])
        tiene_anexo   = archivo_anexo is not None
        if not tiene_anexo:
            st.caption("Recuerda que sin anexo tu solicitud puede quedar sin soporte documental.")
        enviar_solicitud(rfc, nombre, tipo, fi, ff, dias_hab, 0.0,
                         motivo, tiene_anexo, incidencias, archivo_anexo, jefe_inmediato=jefe_pdf)
    # ── REPOSICIÓN DE GUARDIAS ──────────────────
    elif tipo == "RGU":
        st.info("Captura los días que repones a cambio de guardia(s) previamente realizada(s).")
        col1, col2 = st.columns(2)
        with col1:
            fi = st.date_input("Fecha inicio reposición", value=date.today())
        with col2:
            ff = st.date_input("Fecha fin reposición", value=date.today())
        dias_hab = 0
        if ff < fi:
            st.error("La fecha fin no puede ser anterior a la fecha inicio.")
        else:
            festivos = cargar_festivos()
            dias_hab = dias_habiles_entre(fi, ff, festivos)
            st.caption(f"Días a reponer: **{dias_hab}**")
        fecha_guardia = st.text_input(
            "Fecha(s) de la guardia que repones",
            placeholder="27/04/2026",
            help="Indica la fecha de la guardia realizada que compensas con estos días."
        )
        motivo = st.text_area("Observaciones (opcional)", max_chars=300)
        archivo_anexo = st.file_uploader("Adjuntar formato de guardia / soporte (opcional)", type=["pdf","png","jpg","jpeg"])
        tiene_anexo   = archivo_anexo is not None
        motivo_rgu = f"Reposición de guardia | Guardia repuesta: {fecha_guardia}".strip()
        if motivo:
            motivo_rgu += f" | {motivo}"
        if dias_hab == 0:
            st.warning("Selecciona al menos un día hábil para reponer.")
            return
        if not fecha_guardia.strip():
            st.warning("Indica la fecha de la guardia que estás reponiendo.")
            return
        enviar_solicitud(rfc, nombre, tipo, fi, ff, dias_hab, 0.0,
                         motivo_rgu, tiene_anexo, incidencias, archivo_anexo, jefe_inmediato=jefe_pdf)
    # ── CAMBIO DE HORARIO ───────────────────────
    elif tipo == "CHO":
        fecha_inicio_cho = st.date_input("¿A partir de qué fecha aplica el cambio?", value=date.today())
        motivo      = st.text_area("Motivo del cambio de horario", max_chars=300)
        tiene_anexo = st.checkbox("¿Traerás documento de soporte (oficio, etc.)?")
        st.markdown("**Horario solicitado:**")
        st.caption("Deja en blanco los días que no labora.")
        horario_solicitado = {}
        cols_dias = st.columns(7)
        for idx, dia in enumerate(DIAS_SEMANA):
            with cols_dias[idx]:
                st.caption(dia)
                ne = st.text_input("Entrada", value="", key=f"cho_e_{dia}", max_chars=5,
                                   placeholder="08:00")
                ns = st.text_input("Salida",  value="", key=f"cho_s_{dia}", max_chars=5,
                                   placeholder="16:00")
                horario_solicitado[dia] = {"entrada": ne.strip(), "salida": ns.strip()}
        # Serializar horario para guardarlo en MOTIVO
        horario_str = " | ".join(
            f"{dia} {v['entrada']}-{v['salida']}"
            for dia, v in horario_solicitado.items()
            if v["entrada"]
        )
        motivo_completo = f"Horario solicitado: {horario_str} | Motivo: {motivo}".strip()
        enviar_solicitud(rfc, nombre, tipo, fecha_inicio_cho, fecha_inicio_cho, 0, 0.0,
                         motivo_completo, tiene_anexo, incidencias, jefe_inmediato=jefe_pdf)

    # ── CUMPLEAÑOS (oculto hasta autorización) ──
    if HABILITAR_CUMPLEANOS and tipo == "CUM":
        cumple = cumpleanos_laboral(rfc)
        if cumple:
            st.info(f"Tu día de cumpleaños hábil es: **{cumple.strftime('%d/%m/%Y')}**")
            fi = cumple
            ff = cumple
            motivo = "Día de cumpleaños"
            st.caption("Solo el día hábil correspondiente a tu fecha de nacimiento.")
            enviar_solicitud(rfc, nombre, tipo, fi, ff, 1, 0.0, motivo, False, incidencias, jefe_inmediato=jefe_pdf)
        else:
            st.error("No se pudo calcular tu fecha de cumpleaños desde el RFC.")


def enviar_solicitud(rfc, nombre, tipo, fi, ff, dias, horas_pase, motivo, tiene_anexo, incidencias_df, archivo_anexo=None, subtipo_label=None, hora_retorno="", jefe_inmediato=""):
    key_btn = f"btn_registrar_{tipo}_{str(fi)}_{str(ff)}"
    if st.button("Registrar solicitud", type="primary", key=key_btn):
        if st.session_state.get(f"registrado_{key_btn}"):
            st.warning("Esta solicitud ya fue registrada.")
            return

        # Validar duplicado: mismo tipo y misma fecha
        if not incidencias_df.empty:
            dup = incidencias_df[
                (incidencias_df["RFC"].astype(str).str.upper() == rfc.upper()) &
                (incidencias_df["TIPO"].astype(str) == tipo) &
                (incidencias_df["FECHA_INICIO"].astype(str) == str(fi))
            ]
            if not dup.empty:
                folio_dup = dup.iloc[0]["FOLIO"]
                tipo_label_dup = TIPO_LABELS.get(tipo, tipo)
                st.error(f"⚠️ Ya tienes una solicitud **{folio_dup}** de **{tipo_label_dup}** registrada para el **{fi.strftime('%d/%m/%Y')}**. No puedes registrar dos solicitudes del mismo tipo en la misma fecha.")
                return
        # Validar duplicado en ECO
        if tipo == "ECO":
            try:
                sol_eco = cargar_solicitudes_eco()
                if not sol_eco.empty:
                    col_rfc_eco = next((c for c in sol_eco.columns if c.upper() == "RFC"), None)
                    col_fi_eco  = next((c for c in sol_eco.columns if "INICIO" in c.upper()), None)
                    if col_rfc_eco and col_fi_eco:
                        sol_eco["_FDT"] = pd.to_datetime(sol_eco[col_fi_eco], errors="coerce")
                        dup_eco = sol_eco[
                            (sol_eco[col_rfc_eco].astype(str).str.upper() == rfc.upper()) &
                            (sol_eco["_FDT"].dt.date == fi)
                        ]
                        if not dup_eco.empty:
                            col_f = next((c for c in sol_eco.columns if "FOLIO" in c.upper()), None)
                            folio_dup = str(dup_eco.iloc[0].get(col_f, "ECO")) if col_f else "ECO"
                            st.error(f"⚠️ Ya tienes una solicitud **{folio_dup}** de **Día económico** registrada para el **{fi.strftime('%d/%m/%Y')}**. No puedes registrar dos días económicos para la misma fecha.")
                            return
            except: pass
        st.session_state[f"registrado_{key_btn}"] = True
        folio     = generar_folio(tipo, incidencias_df)
        import pytz
        tz_mx = pytz.timezone("America/Mexico_City")
        fecha_sol = datetime.now(pytz.utc).astimezone(tz_mx).strftime("%Y-%m-%d %H:%M")
        link_anexo = ""
        if archivo_anexo is not None:
            with st.spinner("Subiendo anexo a Drive..."):
                link_anexo = subir_anexo_drive(archivo_anexo, folio, rfc)
        datos = {
            "folio":           folio,
            "rfc":             rfc,
            "nombre":          nombre,
            "tipo":            tipo,
            "tipo_label":      TIPO_LABELS[tipo],
            "fecha_solicitud": fecha_sol,
            "fecha_inicio":    str(fi),
            "fecha_fin":       str(ff),
            "dias":            dias,
            "horas_pase":      horas_pase,
            "motivo":          motivo,
            "tiene_anexo":     tiene_anexo or archivo_anexo is not None,
            "link_anexo":      link_anexo,
            "subtipo_label":   subtipo_label if subtipo_label else TIPO_LABELS[tipo],
            "jefe_inmediato":   jefe_inmediato,
            "director_general": st.session_state.get("director_general", ""),
            "hora_retorno":    hora_retorno,
        }
        _guardado_ok = False
        try:
            if tipo == "ECO":
                _guardado_ok = guardar_dia_economico(datos)
            else:
                _guardado_ok = guardar_incidencia(datos)
        except Exception as _e_save:
            _error_amable(_e_save, "al registrar la solicitud")
        if _guardado_ok:
            folio = datos["folio"]  # puede haberse recalculado en vivo al guardar
            pdf_bytes = generar_comprobante_pdf(datos)
            st.success(f"✅ Solicitud registrada con folio **{folio}**")
            st.download_button(
                label="⬇️ Descargar comprobante PDF",
                data=pdf_bytes,
                file_name=f"Comprobante_{folio}.pdf",
                mime="application/pdf",
            )
            st.info("Imprime tu comprobante y preséntalo en RH junto con el documento físico original.")

# ─────────────────────────────────────────────
# VISTA ADMIN
# ─────────────────────────────────────────────
def vista_admin():
    st.markdown("### Panel de administración · Incidencias")

    incidencias = cargar_incidencias()
    horarios_df = cargar_horarios()

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Pendientes", "Historial completo", "Reporte mensual", "🕐 Reloj Checador", "📋 Nómina"])

    with tab1:
        with st.expander("🔄 Sincronizar saldos de días económicos al Sheet"):
            st.caption("Recalcula DIAS DISPONIBLES (= totales − usados aprobados) y lo escribe "
                       "en la hoja Empleados para todos. Útil para dejar el Sheet al día.")
            if st.button("Actualizar todos los saldos ahora"):
                with st.spinner("Recalculando y escribiendo en el Sheet..."):
                    try:
                        n, errs = actualizar_dias_disponibles_sheet()
                        if errs:
                            st.error(" · ".join(errs))
                        else:
                            st.success(f"✅ {n} saldo(s) actualizado(s) en la hoja Empleados.")
                    except Exception as e:
                        st.error(f"No se pudo actualizar: {e}")

        solicitudes_eco = cargar_solicitudes_eco()
        col_aprobado = next((c for c in solicitudes_eco.columns if "Aprobado" in c), None)
        col_folio    = next((c for c in solicitudes_eco.columns if "FOLIO" in c.upper() or "Folio" in c), None)
        eco_pend = pd.DataFrame()
        if col_aprobado and not solicitudes_eco.empty:
            eco_pend = solicitudes_eco[solicitudes_eco[col_aprobado].astype(str).str.strip() == ""]

        pendientes = incidencias[incidencias["ESTADO"] == "PENDIENTE"]
        total_pend = len(pendientes) + len(eco_pend)

        if total_pend == 0:
            st.success("No hay solicitudes pendientes.")
        else:
            st.caption(f"{total_pend} solicitud(es) pendiente(s)")

            # Días económicos primero
            for idx, row in eco_pend.iterrows():
                folio_eco  = str(row.get(col_folio, "")) if col_folio else f"ECO-{idx}"
                nombre_eco = str(row.get("Nombre Completo", row.get("NOMBRE", "")))
                with st.expander(f"**{folio_eco}** · Día económico · {nombre_eco}"):
                    col1, col2 = st.columns(2)
                    col1.write(f"**RFC:** {str(row.get('RFC',''))}")
                    col1.write(f"**Fecha inicio:** {str(row.get('Fecha Inicio',''))}")
                    col1.write(f"**Fecha fin:** {str(row.get('Fecha Fin',''))}")
                    col1.write(f"**Días:** {str(row.get('Dias Solicitados',''))}")
                    col2.write(f"**Motivo:** {str(row.get('Motivo',''))}")
                    col2.write(f"**Registrado:** {str(row.get('Fecha Registro',''))}")
                    obs_eco = st.text_input("Observaciones (para rechazo)", key=f"obs_eco_{idx}")
                    col_a, col_r = st.columns(2)
                    _rfc_eco   = str(row.get("RFC", ""))
                    _fi_eco    = str(row.get("Fecha Inicio", ""))
                    _freg_eco  = str(row.get("Fecha Registro", ""))
                    _folio_real = str(row.get(col_folio, "")) if col_folio else ""
                    with col_a:
                        if st.button("✅ Aprobar", key=f"eco_{idx}", type="primary"):
                            aprobar_dia_economico(st.session_state.get("nombre", "admin"),
                                                  folio=_folio_real, rfc=_rfc_eco,
                                                  fecha_inicio=_fi_eco, fecha_registro=_freg_eco)
                            st.rerun()
                    with col_r:
                        if st.button("❌ Rechazar", key=f"rec_eco_{idx}"):
                            if not obs_eco:
                                st.error("Escribe una observación para rechazar.")
                            else:
                                rechazar_dia_economico(obs_eco,
                                                       folio=_folio_real, rfc=_rfc_eco,
                                                       fecha_inicio=_fi_eco, fecha_registro=_freg_eco)
                                st.rerun()
            for _, row in pendientes.iterrows():
                with st.expander(f"**{row['FOLIO']}** · {TIPO_LABELS.get(row['TIPO'], row['TIPO'])} · {row['NOMBRE']}"):
                    col1, col2 = st.columns(2)
                    col1.write('**Filiación:** ' + rfc_oculto(str(row['RFC'])))
                    col1.write(f"**Fechas:** {row['FECHA_INICIO']} → {row['FECHA_FIN']}")
                    col1.write(f"**Días:** {row['DIAS']}")
                    col1.write(f"**Horas pase:** {row.get('HORAS_PASE','')}")
                    col2.write(f"**Motivo:** {row['MOTIVO']}")
                    col2.write(f"**Anexo:** {row['TIENE_ANEXO']}")
                    col2.write(f"**Registrado:** {row['FECHA_SOLICITUD']}")

                    obs = st.text_input("Observaciones", key=f"obs_{row['FOLIO']}")
                    col_a, col_r = st.columns(2)
                    with col_a:
                        if st.button("✅ Autorizar", key=f"aut_{row['FOLIO']}", type="primary"):
                            if row["TIPO"] == "CHO":
                                st.warning("Para cambio de horario usa el editor de horario abajo.")
                            else:
                                autorizar_incidencia(row["FOLIO"], obs)
                                st.success(f"Folio {row['FOLIO']} autorizado.")
                                st.rerun()
                    with col_r:
                        if st.button("❌ Rechazar", key=f"rec_{row['FOLIO']}"):
                            if not obs:
                                st.error("Escribe una observación para rechazar.")
                            else:
                                rechazar_incidencia(row["FOLIO"], obs)
                                st.warning(f"Folio {row['FOLIO']} rechazado.")
                                st.rerun()

                    if row["TIPO"] == "CHO":
                        st.divider()
                        st.markdown("**Ajustar horario en el sistema:**")
                        emp_hor = horarios_df[horarios_df["RFC"].astype(str).str.upper().str.strip() == str(row["RFC"]).upper().strip()]

                        # Parsear horario solicitado del MOTIVO
                        motivo_cho = str(row.get("MOTIVO", ""))
                        horario_parsed = {}
                        for parte in motivo_cho.split(" | "):
                            for dia in DIAS_SEMANA:
                                if parte.strip().startswith(dia + " "):
                                    try:
                                        horas = parte.strip()[len(dia)+1:]
                                        e, s = horas.split("-")
                                        horario_parsed[dia] = {"entrada": e.strip(), "salida": s.strip()}
                                    except Exception:
                                        pass

                        nuevo_horario = {}
                        cols_dias = st.columns(5)
                        for idx, dia in enumerate(DIAS_SEMANA):
                            col_e, col_s = COLUMNAS_HORARIO[dia]
                            # Prellenar con horario solicitado, si no con el actual del Sheet
                            if dia in horario_parsed:
                                default_e = horario_parsed[dia]["entrada"]
                                default_s = horario_parsed[dia]["salida"]
                            elif not emp_hor.empty:
                                default_e = str(emp_hor.iloc[0].get(col_e, ""))
                                default_s = str(emp_hor.iloc[0].get(col_s, ""))
                            else:
                                default_e = ""
                                default_s = ""
                            with cols_dias[idx]:
                                st.caption(dia)
                                ne = st.text_input("Entrada", value=default_e, key=f"ne_{row['FOLIO']}_{dia}", max_chars=5)
                                ns = st.text_input("Salida",  value=default_s, key=f"ns_{row['FOLIO']}_{dia}", max_chars=5)
                                nuevo_horario[dia] = {"entrada": ne, "salida": ns}
                        if st.button("💾 Guardar horario y autorizar", key=f"hor_{row['FOLIO']}", type="primary"):
                            autorizar_cambio_horario(row["RFC"], nuevo_horario, row["FOLIO"])
                            st.success("Horario actualizado y solicitud autorizada.")
                            st.rerun()

    with tab2:
        sol_eco_hist = cargar_solicitudes_eco()
        # Normalizar ECO para unir con incidencias
        df_eco_hist = pd.DataFrame()
        if not sol_eco_hist.empty:
            df_eco_hist = sol_eco_hist.rename(columns={
                "Tipo Permiso":"TIPO","Fecha Inicio":"FECHA_INICIO","Fecha Fin":"FECHA_FIN",
                "Dias Solicitados":"DIAS","Aprobado Por":"AUTORIZADO_POR",
                "Nombre Completo":"NOMBRE","Fecha Registro":"FECHA_SOLICITUD"
            }).copy()
            col_f = next((c for c in df_eco_hist.columns if "FOLIO" in c.upper()), None)
            if col_f and col_f != "FOLIO": df_eco_hist["FOLIO"] = df_eco_hist[col_f]
            df_eco_hist["TIPO"] = "ECO"
            df_eco_hist["ESTADO"] = df_eco_hist["AUTORIZADO_POR"].apply(
                lambda x: "🔴 RECHAZADO" if str(x).strip().upper().startswith("RECHAZADO") else ("✅ AUTORIZADO" if str(x).strip() else "🟡 PENDIENTE")
            )
            df_eco_hist["ID"] = ""
            df_eco_hist["HORAS_PASE"] = ""

        df_hist_completo = pd.concat([incidencias, df_eco_hist[[c for c in incidencias.columns if c in df_eco_hist.columns]]], ignore_index=True) if not df_eco_hist.empty else incidencias.copy()

        filtro_tipo   = st.selectbox("Filtrar por tipo", ["TODOS"] + list(TIPO_LABELS.keys()))
        filtro_estado = st.selectbox("Filtrar por estado", ["TODOS", "PENDIENTE", "AUTORIZADO", "RECHAZADO"])
        df_hist = df_hist_completo.copy()
        if filtro_tipo   != "TODOS": df_hist = df_hist[df_hist["TIPO"] == filtro_tipo]
        if filtro_estado != "TODOS": df_hist = df_hist[df_hist["ESTADO"].astype(str).str.contains(filtro_estado, case=False)]
        st.dataframe(df_hist, use_container_width=True, hide_index=True)

        # ── Export Excel padrón mensual ──────────────
        import pytz as _pytz
        _tz = _pytz.timezone("America/Mexico_City")
        _ahora = datetime.now(_pytz.utc).astimezone(_tz)
        mes_sel = st.selectbox("📅 Mes para exportar padrón", 
            options=list(range(1,13)),
            format_func=lambda m: ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][m],
            index=_ahora.month-1)

        if st.button("📊 Generar padrón Excel del mes", type="primary"):
            import io as _io
            from openpyxl import Workbook as _WB
            from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align
            sol_eco = cargar_solicitudes_eco()
            # Filtrar incidencias del mes
            df_mes = incidencias.copy()
            df_mes["_FDT"] = pd.to_datetime(df_mes["FECHA_INICIO"], errors="coerce")
            df_mes = df_mes[(df_mes["_FDT"].dt.year == _ahora.year) & (df_mes["_FDT"].dt.month == mes_sel)]
            # Filtrar ECO del mes
            df_eco = sol_eco.copy() if not sol_eco.empty else pd.DataFrame()
            if not df_eco.empty:
                col_fi = next((c for c in df_eco.columns if "Inicio" in c or "INICIO" in c), None)
                if col_fi:
                    df_eco["_FDT"] = pd.to_datetime(df_eco[col_fi], errors="coerce")
                    df_eco = df_eco[(df_eco["_FDT"].dt.year == _ahora.year) & (df_eco["_FDT"].dt.month == mes_sel)]
                    df_eco = df_eco.rename(columns={
                        "Tipo Permiso":"TIPO","Fecha Inicio":"FECHA_INICIO","Fecha Fin":"FECHA_FIN",
                        "Dias Solicitados":"DIAS","Aprobado Por":"AUTORIZADO_POR",
                        "Nombre Completo":"NOMBRE","Fecha Registro":"FECHA_SOLICITUD"
                    })
                    col_f = next((c for c in df_eco.columns if "FOLIO" in c.upper()), None)
                    if col_f: df_eco["FOLIO"] = df_eco[col_f]
            # Unir
            cols_exp = ["FOLIO","NOMBRE","RFC","TIPO","FECHA_INICIO","FECHA_FIN","DIAS","HORAS_PASE","MOTIVO","ESTADO","AUTORIZADO_POR","FECHA_AUTORIZACION"]
            frames_exp = []
            if not df_mes.empty: frames_exp.append(df_mes[[c for c in cols_exp if c in df_mes.columns]])
            if not df_eco.empty: frames_exp.append(df_eco[[c for c in cols_exp if c in df_eco.columns]])
            if frames_exp:
                df_export = pd.concat(frames_exp, ignore_index=True)
                df_export["TIPO"] = df_export["TIPO"].apply(lambda x: "ECO" if str(x).lower() in ["economico","económico"] else x)
                df_export["TIPO"] = df_export["TIPO"].map(TIPO_LABELS).fillna(df_export["TIPO"])
                df_export = df_export.sort_values("NOMBRE")
                wb = _WB()
                ws = wb.active
                ws.title = "Padrón Solicitudes"
                mes_nombre = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][mes_sel]
                ws.merge_cells("A1:L1")
                ws["A1"] = f"PADRÓN DE SOLICITUDES — {mes_nombre.upper()} {_ahora.year} · DFC · SEJ"
                ws["A1"].font = _Font(bold=True, size=12, color="002F6C")
                ws["A1"].alignment = _Align(horizontal="center")
                hdrs = ["Folio","Nombre","RFC","Tipo","Fecha Inicio","Fecha Fin","Días","Horas Pase","Motivo","Estado","Autorizado Por","Fecha Autorización"]
                ws.append(hdrs)
                for cell in ws[2]: cell.font = _Font(bold=True, color="FFFFFF"); cell.fill = _Fill("solid", fgColor="002F6C")
                for _, row in df_export.iterrows():
                    ws.append([str(row.get(c,"")) for c in ["FOLIO","NOMBRE","RFC","TIPO","FECHA_INICIO","FECHA_FIN","DIAS","HORAS_PASE","MOTIVO","ESTADO","AUTORIZADO_POR","FECHA_AUTORIZACION"]])
                ws.column_dimensions["A"].width = 18
                ws.column_dimensions["B"].width = 35
                ws.column_dimensions["I"].width = 40
                buf_xl = _io.BytesIO()
                wb.save(buf_xl)
                st.download_button(f"⬇️ Descargar padrón {mes_nombre} {_ahora.year}",
                    data=buf_xl.getvalue(),
                    file_name=f"Padron_Solicitudes_{mes_nombre}_{_ahora.year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary")
            else:
                st.info(f"No hay solicitudes registradas para ese mes.")

        st.divider()
        st.markdown("#### 🔍 Regenerar PDF por folio")
        folio_buscar = st.text_input("Ingresa el folio", placeholder="ECO-2026-0001 / PSE-2026-0001", key="folio_regen")
        if folio_buscar:
            folio_buscar = folio_buscar.strip().upper()
            # Buscar en Incidencias
            match_inc = incidencias[incidencias["FOLIO"].astype(str).str.upper() == folio_buscar]
            # Buscar en Solicitudes ECO
            sol_eco = cargar_solicitudes_eco()
            col_f = next((c for c in sol_eco.columns if "FOLIO" in c.upper()), None)
            match_eco = sol_eco[sol_eco[col_f].astype(str).str.upper() == folio_buscar] if col_f else pd.DataFrame()
            if not match_inc.empty:
                row = match_inc.iloc[0]
                # Detectar subtipo real desde el motivo para registros viejos PSE
                motivo_str = str(row["MOTIVO"])
                tipo_real  = row["TIPO"]
                if tipo_real == "PSE":
                    if "Pase de entrada" in motivo_str:
                        tipo_real = "PEN"
                    elif "sin retorno" in motivo_str.lower() or "Salida" in motivo_str:
                        tipo_real = "PSE"
                    elif "Retorno" in motivo_str:
                        tipo_real = "PSR"
                # Buscar jefe inmediato del empleado en tab Usuarios
                jefe_pdf_regen = ""
                try:
                    usuarios_df = cargar_usuarios()
                    match_usr = usuarios_df[usuarios_df["RFC"].astype(str).str.upper().str.strip() == str(row["RFC"]).upper().strip()]
                    if not match_usr.empty:
                        jefe_col = match_usr.iloc[0].get("JEFE_INMEDIATO", "")
                        jefe_pdf_regen = formato_jefe_pdf(str(jefe_col).strip())
                except: pass
                datos_pdf = {
                    "folio":          row["FOLIO"],
                    "rfc":            row["RFC"],
                    "nombre":         row["NOMBRE"],
                    "tipo":           tipo_real,
                    "tipo_label":     TIPO_LABELS.get(tipo_real, tipo_real),
                    "subtipo_label":  TIPO_LABELS.get(tipo_real, tipo_real),
                    "fecha_solicitud":str(row["FECHA_SOLICITUD"]),
                    "fecha_inicio":   str(row["FECHA_INICIO"]),
                    "fecha_fin":      str(row["FECHA_FIN"]),
                    "dias":           row["DIAS"],
                    "horas_pase":     row.get("HORAS_PASE", 0),
                    "motivo":         motivo_str,
                    "tiene_anexo":    row.get("TIENE_ANEXO", "NO") == "SÍ",
                    "link_anexo":     row.get("LINK_ANEXO", ""),
                    "estado":         str(row["ESTADO"]),
                    "jefe_inmediato": jefe_pdf_regen,
                }
                st.success(f"Folio encontrado: {folio_buscar}")
                pdf_bytes = generar_comprobante_pdf(datos_pdf)
                st.download_button("⬇️ Descargar PDF", data=pdf_bytes,
                    file_name=f"Comprobante_{folio_buscar}.pdf", mime="application/pdf", type="primary")
            elif not match_eco.empty:
                row = match_eco.iloc[0]
                datos_pdf = {
                    "folio":          str(row.get(col_f, folio_buscar)),
                    "rfc":            str(row.get("RFC", "")),
                    "nombre":         str(row.get("Nombre Completo", "")),
                    "tipo":           "ECO",
                    "tipo_label":     "Día económico",
                    "subtipo_label":  "Día económico",
                    "fecha_solicitud":str(row.get("Fecha Registro", "")),
                    "fecha_inicio":   str(row.get("Fecha Inicio", "")),
                    "fecha_fin":      str(row.get("Fecha Fin", "")),
                    "dias":           row.get("Dias Solicitados", 0),
                    "horas_pase":     0,
                    "motivo":         str(row.get("Motivo", "")),
                    "tiene_anexo":    False,
                    "link_anexo":     str(row.get("LINK_ANEXO", "")),
                    "estado":         "AUTORIZADO" if str(row.get("Aprobado Por","")).strip() else "PENDIENTE",
                    "jefe_inmediato": str(row.get("Aprobado Por", "")),
                }
                st.success(f"Folio encontrado: {folio_buscar}")
                pdf_bytes = generar_comprobante_pdf(datos_pdf)
                st.download_button("⬇️ Descargar PDF", data=pdf_bytes,
                    file_name=f"Comprobante_{folio_buscar}.pdf", mime="application/pdf", type="primary")
            else:
                st.error(f"Folio {folio_buscar} no encontrado.")


    with tab4:
        render_checador()

    with tab5:
        if render_pendientes_nomina is None:
            st.error(f"El módulo de nómina no está disponible: {_ERROR_NOMINA}")
        else:
            render_pendientes_nomina(cargar_directorio_nomina)

    with tab3:
        # ── Alerta días económicos por agotarse ──────
        st.markdown("#### ⚠️ Empleados con 5 días económicos o menos")
        try:
            import pytz as _ptz
            _tz   = _ptz.timezone("America/Mexico_City")
            _anio = datetime.now(_ptz.utc).astimezone(_tz).year
            sol_eco = cargar_solicitudes_eco()
            if sol_eco.empty:
                st.info("No hay solicitudes registradas.")
            else:
                col_rfc  = next((c for c in sol_eco.columns if c.upper() == "RFC"), None)
                col_nom  = next((c for c in sol_eco.columns if "NOMBRE" in c.upper()), None)
                col_dias = next((c for c in sol_eco.columns if "DIAS" in c.upper() or "SOLICIT" in c.upper()), None)
                col_apr  = next((c for c in sol_eco.columns if "APROBADO" in c.upper()), None)
                col_fi   = next((c for c in sol_eco.columns if "INICIO" in c.upper()), None)
                if col_rfc and col_dias and col_fi:
                    sol_eco["_FDT"] = pd.to_datetime(sol_eco[col_fi], errors="coerce")
                    sol_anio = sol_eco[sol_eco["_FDT"].dt.year == _anio].copy()
                    # Excluir rechazados
                    if col_apr:
                        sol_anio = sol_anio[~sol_anio[col_apr].astype(str).str.strip().str.upper().str.startswith("RECHAZADO")]
                    sol_anio["_DIAS"] = pd.to_numeric(sol_anio[col_dias], errors="coerce").fillna(0).astype(int)
                    resumen = sol_anio.groupby(col_rfc).agg(
                        Usados=("_DIAS", "sum")
                    ).reset_index()
                    resumen.rename(columns={col_rfc: "RFC"}, inplace=True)
                    if col_nom:
                        nombres = sol_anio.groupby(col_rfc)[col_nom].first().reset_index()
                        nombres.rename(columns={col_rfc:"RFC", col_nom:"Nombre"}, inplace=True)
                        resumen = resumen.merge(nombres, on="RFC", how="left")
                    resumen["Disponibles"] = (9 - resumen["Usados"]).clip(lower=0)
                    alertas = resumen[resumen["Disponibles"] <= 5].sort_values("Disponibles")
                    cols_show = ["Nombre","Usados","Disponibles"] if col_nom else ["RFC","Usados","Disponibles"]
                    if alertas.empty:
                        st.success("Ningún empleado tiene 5 días o menos disponibles.")
                    else:
                        st.dataframe(alertas[[c for c in cols_show if c in alertas.columns]], use_container_width=True, hide_index=True)
                    st.warning("No se encontraron las columnas necesarias en Solicitudes.")
        except Exception as e:
            st.warning(f"No se pudo calcular: {e}")
        st.divider()
        st.markdown("Reporte consolidado para la directora.")
        anio_rep = st.number_input("Año",  value=datetime.now().year, step=1)
        mes_rep  = st.selectbox("Mes", range(1, 13), index=datetime.now().month - 1,
                                format_func=lambda m: ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                                                        "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][m-1])
        if st.button("Generar reporte", type="primary"):
            df_rep = incidencias.copy()
            df_rep["FECHA_DT"] = pd.to_datetime(df_rep["FECHA_SOLICITUD"], errors="coerce")
            df_rep = df_rep[(df_rep["FECHA_DT"].dt.year == anio_rep) & (df_rep["FECHA_DT"].dt.month == mes_rep)]
            if df_rep.empty:
                st.info("No hay incidencias en ese período.")
            else:
                mes_nombre = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto",
                              "Septiembre","Octubre","Noviembre","Diciembre"][mes_rep-1]
                st.markdown(f"**Resumen {mes_nombre} {anio_rep}**")
                st.dataframe(df_rep.groupby(["TIPO","ESTADO"]).size().reset_index(name="Total"),
                             use_container_width=True, hide_index=True)
                detalle = df_rep.groupby(["NOMBRE","TIPO"]).agg(
                    Total=("FOLIO","count"),
                    Dias=("DIAS","sum"),
                    Horas_pase=("HORAS_PASE","sum"),
                    Autorizadas=("ESTADO", lambda x: (x=="AUTORIZADO").sum()),
                    Rechazadas=("ESTADO",  lambda x: (x=="RECHAZADO").sum()),
                ).reset_index()
                st.dataframe(detalle, use_container_width=True, hide_index=True)
                st.download_button("⬇️ Descargar CSV",
                                   data=df_rep.to_csv(index=False).encode("utf-8"),
                                   file_name=f"Reporte_{mes_nombre}_{anio_rep}.csv",
                                   mime="text/csv")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def vista_calendario():
    st.markdown("## 📅 Calendario de Pagos y Prestaciones 2026")
    st.markdown("""
    <div style='display:flex;gap:16px;flex-wrap:wrap;margin-bottom:12px;font-size:12px'>
        <span>🟡 Docentes Básica</span>
        <span>🟣 Docentes Nivel Superior</span>
        <span>🟢 Personal de Apoyo Básica</span>
        <span>🔵 Personal de Apoyo No Docente Nivel Superior</span>
    </div>""", unsafe_allow_html=True)

    PAGOS = [
        ("Enero", [
            ("14 Ene", "Q-01", [("Estímulo puntualidad y asistencia 2ª parte", "🟣🔵"), ("Prima Dominical", "🟣🔵")]),
            ("29 Ene", "Q-02", [("Compensación Nacional Única 1ª Parte", "🟡🟢")]),
        ]),
        ("Febrero", [
            ("12 Feb", "Q-03", [("Sueldo ordinario", "")]),
            ("26 Feb", "Q-04", [("Sueldo ordinario", "")]),
        ]),
        ("Marzo", [
            ("12 Mar", "Q-05", [("Sueldo ordinario", "")]),
            ("26 Mar", "Q-06", [("1ª Parte Aguinaldo", "🟢🔵")]),
        ]),
        ("Abril", [
            ("14 Abr", "Q-07", [("Sueldo ordinario", "")]),
            ("29 Abr", "Q-08", [("Sueldo ordinario", "")]),
        ]),
        ("Mayo", [
            ("14 May", "Q-09", [
                ("1ª Parte Aguinaldo", "🟡"),
                ("Gratificación Día del Maestro", "🟡"),
                ("1ª Parte Aguinaldo", "🟣"),
                ("Reconocimiento Docentes Nivel Superior", "🟣"),
                ("Ayuda para Libros", "🟣"),
            ]),
            ("28 May", "Q-10", [("Sueldo ordinario", "")]),
        ]),
        ("Junio", [
            ("12 Jun", "Q-11", [("Estímulo puntualidad y asistencia 1ª parte", "🟣🔵")]),
            ("29 Jun", "Q-12", [("Reconocimiento a Directores", "🟡")]),
        ]),
        ("Julio", [
            ("14 Jul", "Q-13", [("Sueldo ordinario", "")]),
            ("30 Jul", "Q-14", [("Gratificación por el trabajo", "🟢")]),
        ]),
        ("Agosto", [
            ("13 Ago", "Q-15", [("Organización Escolar", "🟡"), ("Ayuda para gastos escolares", "🟢")]),
            ("28 Ago", "Q-16", [("Compensación Nacional Única 2ª Parte", "🟡🟢"), ("Medida Económica Única", "🟣🔵")]),
        ]),
        ("Septiembre", [
            ("14 Sep", "Q-17", [
                ("Estímulo a la Actividad Docente", "🟡"),
                ("Estímulo a Directores", "🟡"),
                ("Gratificación Única 1ª Parte", "🔵"),
            ]),
            ("29 Sep", "Q-18", [
                ("Gratificación Fortalecimiento Académico", "🟡"),
                ("Bono Extraordinario superación académica 1ª Parte", "🟣"),
            ]),
        ]),
        ("Octubre", [
            ("14 Oct", "Q-19", [("Sueldo ordinario", "")]),
            ("29 Oct", "Q-20", [("Fortalecimiento CC según ajustes salariales", "🟡"), ("Fortalecimiento CT según ajustes salariales", "🟢")]),
        ]),
        ("Noviembre", [
            ("12 Nov", "Q-21", [("Sueldo ordinario", "")]),
            ("27 Nov", "Q-22", [("Bono anual 24 días inicial", "🟡"), ("Apoyo a la integración educativa especial", "🟡")]),
        ]),
        ("Diciembre", [
            ("Por definir", "", [("Fecha de pago por definir", "")]),
        ]),
    ]

    import pytz
    from datetime import datetime
    tz_mx = pytz.timezone("America/Mexico_City")
    mes_actual = datetime.now(tz_mx).month
    MESES_NUM = {"Enero":1,"Febrero":2,"Marzo":3,"Abril":4,"Mayo":5,"Junio":6,
                 "Julio":7,"Agosto":8,"Septiembre":9,"Octubre":10,"Noviembre":11,"Diciembre":12}

    def render_mes(col, mes, quincenas, es_actual):
        borde = "#F97316" if es_actual else "var(--color-border-tertiary)"
        fondo = "#FFF7ED" if es_actual else "var(--color-background-primary)"
        col.markdown(f"<div style='border:2px solid {borde};border-radius:12px;padding:12px;margin-bottom:12px;background:{fondo}'>", unsafe_allow_html=True)
        col.markdown(f"**{'🟠 ' if es_actual else ''}{mes.upper()}**")
        for fecha, qna, conceptos in quincenas:
            col.markdown(f"📅 **{fecha}**{'  ·  ' + qna if qna else ''}")
            for concepto, cats in conceptos:
                col.caption(f"{cats}  {concepto}" if cats else concepto)
        col.markdown("</div>", unsafe_allow_html=True)

    for fila in range(0, len(PAGOS), 3):
        grupo = PAGOS[fila:fila+3]
        cols = st.columns(3)
        for j, (mes, quincenas) in enumerate(grupo):
            es_actual = MESES_NUM.get(mes, 0) == mes_actual
            render_mes(cols[j], mes, quincenas, es_actual)

    st.divider()
    try:
        with open("1780634863696_calendario_pagos_est_2026.ics", "rb") as f_ics:
            ics_bytes = f_ics.read()
        st.download_button(
            "⬇️ Descargar calendario (.ics)",
            data=ics_bytes,
            file_name="calendario_pagos_2026.ics",
            mime="text/calendar",
            use_container_width=True,
            type="primary"
        )
    except:
        st.info("Archivo .ics no disponible.")

    st.markdown("### 📲 ¿Cómo agregar este calendario a tu dispositivo?")
    with st.expander("📱 iPhone / iPad"):
        st.markdown("""
1. Descarga el archivo **.ics** con el botón de arriba
2. Abre la app **Archivos** en tu iPhone
3. Toca el archivo descargado
4. Aparecerá una pantalla para **Agregar todos** los eventos — confirma
5. Los pagos aparecerán en tu app **Calendario**
        """)
    with st.expander("🤖 Android / Google Calendar"):
        st.markdown("""
1. Descarga el archivo **.ics**
2. Ve a [calendar.google.com](https://calendar.google.com) desde tu navegador
3. En la barra lateral toca **Otros calendarios** → **+** → **Importar**
4. Selecciona el archivo **.ics** descargado y confirma
        """)
    with st.expander("💻 Outlook (PC o Mac)"):
        st.markdown("""
1. Descarga el archivo **.ics**
2. Abre **Outlook** → **Archivo** → **Abrir y exportar** → **Importar o exportar**
3. Selecciona **Importar un archivo iCalendar (.ics)**
4. Busca el archivo descargado y confirma
        """)


def vista_nomina():
    st.markdown("## 💰 Mis comprobantes de nómina")
    st.caption("Selecciona el portal según tu tipo de plaza.")

    col1, col2 = st.columns(2)

    with col1:
        with st.container(border=True):
            st.markdown("### 🏛️ Personal Estatal")
            st.markdown("""
**Paso 1** — Haz clic en el botón de abajo  
**Paso 2** — Se abrirá el portal del Gobierno de Jalisco en nueva pestaña  
**Paso 3** — Inicia sesión con tu **RFC** y contraseña institucional  
**Paso 4** — Selecciona el período a consultar
            """)
            st.info("⚠️ Si no carga, cópialo en tu navegador: **miscomprobantesnomina.jalisco.gob.mx/login**")
            st.link_button("🔗 Portal Estatal", "https://miscomprobantesnomina.jalisco.gob.mx/login", type="primary", use_container_width=True)

    with col2:
        with st.container(border=True):
            st.markdown("### Personal Federalizado")
            st.markdown("""
**Paso 1** — Haz clic en el botón de abajo  
**Paso 2** — Se abrirá el portal de la SEP Federal en nueva pestaña  
**Paso 3** — Inicia sesión con tus credenciales federales  
**Paso 4** — Selecciona el período a consultar
            """)
            st.info("⚠️ Si no carga, accede desde **miportal.fone.sep.gob.mx**")
            st.link_button("🔗 Portal Federal SEP", "https://miportal.fone.sep.gob.mx", type="primary", use_container_width=True)


def vista_directorio():
    import pytz

    @st.cache_data(ttl=300)
    def cargar_directorio():
        client = get_client()
        sh = client.open_by_key(st.secrets["sheet_checador_id"])
        ws = sh.worksheet("Directorio")
        data = ws.get_all_records(numericise_ignore=["all"])
        return pd.DataFrame(data).fillna("")

    st.markdown("## 📞 Directorio interno DFC 2026")
    st.caption("Toca un área para ver su equipo · Busca por nombre, extensión o correo")

    df = cargar_directorio()
    if df.empty:
        st.warning("No se pudo cargar el directorio.")
        return

    # ── Separar Centros de Maestros del personal DFC ──
    # Las filas cuya AREA contiene "Centro de Maestros" van a su propia
    # sección (solo visible para RFCs autorizados en secrets); la vista
    # DFC de siempre queda intacta y sin mezclarse.
    _mask_cm = df["AREA"].astype(str).str.contains("Centro de Maestros", case=False, na=False)
    df_cm = df[_mask_cm].copy()
    df = df[~_mask_cm].copy()

    busq = st.text_input("", placeholder="🔍 Busca por nombre, extensión o correo...", label_visibility="collapsed")

    COLORES = {
        "Dirección General":        ("#E1F5EE", "#0F6E56"),
        "Dir. Gestión y Evaluación": ("#EEEDFE", "#534AB7"),
        "Dir. Desarrollo Académico": ("#E6F1FB", "#185FA5"),
    }

    def color(area):
        for k, v in COLORES.items():
            if k in area:
                return v
        return ("#F4F6F9", "#444")

    def tarjeta(row):
        bg, tc = color(row["AREA"])
        ini = (str(row["NOMBRE"])[0] + (str(row["NOMBRE"]).split()[1][0] if len(str(row["NOMBRE"]).split()) > 1 else "")).upper()
        ext_html = f"📞 **{row['EXTENSION']}**" if row["EXTENSION"] else ""
        email_html = row["CORREO"] if row["CORREO"] else ""
        dept_html = f" · *{row['DEPARTAMENTO']}*" if row["DEPARTAMENTO"] else ""

        col1, col2 = st.columns([3,1])
        with col1:
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:10px;padding:4px 0'>"
                f"<div style='width:34px;height:34px;border-radius:50%;background:{bg};color:{tc};"
                f"display:flex;align-items:center;justify-content:center;font-weight:600;font-size:12px;flex-shrink:0'>{ini}</div>"
                f"<div><div style='font-size:14px;font-weight:500'>{row['NOMBRE']}</div>"
                f"<div style='font-size:12px;color:gray'>{row['AREA']}{dept_html}</div></div></div>",
                unsafe_allow_html=True
            )
        with col2:
            if ext_html: st.markdown(ext_html)
            if email_html: st.caption(row["CORREO"])
        st.divider()

    if busq:
        q = busq.lower().strip()
        resultados = df[
            df["NOMBRE"].str.lower().str.contains(q) |
            df["EXTENSION"].astype(str).str.contains(q) |
            df["CORREO"].str.lower().str.contains(q)
        ]
        if resultados.empty:
            st.info("Sin resultados.")
        else:
            st.caption(f"{len(resultados)} resultado(s)")
            for _, row in resultados.iterrows():
                tarjeta(row)
    else:
        # Agrupar por DEPARTAMENTO como área visible (7 secciones)
        ORDEN_AREAS = [
            "Despacho",
            "Recursos Humanos",
            "Viáticos",
            "Recursos Materiales",
            "Licitaciones y Presupuestos",
            "Comunicación Social",
            "Dir. Desarrollo Académico",
        ]
        ICONOS = {
            "Despacho":                    "🏢",
            "Recursos Humanos":            "👥",
            "Viáticos":                    "🧾",
            "Recursos Materiales":         "📦",
            "Licitaciones y Presupuestos": "📋",
            "Comunicación Social":         "📢",
            "Dir. Desarrollo Académico":   "👩‍🏫",
        }
        # Normalizar: si DEPARTAMENTO está vacío usar AREA como dept
        df["DEPT_VISTA"] = df.apply(lambda r: r["DEPARTAMENTO"] if r["DEPARTAMENTO"] else r["AREA"], axis=1)

        depts_en_datos = df["DEPT_VISTA"].unique().tolist()
        # Mostrar en orden acordado + cualquier otro que no esté en la lista
        orden_final = [d for d in ORDEN_AREAS if d in depts_en_datos] + [d for d in depts_en_datos if d not in ORDEN_AREAS]

        for dept in orden_final:
            personas = df[df["DEPT_VISTA"] == dept]
            icono = ICONOS.get(dept, "📁")
            bg, tc = color(df[df["DEPT_VISTA"]==dept]["AREA"].iloc[0])
            with st.expander(f"{icono} {dept}  ·  {len(personas)} personas"):
                for _, row in personas.iterrows():
                    tarjeta(row)

    # ── 🏫 Directorio de Centros de Maestros ──────────────────────────
    # Solo visible para el admin y los RFCs listados en secrets:
    #   rfcs_directorio_cm = ["RFC1", "RFC2", ...]
    # (en secrets, NO en el código: los RFC son dato personal)
    try:
        _rfcs_cm = [str(r).upper().strip() for r in st.secrets.get("rfcs_directorio_cm", [])]
    except Exception:
        _rfcs_cm = []
    _rfc_actual = str(st.session_state.get("rfc", "")).upper().strip()
    _permiso_cm = (st.session_state.get("rol") == "admin") or (_rfc_actual and _rfc_actual in _rfcs_cm)

    if _permiso_cm:
        st.markdown("---")
        st.markdown("## 🏫 Centros de Maestros")
        if df_cm.empty:
            st.info("Aún no hay personal de Centros de Maestros en la tab Directorio. "
                    "Agrega filas con AREA = 'Centro de Maestros <nombre>' y "
                    "DEPARTAMENTO = 'Responsable' o 'Asesor'.")
        else:
            busq_cm = st.text_input("", placeholder="🔍 Busca por nombre o centro...",
                                    label_visibility="collapsed", key="busq_cm")
            _df_v = df_cm
            if busq_cm:
                _q = busq_cm.lower().strip()
                _df_v = df_cm[df_cm["NOMBRE"].str.lower().str.contains(_q, na=False) |
                              df_cm["AREA"].str.lower().str.contains(_q, na=False)]
            if _df_v.empty:
                st.info("Sin resultados.")
            else:
                for _centro in sorted(_df_v["AREA"].unique()):
                    _gente = _df_v[_df_v["AREA"] == _centro]
                    # Responsable primero, con corona
                    _es_resp = _gente["DEPARTAMENTO"].astype(str).str.upper().str.contains("RESPONSABLE", na=False)
                    _gente = pd.concat([_gente[_es_resp], _gente[~_es_resp]])
                    _etq = str(_centro).replace("Centro de Maestros", "").strip()
                    with st.expander(f"🏫 {_etq}  ·  {len(_gente)} persona(s)"):
                        for _, _r in _gente.iterrows():
                            _rol = str(_r.get("DEPARTAMENTO", "")).strip()
                            _icon = "🧑‍💼" if "RESPONSABLE" in _rol.upper() else "•"
                            _extra = f" · 📞 {_r['EXTENSION']}" if str(_r.get("EXTENSION", "")).strip() else ""
                            _correo = f" · {_r['CORREO']}" if str(_r.get("CORREO", "")).strip() else ""
                            st.markdown(f"{_icon} **{_r['NOMBRE']}** — {_rol or 'Asesor'}{_extra}{_correo}")


# ═══════════════════════════════════════════════════════════════════
# PROTOCOLO DE EMERGENCIA — Contacto de emergencia (acceso directo + log)
# ═══════════════════════════════════════════════════════════════════
# Reutiliza la misma lista de RFCs autorizados que el directorio de
# Centros de Maestros (secrets: rfcs_directorio_cm). No hay aprobación
# en tiempo real: el control de acceso YA ES esa lista, decidida de
# antemano. Cada consulta se registra sola en la bitácora.
#
# EN TU GOOGLE SHEET: agrega 2 columnas nuevas a la tab "Directorio":
#   CONTACTO_EMERGENCIA_NOMBRE   |   CONTACTO_EMERGENCIA_TEL

EMERGENCIA_TAB = "Bitacora_Emergencias"
EMERGENCIA_HEADERS = ["ID", "CONSULTOR_RFC", "CONSULTOR_NOMBRE", "COMPAÑERO_CONSULTADO", "FECHA_HORA"]

@st.cache_data(ttl=300)
def _cargar_directorio_completo():
    """Loader a nivel módulo (independiente del anidado dentro de
    vista_directorio) para que el protocolo de emergencia no dependa
    de haber entrado antes a esa vista."""
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    ws = sh.worksheet("Directorio")
    data = ws.get_all_records(numericise_ignore=["all"])
    return pd.DataFrame(data).fillna("")

def _ws_emergencias():
    client = get_client()
    sh = client.open_by_key(st.secrets["sheet_checador_id"])
    try:
        return sh.worksheet(EMERGENCIA_TAB)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(EMERGENCIA_TAB, rows=2, cols=len(EMERGENCIA_HEADERS))
        ws.append_row(EMERGENCIA_HEADERS)
        return ws

def registrar_consulta_emergencia(rfc: str, nombre: str, companero: str):
    """Escribe el log EN EL MOMENTO en que se revela el contacto.
    Nunca debe bloquear una emergencia real por un fallo de red."""
    try:
        ws = _ws_emergencias()
        todas = ws.get_all_records(numericise_ignore=["all"])
        nuevo_id = len(todas) + 1
        ahora = datetime.now(pytz.timezone("America/Mexico_City")).strftime("%Y-%m-%d %H:%M")
        ws.append_row([nuevo_id, rfc, nombre, companero, ahora], value_input_option="USER_ENTERED")
    except Exception:
        pass

def _rfcs_secret(clave: str) -> list[str]:
    try:
        return [str(r).upper().strip() for r in st.secrets.get(clave, [])]
    except Exception:
        return []

def _normalizar_busqueda(texto: str) -> str:
    """Mayúsculas y sin acentos, para comparar sin importar cómo esté
    capturado el nombre en el Sheet (con/sin acentos, mayús/minús).
    'Ángel' y 'angel' deben coincidir; antes no lo hacían."""
    import unicodedata
    t = str(texto or "").upper().strip()
    t = unicodedata.normalize("NFKD", t)
    return "".join(c for c in t if not unicodedata.combining(c))

def vista_contacto_emergencia():
    rfc_actual = str(st.session_state.get("rfc", "")).upper().strip()
    autorizados = _rfcs_secret("rfcs_directorio_cm")
    if st.session_state.get("rol") != "admin" and rfc_actual not in autorizados:
        st.error("No tienes permiso para esta sección.")
        return

    st.markdown("## 🚨 Contacto de Emergencia")

    # ── Disclaimer con aceptación explícita: se pide en CADA sesión ──
    st.error(
        "**AVISO IMPORTANTE — Uso restringido**\n\n"
        "Esta sección contiene datos personales confidenciales (contacto de "
        "emergencia) y debe usarse ÚNICAMENTE en caso de una emergencia real.\n\n"
        "Cada consulta queda registrada de forma permanente con tu nombre, RFC, "
        "la persona consultada y la fecha/hora exacta.\n\n"
        "El uso de esta información para fines distintos a una emergencia "
        "constituye una falta administrativa y, en su caso, puede derivar en "
        "responsabilidad legal conforme a la normatividad aplicable en materia "
        "de protección de datos personales."
    )
    acepto = st.checkbox(
        "Entiendo el aviso anterior, confirmo que se trata de una emergencia real "
        "y acepto que esta consulta quedará registrada con mis datos."
    )
    if not acepto:
        st.info("Marca la casilla anterior para continuar.")
        return

    try:
        df = _cargar_directorio_completo()
    except Exception:
        df = pd.DataFrame()
    if df.empty or "NOMBRE" not in df.columns:
        st.info("No se pudo cargar el directorio.")
        return

    # Mostrar SIEMPRE en mayúsculas, sin importar cómo esté capturado en el
    # Sheet (mezcla de mayúsculas/minúsculas rompía la vista y la búsqueda).
    df = df.copy()
    df["NOMBRE"] = df["NOMBRE"].astype(str).str.upper().str.strip()
    df["_NOMBRE_BUSQUEDA"] = df["NOMBRE"].apply(_normalizar_busqueda)

    busq = st.text_input("🔍 Busca por nombre", placeholder="Escribe nombre o apellido, con o sin acentos...")
    if busq:
        q = _normalizar_busqueda(busq)
        opciones = df[df["_NOMBRE_BUSQUEDA"].str.contains(q, na=False)]["NOMBRE"].tolist()
        st.caption(f"{len(opciones)} resultado(s)")
    else:
        opciones = df["NOMBRE"].tolist()

    companero = st.selectbox("Selecciona al compañero", [""] + sorted(set(opciones)))

    if companero and st.button("🚨 VER CONTACTO DE EMERGENCIA", type="primary", use_container_width=True):
        fila = df[df["NOMBRE"] == companero]
        if fila.empty:
            st.error("No se encontró a esa persona en el directorio.")
        else:
            r = fila.iloc[0]
            c_nombre = str(r.get("CONTACTO_EMERGENCIA_NOMBRE", "")).strip()
            c_tel = str(r.get("CONTACTO_EMERGENCIA_TEL", "")).strip()
            registrar_consulta_emergencia(rfc_actual, st.session_state.get("nombre", ""), companero)
            if c_nombre or c_tel:
                nombre_txt = c_nombre or "(sin nombre registrado)"
                tel_txt = c_tel or "(sin teléfono registrado)"
                st.success(f"**Contacto de {companero}:**\n\n👤 {nombre_txt}\n\n📞 {tel_txt}")
            else:
                st.warning(f"{companero} no tiene contacto de emergencia registrado en el sistema. "
                          "Contacta directamente a RH.")
            _tz_emer = pytz.timezone("America/Mexico_City")
            _sello = datetime.now(_tz_emer).strftime("%d/%m/%Y %H:%M")
            _nombre_ses = st.session_state.get("nombre", "")
            st.caption(f"Consulta registrada — {_nombre_ses} · {rfc_actual} · {_sello}")


def vista_ari():
    """ARI embebida: chat de dudas de RH con contexto del usuario autenticado."""
    try:
        import ari_module
    except ImportError:
        st.error("El módulo de ARI (ari_module.py) no está en el proyecto.")
        return
    if "GEMINI_API_KEY" not in st.secrets:
        st.warning("Falta GEMINI_API_KEY en los secrets para habilitar a ARI.")
        return

    contexto = ""
    if st.session_state.get("rol") == "empleado":
        rfc    = st.session_state.get("rfc", "")
        nombre = st.session_state.get("nombre", "")
        # Minimización de datos hacia la API externa: solo el nombre de pila.
        # Convención del Sheet: APELLIDO_P APELLIDO_M NOMBRE(S).
        _tk = str(nombre).split()
        nombre_pila = " ".join(_tk[2:]) if len(_tk) >= 3 else (_tk[-1] if _tk else "")
        partes = [f"Nombre: {nombre_pila}"]
        try:
            empleados   = cargar_empleados()
            solicitudes = cargar_solicitudes_eco()
            incidencias = cargar_incidencias()
            emp_row = empleados[empleados["RFC"].astype(str).str.upper() == rfc]
            if not emp_row.empty:
                dias_totales = int(emp_row["DIAS TOTALES"].iloc[0])
                usados       = dias_economicos_usados(rfc, solicitudes)
                partes.append(f"Días económicos: {dias_totales - usados} disponibles de {dias_totales} este año")
            if not incidencias.empty and "RFC" in incidencias.columns:
                propias = incidencias[incidencias["RFC"].astype(str).str.upper() == rfc]
                pend = propias[propias["ESTADO"].astype(str).str.contains("PENDIENTE", na=False)]
                if not pend.empty:
                    folios = ", ".join(pend["FOLIO"].astype(str).tolist()[:5])
                    partes.append(f"Solicitudes pendientes de autorización: {folios}")
        except Exception:
            pass
        contexto = "\n".join(partes)
    elif st.session_state.get("rol") == "admin":
        contexto = "El usuario actual es el administrador de RH de la DFC."

    ari_module.render_ari(contexto)


def main():
    st.set_page_config(page_title="Incidencias DFC · RH", page_icon="📋", layout="wide")

    # ── Interceptor QR de validación ─────────────
    if "validar_folio" in st.query_params:
        folio_a_buscar = st.query_params["validar_folio"]
        st.markdown("## 🔍 Verificación de Autenticidad")
        st.caption("Módulo de Control Interno · Dirección de Formación Continua · SEJ")
        st.divider()
        incidencias   = cargar_incidencias()
        solicitudes   = cargar_solicitudes_eco()
        # Buscar en Incidencias (PSE, COM, CHO)
        match = incidencias[incidencias["FOLIO"].astype(str) == folio_a_buscar]
        # Si no está, buscar en Solicitudes (ECO)
        if match.empty and not solicitudes.empty and "FOLIO" in solicitudes.columns:
            match_eco = solicitudes[solicitudes["FOLIO"].astype(str) == folio_a_buscar]
            encontrado_en = "solicitudes"
        else:
            match_eco = pd.DataFrame()
            encontrado_en = "incidencias"

        registro = match if not match.empty else match_eco

        if not registro.empty:
            row = registro.iloc[0]
            st.success("✅ DOCUMENTO AUTÉNTICO Y REGISTRADO EN SISTEMA")
            with st.container(border=True):
                st.write(f"**Folio Oficial:** {folio_a_buscar}")
                nombre     = row.get("NOMBRE") or row.get("Nombre Completo", "")
                rfc_reg    = row.get("RFC", "")
                tipo_raw   = row.get("TIPO") or row.get("Tipo Permiso", "")
                tipo_reg   = "ECO" if str(tipo_raw).lower() in ["economico", "económico"] else tipo_raw
                fi_reg     = row.get("FECHA_INICIO") or row.get("Fecha Inicio", "")
                ff_reg     = row.get("FECHA_FIN") or row.get("Fecha Fin", "")
                aprobado   = row.get("AUTORIZADO_POR") or row.get("Aprobado Por", "")
                estado_reg = row.get("ESTADO") or ("AUTORIZADO" if str(aprobado).strip() != "" else "PENDIENTE")
                motivo_reg = row.get("MOTIVO") or row.get("Motivo", "")
                st.write(f"**Servidor Público:** {nombre}")
                st.write(f"**Filiación:** {rfc_oculto(str(rfc_reg))}")
                st.write(f"**Incidencia:** {TIPO_LABELS.get(tipo_reg, tipo_reg)}")
                st.write(f"**Periodo:** {fi_reg} al {ff_reg}")
                st.write(f"**Estado:** {estado_reg}")
                # PRIVACIDAD: el motivo NO se muestra en el validador público.
                # Los folios son enumerables en la URL y el motivo puede contener
                # información médica/familiar. Para validar autenticidad basta
                # folio + tipo + periodo + estado. El motivo se consulta con sesión.
        else:
            st.error("🚨 ALERTA: DOCUMENTO NO ENCONTRADO O ALTERADO")
            st.warning("Este folio no existe en los registros oficiales de la DFC. El formato impreso podría ser falso o modificado de manera ilícita.")
        if st.button("Ir al inicio de sesión"):
            st.query_params.clear()
            st.rerun()
        return

    if "rol" not in st.session_state:
        st.markdown("<style>[data-testid='stSidebar']{display:none}</style>", unsafe_allow_html=True)
        login()
        return

    with st.sidebar:
        st.image("dfc_logo.png", use_container_width=True)
        st.markdown("**Portal de Gestión de Incidencias**")
        st.caption('👤 Servidor(a) Público(a): ' + st.session_state['nombre'])
        st.caption('🏷️ Nivel de acceso: ' + ('Administrador RH' if st.session_state['rol'] == 'admin' else 'Personal de la Dirección'))
        st.divider()
        if st.session_state.get("rol") == "admin":
            if st.button("🔄 Limpiar caché"):
                st.cache_data.clear()
                st.success("Caché limpiado.")
        if st.button("📅 Calendario de pagos"):
            st.session_state["vista"] = "calendario"
            st.rerun()
        if st.button("💰 Mis comprobantes de nómina"):
            st.session_state["vista"] = "nomina"
            st.rerun()
        if st.button("📞 Directorio DFC"):
            st.session_state["vista"] = "directorio"
            st.rerun()
        _rfc_sb = str(st.session_state.get("rfc", "")).upper().strip()
        _autoriz_emer = _rfcs_secret("rfcs_directorio_cm")
        if st.session_state.get("rol") == "admin" or _rfc_sb in _autoriz_emer:
            if st.button("🚨 Contacto de Emergencia", key="btn_emer_sidebar"):
                st.session_state["vista"] = "emergencia"
                st.rerun()
            # Colorea SOLO este botón (busca por su texto exacto, no depende
            # de clases internas de Streamlit que cambian entre versiones).
            import streamlit.components.v1 as components
            components.html("""
                <script>
                function _colorearBotonEmergencia() {
                    try {
                        const botones = window.parent.document.querySelectorAll('button');
                        botones.forEach(function(b) {
                            if (b.innerText && b.innerText.indexOf("Contacto de Emergencia") !== -1) {
                                b.style.backgroundColor = "#FF6B00";
                                b.style.color = "#FFFFFF";
                                b.style.fontWeight = "bold";
                                b.style.border = "none";
                            }
                        });
                    } catch (e) {}
                }
                _colorearBotonEmergencia();
                if (!window.parent._emerObserverSet) {
                    window.parent._emerObserverSet = true;
                    new MutationObserver(_colorearBotonEmergencia)
                        .observe(window.parent.document.body, {childList: true, subtree: true});
                }
                </script>
            """, height=0)
        if st.button("🤖 🧠 Pregúntale a ARI la IA de RH"):
            st.session_state["vista"] = "ari"
            st.rerun()
        if st.button("🏠 Inicio"):
            st.session_state["vista"] = "inicio"
            st.rerun()
        if st.button("Cerrar sesión"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    vista = st.session_state.get("vista", "inicio")
    vista = st.session_state.get("vista", "inicio")
    if vista == "ari":
        vista_ari()
    elif vista == "directorio":
        vista_directorio()
    elif vista == "emergencia":
        vista_contacto_emergencia()
    elif vista == "calendario":
        vista_calendario()
    elif vista == "nomina":
        vista_nomina()
    elif st.session_state["rol"] == "admin":
        vista_admin()
    else:
        vista_empleado()

if __name__ == "__main__":
    main()