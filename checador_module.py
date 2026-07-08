# -*- coding: utf-8 -*-
"""Módulo del reloj checador — DFC/SEJ.
Extraído de app_incidencias.py sin cambios de comportamiento (motor v2.2).
Recibe sus dependencias por parámetro: NO importa nada de app_incidencias
(evita imports circulares). Contiene: parser del reporte del checador,
compensaciones, salidas anticipadas, justificaciones de Dirección,
Reporte Ejecutivo en Excel y PDF, y el render completo de la pestaña.
"""
import os
import io
import streamlit as st
import pandas as pd
import pytz
from datetime import datetime, date, timedelta
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER


def render_checador(deps):
    # ── Dependencias inyectadas desde app_incidencias (sin import circular) ──
    cargar_festivos = deps["cargar_festivos"]
    festivos_en_periodo = deps["festivos_en_periodo"]
    get_client = deps["get_client"]
    cargar_justif_direccion = deps["cargar_justif_direccion"]
    guardar_justif_direccion = deps["guardar_justif_direccion"]
    eliminar_justif_direccion = deps["eliminar_justif_direccion"]
    cargar_observaciones = deps["cargar_observaciones"]
    guardar_observaciones = deps["guardar_observaciones"]
    guardar_asistencia_mes = deps["guardar_asistencia_mes"]
    DIRECTORA_NOMBRE = deps["DIRECTORA_NOMBRE"]
    DIRECTORA_CARGO = deps["DIRECTORA_CARGO"]

    import xlrd
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DIAS_NUM_CH  = {0:"LUN",1:"MAR",2:"MIE",3:"JUE",4:"VIE",5:"SAB",6:"DOM"}
    DIAS_NOMBRE  = {"LUN":"Lunes","MAR":"Martes","MIE":"Miércoles","JUE":"Jueves","VIE":"Viernes","SAB":"Sábado","DOM":"Domingo"}
    # Si este texto NO aparece en el reporte generado, la app está corriendo código viejo.
    VERSION_REPORTE = "motor v2.2 (05-jul-2026) · módulo"

    def get_client_ch():
        return get_client()

    @st.cache_data(ttl=300)
    def load_emp_ch():
        client = get_client_ch()
        ws = client.open_by_key(st.secrets["sheet_checador_id"]).worksheet("empleados")
        data = ws.get_all_records(numericise_ignore=["all"])
        df = pd.DataFrame(data).fillna("")
        df["ID"] = df["ID"].astype(str).str.strip()
        if "ACTIVO" not in df.columns:
            df["ACTIVO"] = "SI"
        return df

    @st.cache_data(ttl=300)
    def load_fest_ch():
        client = get_client_ch()
        ws = client.open_by_key(st.secrets["sheet_checador_id"]).worksheet("festivos")
        data = ws.get_all_records()
        df = pd.DataFrame(data).fillna("")
        fechas = set()
        if "FECHA_INICIO" in df.columns:
            for _, r in df.iterrows():
                fi = pd.to_datetime(r.get("FECHA_INICIO",""), errors="coerce")
                ff = pd.to_datetime(r.get("FECHA_FIN","") or r.get("FECHA_INICIO",""), errors="coerce")
                if pd.isna(fi): continue
                if pd.isna(ff): ff = fi
                d = fi
                while d <= ff:
                    fechas.add(d.date())
                    d += timedelta(days=1)
        return fechas

    @st.cache_data(ttl=300)
    def load_hist_ch():
        try:
            client = get_client_ch()
            ws = client.open_by_key(st.secrets["sheet_checador_id"]).worksheet("HISTORIAL_HORARIOS")
            data = ws.get_all_records(numericise_ignore=["all"])
            return pd.DataFrame(data).fillna("") if data else pd.DataFrame()
        except:
            return pd.DataFrame()

    @st.cache_data(ttl=300)
    def load_economicos_ch(fi_str, ff_str):
        try:
            client = get_client_ch()
            ws = client.open_by_key(st.secrets["sheet_economicos_id"]).worksheet("Solicitudes")
            data = ws.get_all_records(numericise_ignore=["all"])
            df = pd.DataFrame(data).fillna("")
            if df.empty: return pd.DataFrame()
            df["Fecha Inicio"] = pd.to_datetime(df["Fecha Inicio"], errors="coerce", dayfirst=True)
            df["Fecha Fin"]    = pd.to_datetime(df["Fecha Fin"],    errors="coerce", dayfirst=True)
            fi = datetime.strptime(fi_str, "%Y-%m-%d")
            ff = datetime.strptime(ff_str, "%Y-%m-%d")
            mask = (df["Fecha Inicio"] <= ff) & (df["Fecha Fin"] >= fi)
            # Solo aprobados
            aprobados = df[mask & (df["Aprobado Por"].astype(str).str.strip() != "")]
            return aprobados
        except Exception as e:
            return pd.DataFrame()

    @st.cache_data(ttl=300)
    def load_incapacidades_ch(fi_str, ff_str):
        try:
            client = get_client_ch()
            ws = client.open_by_key(st.secrets["sheet_economicos_id"]).worksheet("Incapacidades")
            data = ws.get_all_records(numericise_ignore=["all"])
            df = pd.DataFrame(data).fillna("")
            if df.empty: return pd.DataFrame()
            df["Fecha Inicio"]  = pd.to_datetime(df["Fecha Inicio"],  errors="coerce", dayfirst=True)
            df["Fecha Termino"] = pd.to_datetime(df["Fecha Termino"], errors="coerce", dayfirst=True)
            fi = datetime.strptime(fi_str, "%Y-%m-%d")
            ff = datetime.strptime(ff_str, "%Y-%m-%d")
            mask = (df["Fecha Inicio"] <= ff) & (df["Fecha Termino"] >= fi)
            return df[mask].copy()
        except:
            return pd.DataFrame()

    @st.cache_data(ttl=300)
    def load_incidencias_ch(fi_str, ff_str):
        """Carga pases y comisiones autorizados del Sheet del checador."""
        try:
            client = get_client_ch()
            ws = client.open_by_key(st.secrets["sheet_checador_id"]).worksheet("Incidencias")
            data = ws.get_all_records(numericise_ignore=["all"])
            df = pd.DataFrame(data).fillna("")
            if df.empty: return pd.DataFrame()
            df = df[df["ESTADO"] == "AUTORIZADO"]
            df["FECHA_INICIO"] = pd.to_datetime(df["FECHA_INICIO"], errors="coerce")
            df["FECHA_FIN"]    = pd.to_datetime(df["FECHA_FIN"],    errors="coerce")
            fi = datetime.strptime(fi_str, "%Y-%m-%d")
            ff = datetime.strptime(ff_str, "%Y-%m-%d")
            mask = (df["FECHA_INICIO"] <= ff) & (df["FECHA_FIN"] >= fi)
            return df[mask].copy()
        except:
            return pd.DataFrame()

    def extraer_motivo_usuario(motivo_raw: str) -> str:
        """Extrae el texto que escribió el empleado, quitando solo las etiquetas
        técnicas exactas. Si tras limpiar queda algo, lo devuelve; si no queda
        nada pero el usuario sí escribió, devuelve el texto completo (fallback).
        Nunca borra información del usuario."""
        import re as _re
        if not motivo_raw:
            return ""
        raw = str(motivo_raw).strip()
        segmentos = [s.strip() for s in raw.replace("|", "\n").split("\n") if s.strip()]
        utiles = []
        for s in segmentos:
            sl = s.lower()
            # descartar SOLO etiquetas técnicas exactas
            if sl in ("pase de entrada", "pase de salida", "pase de salida sin retorno",
                      "pase de salida con retorno", "comisión", "comision"):
                continue
            # descartar "Entrada: 10:30" / "Salida: 14:00" / "Retorno: 15:00" (etiqueta + hora)
            if _re.match(r'^(entrada|salida|retorno|hora)\s*:?\s*\d{1,2}[:.]\d{2}\s*$', sl):
                continue
            utiles.append(s)
        if utiles:
            return " ".join(utiles).strip()
        # No quedó texto del usuario tras limpiar etiquetas técnicas → vacío.
        # (No se devuelve el raw para no mostrar "Pase de entrada | Entrada: 10:30"
        #  como si fuera un motivo escrito por el empleado.)
        return ""

    def build_justificantes_rfc(economicos, incapacidades, incidencias_df, fi, ff, eventos_inst=None):
        """
        Retorna dict: {RFC_upper -> {date -> tipo_justificante}}
        Incluye: días económicos, incapacidades, comisiones, pases de entrada
        """
        resultado = {}

        # Días económicos aprobados
        if not economicos.empty:
            col_rfc    = next((c for c in economicos.columns if "RFC" in c.upper()), None)
            col_motivo = next((c for c in economicos.columns if "MOTIVO" in c.upper() or "Motivo" in c), None)
            col_fi     = next((c for c in economicos.columns if "INICIO" in c.upper()), None)
            col_ff     = next((c for c in economicos.columns if "FIN" in c.upper() and "INICIO" not in c.upper()), None)
            if col_rfc:
                for _, r in economicos.iterrows():
                    rfc = str(r.get(col_rfc,"")).strip().upper()
                    if not rfc: continue
                    fechas_just = []
                    # Intentar leer fechas exactas del motivo
                    motivo = str(r.get(col_motivo, "")) if col_motivo else ""
                    if "Fechas:" in motivo:
                        try:
                            parte = motivo.split("Fechas:")[1].strip()
                            for f_str in parte.split(","):
                                f_str = f_str.strip()
                                for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
                                    try:
                                        fechas_just.append(datetime.strptime(f_str, fmt).date())
                                        break
                                    except: pass
                        except: pass
                    # Si no hay fechas en motivo, usar rango
                    if not fechas_just and col_fi and col_ff:
                        try:
                            d = r[col_fi]
                            f = r[col_ff]
                            while d <= f:
                                fechas_just.append(d.date())
                                d += timedelta(days=1)
                        except: pass
                    col_mot_eco = next((c for c in economicos.columns if "MOTIVO" in c.upper()), None)
                    for fecha_eco in fechas_just:
                        if fi <= fecha_eco <= ff:
                            mot_eco = str(r.get(col_mot_eco,"")).strip() if col_mot_eco else ""
                            resultado.setdefault(rfc, {})[fecha_eco] = f"Día económico|motivo:{mot_eco}" if mot_eco else "Día económico"

        # Incapacidades
        if not incapacidades.empty:
            col_rfc  = next((c for c in incapacidades.columns if "RFC" in c.upper()), None)
            col_fi   = next((c for c in incapacidades.columns if "INICIO" in c.upper()), None)
            col_ff   = next((c for c in incapacidades.columns if "TERMINO" in c.upper() or ("FIN" in c.upper() and "INICIO" not in c.upper())), None)
            col_tipo = next((c for c in incapacidades.columns if "TIPO" in c.upper()), None)
            if col_rfc and col_fi and col_ff:
                for _, r in incapacidades.iterrows():
                    rfc = str(r.get(col_rfc,"")).strip().upper()
                    if not rfc: continue
                    tipo = str(r.get(col_tipo,"Incapacidad")).strip() if col_tipo else "Incapacidad"
                    try:
                        d = r[col_fi]
                        f = r[col_ff]
                        while d <= f:
                            if fi <= d.date() <= ff:
                                resultado.setdefault(rfc, {})[d.date()] = tipo or "Incapacidad"
                            d += timedelta(days=1)
                    except: pass

        # Comisiones y pases autorizados desde Incidencias
        if not incidencias_df.empty:
            for _, r in incidencias_df.iterrows():
                rfc  = str(r.get("RFC","")).strip().upper()
                tipo = str(r.get("TIPO","")).strip()
                if not rfc: continue
                try:
                    d = r["FECHA_INICIO"]
                    f = r["FECHA_FIN"]
                    if pd.isna(d) or pd.isna(f): continue
                    while d <= f:
                        if fi <= d.date() <= ff:
                            if tipo == "COM":
                                resultado.setdefault(rfc, {})[d.date()] = "Comisión"
                            elif tipo == "RGU":
                                resultado.setdefault(rfc, {})[d.date()] = "Reposición de guardias"
                            elif tipo == "PEN":
                                # Pase de entrada aprobado SIEMPRE justifica el retardo.
                                # El motivo del empleado (si lo hay) es informativo.
                                motivo_real = extraer_motivo_usuario(r.get("MOTIVO",""))
                                if motivo_real:
                                    resultado.setdefault(rfc, {}).setdefault(d.date(), f"Pase de entrada|motivo:{motivo_real}")
                                else:
                                    resultado.setdefault(rfc, {}).setdefault(d.date(), "Pase de entrada")
                            elif tipo == "PSE":
                                motivo_pse = str(r.get("MOTIVO","")).lower()
                                hora_ret   = str(r.get("HORA_RETORNO","")).strip()
                                mot_real   = extraer_motivo_usuario(r.get("MOTIVO",""))
                                suf = f"|motivo:{mot_real}" if mot_real else ""
                                if "sin retorno" in motivo_pse:
                                    resultado.setdefault(rfc, {})[d.date()] = f"Pase de salida sin retorno{suf}"
                                elif hora_ret:
                                    resultado.setdefault(rfc, {})[d.date()] = f"Pase de salida|retorno:{hora_ret}{suf}"
                                else:
                                    resultado.setdefault(rfc, {}).setdefault(d.date(), f"Pase de salida{suf}")
                        d += timedelta(days=1)
                except: pass

        # Eventos institucionales — aplican a TODOS los empleados
        if eventos_inst:
            from gspread import Client  # solo para verificar que tenemos clientes cargados
            for ev in eventos_inst:
                fecha_ev = ev["fecha"] if isinstance(ev["fecha"], type(fi)) else ev["fecha"]
                if fi <= fecha_ev <= ff:
                    # Agregar a todos los empleados activos
                    for rfc_emp in set(list(resultado.keys()) + ["__GLOBAL__"]):
                        resultado.setdefault("__GLOBAL__", {})[fecha_ev] = ev["motivo"]
        return resultado

    def get_horario_fecha(emp_row, fecha, historial_df):
        rfc = str(emp_row.get("RFC","")).upper().strip()
        if historial_df is None or historial_df.empty or not rfc:
            return emp_row
        hist_emp = historial_df[historial_df["RFC"].astype(str).str.upper().str.strip() == rfc]
        if hist_emp.empty:
            return emp_row
        for _, h in hist_emp.iterrows():
            try:
                raw_fi = str(h.get("FECHA_INICIO","")).strip()
                raw_ff = str(h.get("FECHA_FIN","")).strip()
                ts_fi = pd.to_datetime(raw_fi, errors="coerce") if raw_fi else None
                ts_ff = pd.to_datetime(raw_ff, errors="coerce") if raw_ff else None
                fi_h = ts_fi.date() if (ts_fi is not None and not pd.isna(ts_fi)) else None
                ff_h = ts_ff.date() if (ts_ff is not None and not pd.isna(ts_ff)) else None
                # SOLO rangos CERRADOS (con inicio Y fin) pisan el horario.
                # Un registro sin FECHA_FIN NO se considera vigente: el horario
                # actual siempre es el de la tab 'empleados' (evita que un historial
                # viejo abierto pise el horario correcto de hoy).
                if fi_h and ff_h and fi_h <= fecha <= ff_h:
                    return h
                # Sin FECHA_INICIO pero con FECHA_FIN: horario anterior a un cambio.
                if not fi_h and ff_h and fecha <= ff_h:
                    return h
            except: pass
        return emp_row

    def parse_checada_ch(val):
        """Extrae TODAS las marcas de la celda (pueden venir varias pegadas de 5
        chars c/u) y devuelve (primera, última) = (entrada, salida).
        Si solo hay una marca, devuelve (marca, "")."""
        s = str(val).strip()
        if not s or s == "nan":
            return None, None
        marcas = []
        i = 0
        while i + 5 <= len(s):
            m = s[i:i+5]
            if len(m) == 5 and m[2] == ":" and m[:2].isdigit() and m[3:].isdigit():
                marcas.append(m)
                i += 5
            else:
                i += 1
        if not marcas:
            return None, None
        if len(marcas) == 1:
            return marcas[0], ""
        # Primera marca = entrada, última = salida (las del medio son duplicados)
        return marcas[0], marcas[-1]

    def parse_report_ch(file_bytes, empleados, festivos, historial_df, just_rfc, umbral=10,
                        comp_diaria=False, comp_mensual=False):
        def _mins_rango(h1, h2):
            """Minutos de h1 a h2 (mismo día). 0 si no es calculable o es negativo."""
            try:
                a = datetime.strptime(str(h1).strip(), "%H:%M")
                b = datetime.strptime(str(h2).strip(), "%H:%M")
                m = (b.hour - a.hour) * 60 + (b.minute - a.minute)
                return m if m > 0 else 0
            except Exception:
                return 0
        wb = xlrd.open_workbook(file_contents=file_bytes)
        if "Reporte de Asistencia" not in wb.sheet_names():
            raise ValueError("No encontré hoja 'Reporte de Asistencia'.")
        sh = wb.sheet_by_name("Reporte de Asistencia")
        periodo = ""
        for i in range(5):
            for j in range(sh.ncols):
                v = str(sh.cell_value(i,j)).strip()
                if "~" in v and "-" in v:
                    periodo = v; break
        fecha_ini = datetime.strptime(periodo.split("~")[0].strip(), "%Y-%m-%d")
        fecha_fin = datetime.strptime(periodo.split("~")[1].strip(), "%Y-%m-%d")
        dias = []
        d = fecha_ini
        while d <= fecha_fin:
            dias.append(d); d += timedelta(days=1)
        col_to_day = {}
        for i in range(5):
            row = sh.row_values(i)
            if sum(1 for v in row if isinstance(v, float) and 1 <= v <= 31) >= 10:
                for j, v in enumerate(row):
                    if isinstance(v, float) and 1 <= v <= 31:
                        col_to_day[j] = int(v)
                break
        checadas = {}
        i = 0
        while i < sh.nrows:
            row = sh.row_values(i)
            if str(row[0]).strip() == "ID:" and str(row[4]).strip().replace(".0",""):
                uid = str(row[4]).strip().replace(".0","")
                checadas[uid] = {}
                if i+1 < sh.nrows:
                    dr = sh.row_values(i+1)
                    for col, dn in col_to_day.items():
                        val = str(dr[col]).strip() if col < len(dr) else ""
                        if val and val != "nan":
                            e, s = parse_checada_ch(val)
                            if e: checadas[uid][dn] = (e, s)
                i += 2; continue
            i += 1

        activos = empleados[empleados["ACTIVO"]=="SI"]
        resumen, detalle_faltas, detalle_retardos, detalle_omisiones = [], [], [], []

        for _, emp_row in activos.iterrows():
            uid    = emp_row["ID"]
            nombre = emp_row["NOMBRE"]
            rfc    = str(emp_row.get("RFC","")).upper().strip()
            dias_prog = asistidos = faltas = retardos = retardos_min = justif = omisiones = 0
            salidas_antic = retardos_comp = 0
            min_prog_total = min_trab_total = 0
            uid_ch     = checadas.get(uid, {})
            just_emp_personal = just_rfc.get(rfc, {})
            just_emp_global   = just_rfc.get("__GLOBAL__", {})
            just_emp = {**just_emp_global, **just_emp_personal}
            dias_falta = []
            dias_just  = []
            det_retardos_emp = []
            det_omisiones = []

            for dia_dt in dias:
                if dia_dt.date() in festivos:
                    continue
                nd = DIAS_NUM_CH[dia_dt.weekday()]
                hor = get_horario_fecha(emp_row, dia_dt.date(), historial_df)
                ep  = str(hor.get(f"ENTRADA_{nd}","")).strip()
                if not ep:
                    continue
                sp_dia = str(hor.get(f"SALIDA_{nd}","")).strip()
                dias_prog += 1
                ch  = uid_ch.get(dia_dt.day)
                just_tipo = just_emp.get(dia_dt.date())

                if not ch:
                    if just_tipo and not just_tipo.startswith("Pase de entrada"):
                        justif += 1
                        dias_just.append(f"{dia_dt.strftime('%d/%m')} ({just_tipo.split('|')[0]})")
                    else:
                        faltas += 1
                        dias_falta.append(dia_dt.strftime("%d/%m"))
                        # Una falta debe las horas programadas completas del día
                        min_prog_total += _mins_rango(ep, sp_dia)
                    continue

                asistidos += 1
                entrada_real, salida_real = ch
                mins   = 0
                estado = "Asistió"

                # Bug B: si solo hay UNA checada, decidir si es entrada o salida
                # comparándola con el horario programado (entrada vs salida).
                solo_una = bool(entrada_real) and not salida_real
                sp = str(hor.get(f"SALIDA_{nd}","")).strip()
                if solo_una and ep and sp:
                    try:
                        marca_dt = datetime.strptime(entrada_real, "%H:%M")
                        ent_dt   = datetime.strptime(ep, "%H:%M")
                        sal_dt   = datetime.strptime(sp, "%H:%M")
                        dist_ent = abs((marca_dt - ent_dt).total_seconds())
                        dist_sal = abs((marca_dt - sal_dt).total_seconds())
                        if dist_sal < dist_ent:
                            # La única marca es la SALIDA → olvidó checar entrada.
                            # No se inventa retardo: se señala para revisión manual.
                            salida_real  = entrada_real
                            entrada_real = ""
                    except: pass

                # Protección extra: si hay entrada Y salida pero la "entrada" está
                # DESPUÉS de la hora de salida programada (marcas invertidas o dato
                # de horario incorrecto), no calcular un retardo absurdo. Se marca
                # para revisión en vez de inventar cientos de minutos.
                if entrada_real and salida_real and ep and sp:
                    try:
                        er  = datetime.strptime(entrada_real, "%H:%M")
                        srd = datetime.strptime(salida_real, "%H:%M")
                        spd = datetime.strptime(sp, "%H:%M")
                        epd = datetime.strptime(ep, "%H:%M")
                        # Marcas inconsistentes con el horario del día:
                        # (a) la "entrada" cae después de la hora de salida programada, o
                        # (b) la SALIDA real ocurre antes/igual que la hora de ENTRADA
                        #     programada → checó en un turno que NO es el suyo
                        #     (ej. horario vespertino pero marcó toda la mañana).
                        fuera_de_horario = (er >= spd) or (srd <= epd)
                        if fuera_de_horario:
                            estado = "Revisar (checó fuera de su horario)"
                            # VISIBLE: va a la tabla de revisión (df_omis), que sí
                            # se muestra en pantalla y sí se exporta al Excel.
                            omisiones += 1
                            det_omisiones.append({
                                "Empleado": nombre,
                                "Fecha": dia_dt.strftime("%d/%m/%Y"),
                                "Día": DIAS_NOMBRE.get(nd, nd),
                                "Omitió": "⚠ Checó fuera de su horario",
                                "Hora prog. entrada": f"{ep} - {sp}",
                                "Checada registrada": f"{entrada_real} - {salida_real}",
                            })
                            detalle_faltas.append({
                                "nombre": nombre, "fecha": dia_dt.date(), "nd": nd,
                                "prog_entrada": ep, "checada_entrada": entrada_real,
                                "checada_salida": salida_real, "retardo_min": 0,
                                "estado": "Revisar (checó fuera de su horario)",
                                "justificante": just_tipo or ""
                            })
                            continue
                    except: pass

                if not entrada_real:
                    # Si hay un pase de entrada justificado ese día, la entrada está cubierta:
                    # no es omisión.
                    if just_tipo and str(just_tipo).startswith("Pase de entrada"):
                        estado = "Asistió"
                        detalle_faltas.append({
                            "nombre": nombre, "fecha": dia_dt.date(), "nd": nd,
                            "prog_entrada": ep, "checada_entrada": "",
                            "checada_salida": salida_real, "retardo_min": 0,
                            "estado": "Justificado (pase de entrada)", "justificante": just_tipo
                        })
                        continue
                    # Sin checada de entrada y sin justificante: asistió pero falta el dato.
                    omisiones += 1
                    det_omisiones.append({
                        "Empleado": nombre,
                        "Fecha": dia_dt.strftime("%d/%m/%Y"),
                        "Día": DIAS_NOMBRE.get(nd, nd),
                        "Omitió": "Entrada",
                        "Hora prog. entrada": ep,
                        "Checada registrada": f"Salida {salida_real}" if salida_real else "—",
                    })
                    detalle_faltas.append({
                        "nombre": nombre, "fecha": dia_dt.date(), "nd": nd,
                        "prog_entrada": ep, "checada_entrada": "",
                        "checada_salida": salida_real, "retardo_min": 0,
                        "estado": "Sin checada de entrada", "justificante": just_tipo or ""
                    })
                    continue

                try:
                    ep_dt = datetime.strptime(ep, "%H:%M")
                    er_dt = datetime.strptime(entrada_real, "%H:%M")
                    mins  = (er_dt.hour - ep_dt.hour)*60 + (er_dt.minute - ep_dt.minute)
                    if mins > umbral:
                        _jt = str(just_tipo or "")
                        _jt_low = _jt.lower()
                        _jt_up  = _jt.strip().upper()
                        _es_pase_entrada = ("pase de entrada" in _jt_low or _jt_up.startswith("PEN"))
                        _es_pase_salida  = ("pase de salida" in _jt_low or _jt_up.startswith(("PSE","PSR")))
                        _es_dia_completo = ("día económico" in _jt_low or "economico" in _jt_low
                                            or "comisión" in _jt_low or "comision" in _jt_low
                                            or _jt_up.startswith(("ECO","COM","CUM")))
                        if _es_pase_entrada:
                            estado = "Asistió"  # pase de entrada: justifica la llegada tarde
                        elif _es_pase_salida:
                            estado = "Asistió"  # pase de salida ese día: no se penaliza la entrada
                        elif _es_dia_completo:
                            estado = "Asistió"  # día económico/comisión: no se penaliza
                        else:
                            # ── Compensación diaria (opcional): si salió al menos
                            # tantos minutos DESPUÉS de su hora como llegó tarde,
                            # el retardo queda compensado y no se cuenta ──
                            _extra_sal = _mins_rango(sp, salida_real) if (comp_diaria and salida_real and sp) else 0
                            _compensado = comp_diaria and _extra_sal >= mins
                            if _compensado:
                                retardos_comp += 1
                                estado = "Retardo compensado"
                            else:
                                retardos += 1
                                retardos_min += mins
                                estado = "Retardo"
                            _motivo_dia = ""
                            if just_tipo and "motivo:" in str(just_tipo):
                                _motivo_dia = str(just_tipo).split("motivo:")[1].strip()
                            elif just_tipo:
                                _motivo_dia = str(just_tipo).split("|")[0].strip()
                            det_retardos_emp.append({
                                "Empleado": nombre,
                                "Fecha": dia_dt.strftime("%d/%m/%Y"),
                                "Día": DIAS_NOMBRE.get(nd, nd),
                                "Hora prog.": ep,
                                "Hora real": entrada_real,
                                "Minutos tarde": mins,
                                "Observación": ((f"Compensado: salida prog. {sp}, salió {salida_real} "
                                                 f"(+{_extra_sal} min)"
                                                 + (" ⚠ verificar horario en Sheet o checada"
                                                    if _extra_sal > 180 else ""))
                                                if _compensado else ""),
                                "Motivo del día": _motivo_dia
                            })
                except: pass

                # Detectar regreso tardío en pase de salida con retorno
                if just_tipo and "retorno:" in str(just_tipo) and salida_real:
                    try:
                        hora_ret_aut = just_tipo.split("retorno:")[1].split("|")[0].strip()
                        ret_aut_dt   = datetime.strptime(hora_ret_aut, "%H:%M")
                        sal_real_dt  = datetime.strptime(salida_real,  "%H:%M")
                        mins_tarde   = (sal_real_dt.hour - ret_aut_dt.hour)*60 + (sal_real_dt.minute - ret_aut_dt.minute)
                        if mins_tarde > umbral:
                            retardos += 1
                            retardos_min += mins_tarde
                            det_retardos_emp.append({
                                "Empleado": nombre,
                                "Fecha": dia_dt.strftime("%d/%m/%Y"),
                                "Día": DIAS_NOMBRE.get(nd, nd),
                                "Hora prog.": hora_ret_aut,
                                "Hora real": salida_real,
                                "Minutos tarde": mins_tarde,
                                "Observación": "Regreso tardío de pase de salida",
                                "Motivo del día": "Pase de salida con retorno"
                            })
                    except: pass

                # Detectar salida omitida: checó entrada pero no salida.
                # EXCEPCIÓN: si tiene un pase de salida autorizado (PSE/PSR), no marcó
                # salida a propósito (se retiró con permiso) → no es omisión.
                _jts = str(just_tipo or "")
                tiene_pase_salida = ("pase de salida" in _jts.lower() or
                                     _jts.strip().upper().startswith(("PSE","PSR")))
                if entrada_real and not salida_real and sp and not tiene_pase_salida:
                    omisiones += 1
                    det_omisiones.append({
                        "Empleado": nombre,
                        "Fecha": dia_dt.strftime("%d/%m/%Y"),
                        "Día": DIAS_NOMBRE.get(nd, nd),
                        "Omitió": "Salida",
                        "Hora prog. entrada": ep,
                        "Checada registrada": f"Entrada {entrada_real}",
                    })

                # ── SALIDA ANTICIPADA: salió antes de su hora programada ──
                # (caso César: horario 7:00-13:00, checando 7:00-12:00 sin que
                # nada lo detectara — el sistema solo vigilaba la entrada).
                # No aplica si tiene pase de salida o justificante de día completo.
                _jts_up = _jts.strip().upper()
                _just_dia_comp = ("económico" in _jts.lower() or "economico" in _jts.lower()
                                  or "comisión" in _jts.lower() or "comision" in _jts.lower()
                                  or "dirección" in _jts.lower() or "direccion" in _jts.lower()
                                  or _jts_up.startswith(("ECO", "COM", "CUM")))
                if entrada_real and salida_real and sp and not tiene_pase_salida and not _just_dia_comp:
                    _antic = _mins_rango(salida_real, sp)
                    if _antic > umbral:
                        salidas_antic += 1
                        det_omisiones.append({
                            "Empleado": nombre,
                            "Fecha": dia_dt.strftime("%d/%m/%Y"),
                            "Día": DIAS_NOMBRE.get(nd, nd),
                            "Omitió": f"⏰ Salida anticipada ({_antic} min antes)",
                            "Hora prog. entrada": f"Salida prog. {sp}",
                            "Checada registrada": f"Salió {salida_real}",
                        })

                # ── Horas del día: solo días con ambas marcas se comparan
                # (trabajadas vs programadas). Faltas ya sumaron lo programado. ──
                if entrada_real and salida_real and ep and sp:
                    _mp_dia = _mins_rango(ep, sp)
                    if _mp_dia:
                        min_prog_total += _mp_dia
                        min_trab_total += _mins_rango(entrada_real, salida_real)

                detalle_faltas.append({
                    "nombre": nombre, "fecha": dia_dt.date(), "nd": nd,
                    "prog_entrada": ep, "checada_entrada": entrada_real,
                    "checada_salida": salida_real, "retardo_min": max(0,mins),
                    "estado": estado, "justificante": just_tipo or ""
                })

            # ── Compensación mensual (opcional): si trabajó igual o más horas
            # que las programadas del período, sus retardos no se cuentan
            # (quedan visibles en el detalle marcados como compensados) ──
            if comp_mensual and min_prog_total > 0 and min_trab_total >= min_prog_total and retardos > 0:
                _dif_h_emp = (min_trab_total - min_prog_total) / 60
                for _r in det_retardos_emp:
                    if not _r.get("Observación"):
                        _r["Observación"] = f"Compensado: cumplió sus horas del período (+{_dif_h_emp:.1f} h)"
                retardos_comp += retardos
                retardos = 0
                retardos_min = 0
            detalle_retardos.extend(det_retardos_emp)
            det_retardos_emp = []

            detalle_omisiones.extend(det_omisiones)

            pct = round((asistidos+justif)/dias_prog*100) if dias_prog > 0 else 0
            resumen.append({
                "RFC": rfc,
                "Empleado": nombre,
                "Días prog.": dias_prog,
                "Asistidos": asistidos,
                "Faltas": faltas,
                "Justificadas": justif,
                "Retardos": retardos,
                "Min. retardo": retardos_min,
                "Omisiones": omisiones,
                "Dif. horas": round((min_trab_total - min_prog_total)/60, 1),
                "% Asistencia": f"{pct}%",
                "Días que faltó": ", ".join(dias_falta) if dias_falta else "—",
                "Días justificados": ", ".join(dias_just) if dias_just else "—",
            })

        return pd.DataFrame(resumen), pd.DataFrame(detalle_retardos), pd.DataFrame(detalle_omisiones), periodo, fecha_ini, fecha_fin

    def generar_excel_reporte(df_res, df_ret, df_omis, df_obs, periodo, fecha_ini, fecha_fin, umbral, fests_periodo=None):
        wb = Workbook()
        AZUL    = "002F6C"
        BLANCO  = "FFFFFF"
        GRIS    = "F4F6F9"
        VERDE   = "D4EDDA"
        AMARILLO= "FFF3CD"
        ROJO    = "F8D7DA"
        AZUL_C  = "D0E4F7"

        def hdr_font(bold=True, color=BLANCO, sz=11):
            return Font(bold=bold, color=color, size=sz, name="Arial")
        def fill(color):
            return PatternFill("solid", fgColor=color)
        def center():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)
        def left():
            return Alignment(horizontal="left", vertical="center", wrap_text=True)
        def border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)

        # ── Hoja 1: Reporte Ejecutivo ──────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Reporte Ejecutivo"
        mes_nombre = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][fecha_ini.month-1]

        # Logo institucional (si existe el archivo)
        try:
            from openpyxl.drawing.image import Image as XLImage
            if os.path.exists("logos_gris.png"):
                logo_img = XLImage("logos_gris.png")
                logo_img.width  = 220
                logo_img.height = 55
                ws1.add_image(logo_img, "A1")
                ws1.row_dimensions[1].height = 45
        except Exception:
            pass

        ws1.merge_cells("A1:K1")
        ws1["A1"] = "SECRETARÍA DE EDUCACIÓN JALISCO"
        ws1["A1"].font = Font(bold=True, size=13, name="Arial", color=AZUL)
        ws1["A1"].alignment = center()

        ws1.merge_cells("A2:K2")
        ws1["A2"] = "DIRECCIÓN DE FORMACIÓN CONTINUA"
        ws1["A2"].font = Font(bold=True, size=12, name="Arial", color=AZUL)
        ws1["A2"].alignment = center()

        ws1.merge_cells("A3:K3")
        ws1["A3"] = f"REPORTE DE ASISTENCIA — {mes_nombre.upper()} {fecha_ini.year}"
        ws1["A3"].font = Font(bold=True, size=11, name="Arial")
        ws1["A3"].alignment = center()

        ws1.merge_cells("A4:K4")
        ws1["A4"] = (f"Período: {fecha_ini.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}  |  "
                     f"Justificantes: días económicos, incapacidades, comisiones y pases aplicados  |  "
                     f"Tolerancia retardo: {umbral} min")
        ws1["A4"].font = Font(italic=True, size=9, name="Arial", color="555555")
        ws1["A4"].alignment = center()
        ws1.row_dimensions[4].height = 30

        # ── Bloque de metadatos institucionales ──
        _tz_mx_rep = pytz.timezone("America/Mexico_City")
        _fecha_emision = datetime.now(pytz.utc).astimezone(_tz_mx_rep)
        ws1.merge_cells("A5:K5")
        ws1["A5"] = (f"Tipo de Documento: Reporte Ejecutivo Institucional  ·  "
                     f"Fecha de Emisión: {_fecha_emision.strftime('%d/%m/%Y')}  ·  "
                     f"Área Emisora: Control de Personal / Recursos Humanos")
        ws1["A5"].font = Font(size=8, name="Arial", color="555555")
        ws1["A5"].alignment = center()

        ws1.append([])

        hdrs = ["Empleado","Días prog.","Asistidos","Faltas","Justificadas","Retardos","Min. retardo","Omisiones","Dif. horas","% Asistencia","Días que faltó","Días justificados"]
        ws1.append(hdrs)
        hrow = ws1.max_row
        for col, _ in enumerate(hdrs, 1):
            c = ws1.cell(hrow, col)
            c.font      = hdr_font()
            c.fill      = fill(AZUL)
            c.alignment = center()
            c.border    = border()

        for _, r in df_res.sort_values("Faltas", ascending=False).iterrows():
            faltas = r["Faltas"]
            pct    = int(str(r["% Asistencia"]).replace("%","") or 0)
            bg = ROJO if faltas >= 10 or pct < 75 else (AMARILLO if faltas >= 3 or pct < 90 else (VERDE if r["Justificadas"] > 0 else BLANCO))
            row_data = [r["Empleado"], r["Días prog."], r["Asistidos"], r["Faltas"],
                        r["Justificadas"], r["Retardos"], r.get("Min. retardo", 0),
                        r.get("Omisiones", 0), r.get("Dif. horas", ""),
                        r["% Asistencia"], r["Días que faltó"], r["Días justificados"]]
            ws1.append(row_data)
            drow = ws1.max_row
            for col in range(1, len(hdrs) + 1):
                c = ws1.cell(drow, col)
                c.fill      = fill(bg)
                c.border    = border()
                c.font      = Font(size=9, name="Arial")
                c.alignment = center() if col > 1 else left()
            # Dif. horas positiva en verde: lo bueno también se ve
            try:
                if float(r.get("Dif. horas", 0)) > 0:
                    c_dif = ws1.cell(drow, 9)
                    c_dif.fill = fill(VERDE)
                    c_dif.font = Font(size=9, name="Arial", bold=True, color="0F6E56")
            except Exception:
                pass

        ws1.append([])
        ws1.append(["Leyenda:", "0-2 faltas/≥90%", "3-9 faltas/75-89%", "≥10 faltas/<75%", "Justificado"])
        lrow = ws1.max_row
        for col, bg in [(2,BLANCO),(3,AMARILLO),(4,ROJO),(5,VERDE)]:
            c = ws1.cell(lrow, col)
            c.fill = fill(bg)
            c.font = Font(size=8, name="Arial", bold=True)
            c.alignment = center()

        ws1.column_dimensions["A"].width = 35
        for col in ["B","C","D","E","F","G","H"]:
            ws1.column_dimensions[col].width = 12
        ws1.column_dimensions["I"].width = 12
        ws1.column_dimensions["J"].width = 45
        ws1.column_dimensions["K"].width = 45

        # ── Días inhábiles / festivos del período ──
        prox = lrow + 2
        if fests_periodo:
            ws1.merge_cells(start_row=prox, start_column=1, end_row=prox, end_column=4)
            ct = ws1.cell(prox, 1)
            ct.value = "DÍAS INHÁBILES / NO LABORABLES EN EL PERÍODO"
            ct.font = Font(bold=True, size=9, name="Arial", color=AZUL)
            MESES = {1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}
            for k, (fday, desc) in enumerate(fests_periodo, 1):
                rr = prox + k
                ws1.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
                cd = ws1.cell(rr, 1)
                cd.value = f"• {fday.day} de {MESES[fday.month]}: {desc}"
                cd.font = Font(size=8, name="Arial")
            base = prox + len(fests_periodo) + 3
        else:
            base = prox + 3

        # ── Bloque de firma / Vo.Bo. (aval para archivo y auditoría) ──
        ws1.merge_cells(start_row=base, start_column=4, end_row=base, end_column=8)
        cf = ws1.cell(base, 4)
        cf.value = "_______________________________________"
        cf.alignment = center()
        ws1.merge_cells(start_row=base+1, start_column=4, end_row=base+1, end_column=8)
        cn = ws1.cell(base+1, 4)
        cn.value = f"Vo.Bo.  {DIRECTORA_NOMBRE}"
        cn.font = Font(bold=True, size=10, name="Arial", color=AZUL)
        cn.alignment = center()
        ws1.merge_cells(start_row=base+2, start_column=4, end_row=base+2, end_column=8)
        cc = ws1.cell(base+2, 4)
        cc.value = DIRECTORA_CARGO
        cc.font = Font(size=9, name="Arial", color="555555")
        cc.alignment = center()

        # ── Leyenda de confidencialidad (debajo de la firma) ──
        ley_row = base + 4
        ws1.merge_cells(start_row=ley_row, start_column=1, end_row=ley_row, end_column=11)
        cl = ws1.cell(ley_row, 1)
        cl.value = ("Este documento es para uso exclusivo de la Dirección de Formación Continua "
                    "y contiene información confidencial de carácter institucional. "
                    "Su emisión tiene fines de control interno y respaldo administrativo.")
        cl.font = Font(italic=True, size=8, name="Arial", color="555555")
        cl.alignment = center()
        ws1.row_dimensions[ley_row].height = 26

        # ── Pie de generación ──
        pie_row = ley_row + 1
        ws1.merge_cells(start_row=pie_row, start_column=1, end_row=pie_row, end_column=11)
        cp = ws1.cell(pie_row, 1)
        cp.value = (f"Generado el {_fecha_emision.strftime('%d/%m/%Y %H:%M')} · "
                    f"Recursos Humanos — DFC · SEJ")
        cp.font = Font(size=8, name="Arial", color="999999")
        cp.alignment = center()

        # ── Hoja 2: Detalle faltas ──────────────────────────────────────────
        ws2 = wb.create_sheet("Detalle faltas y justificantes")
        ws2.merge_cells("A1:E1")
        ws2["A1"] = f"DETALLE DE FALTAS Y JUSTIFICANTES — {mes_nombre.upper()} {fecha_ini.year}"
        ws2["A1"].font = Font(bold=True, size=11, name="Arial", color=AZUL)
        ws2["A1"].alignment = center()
        ws2.append([])

        hdrs2 = ["Empleado","Días prog.","Faltas","Justificadas","Días que faltó / Justificante","Observación de Dirección"]
        ws2.append(hdrs2)
        hrow2 = ws2.max_row
        for col, _ in enumerate(hdrs2, 1):
            c = ws2.cell(hrow2, col)
            c.font = hdr_font(); c.fill = fill(AZUL); c.alignment = center(); c.border = border()

        df_faltas = df_res[df_res["Faltas"] > 0].sort_values("Faltas", ascending=False)
        for _, r in df_faltas.iterrows():
            detalle_txt = ""
            if r["Días que faltó"] != "—":
                detalle_txt += "FALTA: " + r["Días que faltó"]
            if r["Días justificados"] != "—":
                detalle_txt += (" | " if detalle_txt else "") + "JUST: " + r["Días justificados"]
            # Observaciones del histórico, por cada fecha de falta
            obs_txt = ""
            if df_obs is not None and not df_obs.empty:
                rfc_e = str(r.get("RFC","")).upper()
                fechas_falta = [f.strip() for f in str(r["Días que faltó"]).split(",") if f.strip() and r["Días que faltó"] != "—"]
                partes = []
                for ff_ in fechas_falta:
                    m = df_obs[(df_obs["RFC"].astype(str).str.upper()==rfc_e) & (df_obs["FECHA_FALTA"].astype(str)==ff_)]
                    if not m.empty and str(m.iloc[0]["OBSERVACION"]).strip():
                        partes.append(f"{ff_}: {m.iloc[0]['OBSERVACION']}")
                obs_txt = " | ".join(partes)
            ws2.append([r["Empleado"], r["Días prog."], r["Faltas"], r["Justificadas"], detalle_txt, obs_txt])
            drow = ws2.max_row
            for col in range(1,7):
                c = ws2.cell(drow, col)
                c.border = border(); c.font = Font(size=9, name="Arial")
                c.alignment = center() if col > 1 else left()

        ws2.column_dimensions["A"].width = 35
        for col in ["B","C","D"]:
            ws2.column_dimensions[col].width = 12
        ws2.column_dimensions["E"].width = 50
        ws2.column_dimensions["F"].width = 45

        # ── Hoja 3: Detalle retardos ───────────────────────────────────────
        ws3 = wb.create_sheet("Detalle de retardos")
        ws3.merge_cells("A1:H1")
        ws3["A1"] = f"DETALLE DE RETARDOS — {mes_nombre.upper()} {fecha_ini.year}  (tolerancia {umbral} min) · {VERSION_REPORTE}"
        ws3["A1"].font = Font(bold=True, size=11, name="Arial", color=AZUL)
        ws3["A1"].alignment = center()
        ws3.append([])

        hdrs3 = ["Empleado","Fecha","Día","Hora prog.","Hora real","Minutos tarde","Observación","Motivo del día"]
        ws3.append(hdrs3)
        hrow3 = ws3.max_row
        for col, _ in enumerate(hdrs3, 1):
            c = ws3.cell(hrow3, col)
            c.font = hdr_font(); c.fill = fill(AZUL); c.alignment = center(); c.border = border()

        if not df_ret.empty:
            for _, r in df_ret.iterrows():
                ws3.append([r["Empleado"],r["Fecha"],r["Día"],r["Hora prog."],r["Hora real"],r["Minutos tarde"],r.get("Observación",""),r.get("Motivo del día","")])
                drow = ws3.max_row
                for col in range(1,9):
                    c = ws3.cell(drow, col)
                    c.border = border(); c.font = Font(size=9, name="Arial")
                    c.alignment = center() if col > 1 else left()
                    if r["Minutos tarde"] > 30:
                        c.fill = fill(ROJO)
                    elif r["Minutos tarde"] > 15:
                        c.fill = fill(AMARILLO)

        ws3.column_dimensions["A"].width = 35
        for col in ["B","C","D","E","F","G"]:
            ws3.column_dimensions[col].width = 14
        ws3.column_dimensions["H"].width = 30

        # ── Hoja 4: Entradas/Salidas omitidas ──
        ws4 = wb.create_sheet("Checadas omitidas")
        ws4.merge_cells("A1:F1")
        ws4["A1"] = f"ENTRADAS / SALIDAS OMITIDAS — {mes_nombre.upper()} {fecha_ini.year}"
        ws4["A1"].font = Font(bold=True, size=11, name="Arial", color=AZUL)
        ws4["A1"].alignment = center()
        ws4.append([])
        ws4.append(["NOTA: estos casos NO son retardos. El empleado asistió pero olvidó una checada. Requieren tu revisión."])
        hdrs4 = ["Empleado","Fecha","Día","Omitió","Hora prog. entrada","Checada registrada"]
        ws4.append(hdrs4)
        hrow4 = ws4.max_row
        for col, _ in enumerate(hdrs4, 1):
            c = ws4.cell(hrow4, col)
            c.font = hdr_font(); c.fill = fill(AZUL); c.alignment = center(); c.border = border()
        if df_omis is not None and not df_omis.empty:
            for _, r in df_omis.iterrows():
                ws4.append([r["Empleado"],r["Fecha"],r["Día"],r["Omitió"],r["Hora prog. entrada"],r["Checada registrada"]])
                drow = ws4.max_row
                for col in range(1,7):
                    c = ws4.cell(drow, col)
                    c.border = border(); c.font = Font(size=9, name="Arial")
                    c.alignment = center() if col > 1 else left()
                    c.fill = fill(AMARILLO)
        ws4.column_dimensions["A"].width = 35
        for col in ["B","C","D","E","F"]:
            ws4.column_dimensions[col].width = 16

        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generar_pdf_ejecutivo(df_res, periodo, fecha_ini, fecha_fin, fests_periodo=None):
        from reportlab.lib.pagesizes import landscape, letter as _letter
        mes_nombre = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"][fecha_ini.month-1]
        AZUL = colors.HexColor("#002F6C")
        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(_letter),
                                leftMargin=0.8*cm, rightMargin=0.8*cm, topMargin=0.8*cm, bottomMargin=1.2*cm)
        styles = getSampleStyleSheet()
        elems = []

        # Encabezado con logo — título centrado respecto a la PÁGINA completa
        st_tit = ParagraphStyle("tit", parent=styles["Normal"], fontSize=14, fontName="Helvetica-Bold",
                                textColor=AZUL, alignment=TA_CENTER, leading=17)
        st_sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, fontName="Helvetica-Bold",
                                textColor=AZUL, alignment=TA_CENTER)
        st_per = ParagraphStyle("per", parent=styles["Normal"], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor("#555555"))
        encab = Paragraph("SECRETARÍA DE EDUCACIÓN JALISCO<br/>"
                          "DIRECCIÓN DE FORMACIÓN CONTINUA<br/>"
                          f"<font size=11>REPORTE DE ASISTENCIA — {mes_nombre.upper()} {fecha_ini.year}</font>", st_tit)
        if os.path.exists("logos_gris.png"):
            logo_rl = RLImage("logos_gris.png", width=4.5*cm, height=1.2*cm)
            # logo (izq) | título (centro, simétrico) | espacio igual al logo (der)
            head = Table([[logo_rl, encab, ""]], colWidths=[5*cm, 14*cm, 5*cm])
            head.setStyle(TableStyle([
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("ALIGN",(0,0),(0,0),"LEFT"),
                ("ALIGN",(1,0),(1,0),"CENTER"),
            ]))
            elems.append(head)
        else:
            elems.append(encab)
        elems.append(Spacer(1, 0.15*cm))
        # Línea divisoria bajo el encabezado (aspecto formal)
        from reportlab.platypus import HRFlowable
        elems.append(HRFlowable(width="100%", thickness=1, color=AZUL, spaceBefore=2, spaceAfter=4))
        elems.append(Paragraph(f"Período del reporte: {fecha_ini.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}", st_per))
        # ── Metadatos institucionales ──
        _tz_pdf = pytz.timezone("America/Mexico_City")
        _emision_pdf = datetime.now(pytz.utc).astimezone(_tz_pdf)
        elems.append(Paragraph(
            f"Tipo de Documento: Reporte Ejecutivo Institucional  ·  "
            f"Fecha de Emisión: {_emision_pdf.strftime('%d/%m/%Y')}  ·  "
            f"Área Emisora: Control de Personal / Recursos Humanos", st_per))
        elems.append(Spacer(1, 0.3*cm))

        # Tabla
        st_cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=6.5, leading=8)
        cols = ["Empleado","Días\nprog.","Asist.","Faltas","Justif.","Retardos","Min.\nretardo","Omis.","%","Días que faltó"]
        data = [cols]
        for _, r in df_res.sort_values("Faltas", ascending=False).iterrows():
            dias_falto = r.get("Días que faltó", "—")
            data.append([
                Paragraph(str(r["Empleado"]), st_cell), r["Días prog."], r["Asistidos"], r["Faltas"],
                r["Justificadas"], r["Retardos"], r.get("Min. retardo",0),
                r.get("Omisiones",0), r["% Asistencia"],
                Paragraph(str(dias_falto), st_cell)
            ])
        t = Table(data, colWidths=[5.2*cm,1.3*cm,1.2*cm,1.2*cm,1.3*cm,1.5*cm,1.5*cm,1.2*cm,1.2*cm,4*cm], repeatRows=1)
        estilo = [
            ("BACKGROUND",(0,0),(-1,0), AZUL),
            ("TEXTCOLOR",(0,0),(-1,0), colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),7),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("GRID",(0,0),(-1,-1),0.4, colors.grey),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F4F6F9")]),
        ]
        for i, (_, r) in enumerate(df_res.sort_values("Faltas", ascending=False).iterrows(), start=1):
            pct = int(str(r["% Asistencia"]).replace("%","") or 0)
            if r["Faltas"] >= 10 or pct < 75:
                estilo.append(("BACKGROUND",(0,i),(-1,i), colors.HexColor("#F8D7DA")))
            elif r["Faltas"] >= 3 or pct < 90:
                estilo.append(("BACKGROUND",(0,i),(-1,i), colors.HexColor("#FFF3CD")))
        t.setStyle(TableStyle(estilo))
        elems.append(t)

        # Listado de días inhábiles / no laborables del período
        if fests_periodo:
            elems.append(Spacer(1, 0.5*cm))
            MESES = {1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}
            st_fh = ParagraphStyle("fh", parent=styles["Normal"], fontSize=8, fontName="Helvetica-Bold", textColor=AZUL)
            st_fl = ParagraphStyle("fl", parent=styles["Normal"], fontSize=8)
            elems.append(Paragraph("DÍAS INHÁBILES / NO LABORABLES EN EL PERÍODO", st_fh))
            for fday, desc in fests_periodo:
                elems.append(Paragraph(f"• {fday.day} de {MESES[fday.month]}: {desc}", st_fl))

        # Firma de Vo.Bo.
        elems.append(Spacer(1, 1.2*cm))
        st_fir = ParagraphStyle("fir", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER)
        firma = Paragraph(
            "_______________________________________<br/>"
            f"<b>Vo.Bo.  {DIRECTORA_NOMBRE}</b><br/>"
            f"{DIRECTORA_CARGO}", st_fir)
        ft = Table([[firma]], colWidths=[24*cm])
        ft.setStyle(TableStyle([("ALIGN",(0,0),(-1,-1),"CENTER")]))
        elems.append(ft)

        # ── Leyenda de confidencialidad (debajo de la firma) ──
        elems.append(Spacer(1, 0.5*cm))
        st_ley = ParagraphStyle("ley", parent=styles["Normal"], fontSize=7,
                                alignment=TA_CENTER, textColor=colors.HexColor("#555555"),
                                fontName="Helvetica-Oblique")
        elems.append(Paragraph(
            "Este documento es para uso exclusivo de la Dirección de Formación Continua "
            "y contiene información confidencial de carácter institucional. "
            "Su emisión tiene fines de control interno y respaldo administrativo.", st_ley))

        # Pie de página con numeración "Página X de Y" e identificación del reporte
        pie_txt = (f"Generado el {_emision_pdf.strftime('%d/%m/%Y %H:%M')} · "
                   f"Recursos Humanos — DFC · SEJ")

        from reportlab.pdfgen import canvas as _canvas

        class NumberedCanvas(_canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_states = []
            def showPage(self):
                self._saved_states.append(dict(self.__dict__))
                self._startPage()
            def save(self):
                total = len(self._saved_states)
                for state in self._saved_states:
                    self.__dict__.update(state)
                    self._draw_pie(total)
                    super().showPage()
                super().save()
            def _draw_pie(self, total):
                self.saveState()
                self.setFont("Helvetica", 7)
                self.setFillColor(colors.grey)
                w, _h = landscape(_letter)
                self.drawString(0.8*cm, 0.5*cm, pie_txt)
                self.drawRightString(w - 0.8*cm, 0.5*cm, f"Página {self._pageNumber} de {total}")
                self.restoreState()

        doc.build(elems, canvasmaker=NumberedCanvas)
        return buf.getvalue()

    # ── UI ────────────────────────────────────────────────────────────────
    st.markdown("### 🕐 Reloj Checador")
    col_up, col_tol = st.columns([3,1])
    with col_up:
        uploaded_xls = st.file_uploader("StandardReport.xls del checador", type=["xls"])
    with col_tol:
        umbral_r = st.number_input("Tolerancia retardo (min)", min_value=1, max_value=30, value=10)

    col_cd, col_cm = st.columns(2)
    with col_cd:
        comp_d = st.checkbox("Compensar retardos con salida tardía (mismo día)", value=False,
                             help="Si llegó X min tarde pero salió al menos X min después de su hora, "
                                  "el retardo no se cuenta (queda visible como 'Compensado' en el detalle).")
    with col_cm:
        comp_m = st.checkbox("Compensar retardos si cumple sus horas del período", value=False,
                             help="Si sus horas trabajadas del período igualan o superan las programadas, "
                                  "sus retardos no se cuentan (quedan visibles como compensados).")

    # ── Días justificados por evento institucional ──────────────────────
    with st.expander("📋 Agregar días justificados por evento institucional"):
        st.caption("Agrega fechas donde todos los empleados tienen autorizado faltar o salir temprano por evento oficial (ej: Juegos Magisteriales, Comida del Maestro).")
        col_f, col_m = st.columns([1,2])
        with col_f:
            fecha_evento = st.date_input("Fecha del evento", value=date.today(), key="fecha_evento_inst")
        with col_m:
            motivo_evento = st.text_input("Motivo / Nombre del evento", placeholder="Ej: Juegos Magisteriales", key="motivo_evento_inst")
        if st.button("➕ Agregar fecha justificada", key="btn_add_evento"):
            if motivo_evento:
                if "eventos_institucionales" not in st.session_state:
                    st.session_state["eventos_institucionales"] = []
                st.session_state["eventos_institucionales"].append({"fecha": fecha_evento, "motivo": motivo_evento})
                st.success(f"✅ {fecha_evento.strftime('%d/%m/%Y')} — {motivo_evento} agregado.")
            else:
                st.error("Escribe el motivo del evento.")

        if st.session_state.get("eventos_institucionales"):
            st.markdown("**Fechas justificadas para este reporte:**")
            for j, ev in enumerate(st.session_state["eventos_institucionales"]):
                c1, c2, c3 = st.columns([1,2,1])
                c1.write(ev["fecha"].strftime("%d/%m/%Y"))
                c2.write(ev["motivo"])
                if c3.button("🗑️", key=f"del_ev_{j}"):
                    st.session_state["eventos_institucionales"].pop(j)
                    st.rerun()

    if uploaded_xls:
        # Pre-leer período
        try:
            import xlrd as _xlrd
            _wb = _xlrd.open_workbook(file_contents=uploaded_xls.getvalue())
            _sh = _wb.sheet_by_name("Reporte de Asistencia")
            _periodo = ""
            for _i in range(5):
                for _j in range(_sh.ncols):
                    _v = str(_sh.cell_value(_i,_j)).strip()
                    if "~" in _v and "-" in _v:
                        _periodo = _v; break
            _fi_str, _ff_str = [x.strip() for x in _periodo.split("~")]
        except Exception as e:
            st.error(f"No se pudo leer el archivo: {e}")
            return

        with st.spinner("Cargando datos..."):
            empleados_ch  = load_emp_ch()
            festivos_ch   = load_fest_ch()
            historial_df  = load_hist_ch()
            economicos    = load_economicos_ch(_fi_str, _ff_str)
            incapacidades = load_incapacidades_ch(_fi_str, _ff_str)
            incidencias_p = load_incidencias_ch(_fi_str, _ff_str)

            fi = datetime.strptime(_fi_str, "%Y-%m-%d").date()
            ff = datetime.strptime(_ff_str, "%Y-%m-%d").date()
            eventos_inst = st.session_state.get("eventos_institucionales", [])
            just_rfc = build_justificantes_rfc(economicos, incapacidades, incidencias_p, fi, ff, eventos_inst)

            # ── Justificaciones de Dirección (Vo.Bo.): se aplican ANTES de
            # generar, para que Excel y PDF salgan ya corregidos ──
            try:
                _jd = cargar_justif_direccion()
                if not _jd.empty:
                    for _, _rj in _jd.iterrows():
                        _rfc_j = str(_rj.get("RFC", "")).strip().upper()
                        _f_j = pd.to_datetime(str(_rj.get("FECHA", "")), errors="coerce")
                        if _rfc_j and not pd.isna(_f_j) and fi <= _f_j.date() <= ff:
                            _mot_j = str(_rj.get("MOTIVO", "")).strip()
                            just_rfc.setdefault(_rfc_j, {})[_f_j.date()] = (
                                f"Justificada por Dirección|motivo:{_mot_j}" if _mot_j
                                else "Justificada por Dirección")
            except Exception as _e_jd:
                st.warning(f"No se pudieron leer las justificaciones de Dirección: {_e_jd}")

        try:
            df_res, df_ret, df_omis, periodo, fecha_ini, fecha_fin = parse_report_ch(
                uploaded_xls.getvalue(), empleados_ch, festivos_ch,
                historial_df, just_rfc, umbral_r, comp_d, comp_m)

            mes_nombre = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
                          "Agosto","Septiembre","Octubre","Noviembre","Diciembre"][fecha_ini.month-1]

            st.success(f"📅 {periodo} — {mes_nombre} {fecha_ini.year} · {VERSION_REPORTE}")

            c1,c2,c3,c4,c5 = st.columns(5)
            c1.metric("Empleados activos", len(df_res))
            c2.metric("Asistencia prom.", df_res["% Asistencia"].apply(lambda x: int(str(x).replace("%","") or 0)).mean().__round__(1).__str__() + "%")
            c3.metric("Total faltas",   int(df_res["Faltas"].sum()))
            c4.metric("Total retardos", int(df_res["Retardos"].sum()))
            c5.metric("Justificadas",   int(df_res["Justificadas"].sum()))

            st.dataframe(
                df_res.sort_values("Faltas", ascending=False),
                use_container_width=True, hide_index=True
            )

            if not df_ret.empty:
                with st.expander(f"📋 Detalle de retardos ({len(df_ret)} registros)"):
                    st.dataframe(df_ret, use_container_width=True, hide_index=True)

            if df_omis is not None and not df_omis.empty:
                st.warning(f"⚠️ {len(df_omis)} entrada(s)/salida(s) omitida(s) — el empleado asistió pero olvidó checar. No son retardos; revísalas.")
                with st.expander(f"📋 Checadas omitidas ({len(df_omis)} registros)"):
                    st.dataframe(df_omis, use_container_width=True, hide_index=True)

            # ── Selector de avisos de faltas ────────────────────────────
            con_faltas = df_res[df_res["Faltas"] > 0].copy()
            # ── Justificaciones de Dirección (Vo.Bo.) ─────────────────
            st.markdown("### ✍️ Justificar faltas (Vo.Bo. de Dirección)")
            st.caption("Marca las faltas que la Dirección justifica. Se guardan en el Sheet y "
                       "al instante el reporte se regenera con ellas aplicadas — el Excel y el "
                       "PDF salen ya corregidos, sin editar nada a mano.")
            _con_faltas_j = df_res[df_res["Faltas"] > 0]
            if _con_faltas_j.empty:
                st.info("No hay faltas por justificar en este período.")
            else:
                _emp_sel = st.selectbox("Empleado con falta(s)",
                                        _con_faltas_j["Empleado"].tolist(), key="jd_emp")
                _row_sel = _con_faltas_j[_con_faltas_j["Empleado"] == _emp_sel].iloc[0]
                _rfc_sel = str(_row_sel.get("RFC", "")).upper().strip()
                _fechas_raw = [f.strip() for f in str(_row_sel["Días que faltó"]).split(",")
                               if f.strip() and str(_row_sel["Días que faltó"]) != "—"]
                _fechas_sel = st.multiselect("Fecha(s) de falta a justificar (dd/mm)",
                                             _fechas_raw, key="jd_fechas")
                _motivo_jd = st.text_input("Motivo de la justificación (opcional)",
                                           key="jd_motivo",
                                           placeholder="Ej. Comisión verbal de la Dirección")
                if st.button("✅ Guardar justificación(es)", key="jd_guardar", type="primary"):
                    if not _fechas_sel:
                        st.error("Selecciona al menos una fecha.")
                    else:
                        _nuevas_jd = []
                        for _f in _fechas_sel:
                            try:
                                _d, _m = _f.split("/")
                                _fecha_iso = f"{fecha_ini.year}-{int(_m):02d}-{int(_d):02d}"
                                _nuevas_jd.append({"RFC": _rfc_sel, "NOMBRE": _emp_sel,
                                                   "FECHA": _fecha_iso, "MOTIVO": _motivo_jd})
                            except Exception:
                                st.error(f"Fecha con formato inesperado: {_f}")
                        if _nuevas_jd:
                            guardar_justif_direccion(_nuevas_jd)
                            st.success(f"{len(_nuevas_jd)} justificación(es) guardada(s). Regenerando reporte...")
                            st.rerun()
            # Justificaciones ya registradas en el período, con opción de deshacer
            try:
                _jd_ver = cargar_justif_direccion()
                if not _jd_ver.empty:
                    _jd_ver["_F"] = pd.to_datetime(_jd_ver["FECHA"].astype(str), errors="coerce")
                    _jd_per = _jd_ver[(_jd_ver["_F"].dt.date >= fi) & (_jd_ver["_F"].dt.date <= ff)]
                    if not _jd_per.empty:
                        with st.expander(f"📄 Justificaciones de Dirección registradas ({len(_jd_per)})"):
                            for _, _rj in _jd_per.iterrows():
                                c_txt, c_btn = st.columns([5, 1])
                                c_txt.write(f"**{_rj['NOMBRE']}** · {_rj['FECHA']} · {_rj['MOTIVO'] or 'sin motivo'}")
                                if c_btn.button("🗑️", key=f"jd_del_{_rj['RFC']}_{_rj['FECHA']}"):
                                    eliminar_justif_direccion(_rj["RFC"], _rj["FECHA"])
                                    st.rerun()
            except Exception:
                pass
            st.divider()

            st.markdown("### 📨 Avisos de faltas a empleados")
            if con_faltas.empty:
                st.success("No hay empleados con faltas en este período.")
                avisar_set = set()
            else:
                st.caption("Marca a quién mostrarle el aviso de falta en su portal y, si quieres, "
                           "escribe una observación (ej. 'Dirección autorizó', 'final juegos magisteriales'). "
                           "Las observaciones se guardan por fecha y se conservan entre reportes.")
                obs_hist = cargar_observaciones()
                avisar_set = set()
                obs_capturadas = []  # {RFC, NOMBRE, FECHA_FALTA, OBSERVACION}
                for _, r in con_faltas.iterrows():
                    rfc_e = str(r["RFC"]).upper()
                    dias_falta_txt = r["Días que faltó"] if r["Días que faltó"] != "—" else ""
                    st.markdown(f"**{r['Empleado']}** — {int(r['Faltas'])} falta(s): {dias_falta_txt}")
                    ca, cb = st.columns([1, 3])
                    with ca:
                        if st.checkbox("Avisar en portal", key=f"avisar_{rfc_e}"):
                            avisar_set.add(rfc_e)
                    with cb:
                        # Una observación por cada fecha de falta del empleado
                        fechas_falta = [f.strip() for f in dias_falta_txt.split(",") if f.strip()]
                        for ffalta in fechas_falta:
                            prev = ""
                            if not obs_hist.empty:
                                m = obs_hist[
                                    (obs_hist["RFC"].astype(str).str.upper() == rfc_e) &
                                    (obs_hist["FECHA_FALTA"].astype(str) == ffalta)
                                ]
                                if not m.empty:
                                    prev = str(m.iloc[0]["OBSERVACION"])
                            val = st.text_input(
                                f"Observación {ffalta}",
                                value=prev,
                                key=f"obs_{rfc_e}_{ffalta}"
                            )
                            obs_capturadas.append({
                                "RFC": rfc_e, "NOMBRE": r["Empleado"],
                                "FECHA_FALTA": ffalta, "OBSERVACION": val
                            })
                    st.divider()

            if st.button("💾 Guardar reporte, avisos y observaciones", type="primary"):
                filas = []
                for _, r in df_res.iterrows():
                    rfc_e = str(r.get("RFC","")).upper()
                    filas.append({
                        "RFC": rfc_e,
                        "NOMBRE": r["Empleado"],
                        "FALTAS": int(r["Faltas"]),
                        "JUSTIFICADAS": int(r["Justificadas"]),
                        "NO_JUSTIFICADAS": int(r["Faltas"]),
                        "RETARDOS": int(r["Retardos"]),
                        "DIAS_FALTA": r["Días que faltó"] if r["Días que faltó"] != "—" else "",
                        "AVISAR": "SI" if rfc_e in avisar_set else "NO",
                    })
                try:
                    guardar_asistencia_mes(filas, periodo)
                    guardar_observaciones(obs_capturadas if not con_faltas.empty else [])
                    st.success(f"Guardado. {len(avisar_set)} aviso(s) en portal · observaciones conservadas por fecha.")
                except Exception as e:
                    st.error(f"No se pudo guardar: {e}")

            try:
                festivos_df_desc = cargar_festivos()
            except Exception as e:
                festivos_df_desc = pd.DataFrame()
                st.error(f"⚠️ Error leyendo la tab 'festivos': {e}")
            fests_periodo = festivos_en_periodo(festivos_df_desc, fecha_ini, fecha_fin)
            if fests_periodo:
                st.caption(f"🗓️ Días inhábiles en el período incluidos en el reporte: "
                           + ", ".join(f"{d.day}/{d.month}" for d, _ in fests_periodo))
            else:
                st.caption("🗓️ No se detectaron días inhábiles dentro del período del reporte.")
                # Diagnóstico: mostrar qué se leyó realmente del Sheet
                with st.expander("🔍 Diagnóstico de festivos (por qué no se detectaron)"):
                    if festivos_df_desc is None or festivos_df_desc.empty:
                        st.write(f"La tab 'festivos' se leyó con **{0 if festivos_df_desc is None else len(festivos_df_desc)} filas**.")
                        st.write("Columnas:", list(festivos_df_desc.columns) if festivos_df_desc is not None else "N/A")
                        st.write("Voy a leer la tab SIN procesar para ver qué hay:")
                        try:
                            _cli = get_client()
                            _ws = _cli.open_by_key(st.secrets["sheet_checador_id"]).worksheet("festivos")
                            _valores = _ws.get_all_values()
                            st.write(f"get_all_values() devuelve **{len(_valores)} filas** (incluyendo encabezado):")
                            for _fila in _valores[:8]:
                                st.text(str(_fila))
                        except Exception as _e:
                            st.error(f"Error leyendo valores crudos: {_e}")
                    else:
                        st.write(f"Filas leídas: **{len(festivos_df_desc)}**")
                        st.write("Columnas:", list(festivos_df_desc.columns))
                        for _, _r in festivos_df_desc.head(8).iterrows():
                            st.text(f"INICIO={_r.get('FECHA_INICIO')!r} | FIN={_r.get('FECHA_FIN')!r} | {_r.get('DESCRIPCION')}")
                        st.write(f"Período del reporte: {fecha_ini} a {fecha_fin}")
            excel_bytes = generar_excel_reporte(df_res, df_ret, df_omis, cargar_observaciones(), periodo, fecha_ini, fecha_fin, umbral_r, fests_periodo)
            pdf_bytes   = generar_pdf_ejecutivo(df_res, periodo, fecha_ini, fecha_fin, fests_periodo)
            cdl1, cdl2 = st.columns(2)
            with cdl1:
                st.download_button(
                    "⬇️ Descargar reporte Excel (4 hojas)",
                    data=excel_bytes,
                    file_name=f"Asistencia_{mes_nombre}_{fecha_ini.year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with cdl2:
                st.download_button(
                    "📄 Descargar ejecutivo PDF (con logo y firma)",
                    data=pdf_bytes,
                    file_name=f"Asistencia_Ejecutivo_{mes_nombre}_{fecha_ini.year}.pdf",
                    mime="application/pdf"
                )

        except Exception as e:
            st.error(f"Error procesando el archivo: {e}")
    else:
        st.info("Sube el StandardReport.xls para procesar la asistencia del período.")